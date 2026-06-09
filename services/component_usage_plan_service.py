import math
import os
import re
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from services.errors import ServiceError


PLAN_HISTORY_COLUMN = 22
PLAN_HISTORY_COLUMN_LETTER = "V"
PLAN_LINE_COLUMN = 1
PLAN_WO_SUPPLY_COLUMN = 3
PLAN_DMS_PART_COLUMN = 7
PLAN_PCB_COLUMN = 9
TARGET_SHEET_NAME = "BOM"
EXCEL_EXTENSIONS = (".xls", ".xlsx", ".xlsm")
OPENPYXL_EXTENSIONS = (".xlsx", ".xlsm")
PLAN_EXTENSIONS = (".xls", ".xlsx", ".xlsm")

PCB_PATTERN = re.compile(r"(?<![A-Za-z0-9])(EAX[A-Za-z0-9]{8})(?:\s*\(([^()]*)\))?", re.IGNORECASE)
MODEL_PATTERN = re.compile(r"(?<![A-Za-z0-9])(?:EBU|EBT|EBR)[A-Za-z0-9-]{5,}", re.IGNORECASE)


@dataclass
class ComponentUsagePlanConfig:
    component_part_number: str
    plan_file: str
    source_folder: str


@dataclass
class PlanHistoryEntry:
    sheet_name: str
    row_number: int
    line: str
    wo_supply: str
    dms_part_number: str
    plan_pcb_part_number: str
    history_text: str
    main_part_number: str
    pcb_part_number: str
    revision: str


@dataclass
class PlanProgramTarget:
    main_part_number: str
    pcb_part_number: str
    revision: str
    plan_entries: list[PlanHistoryEntry] = field(default_factory=list)
    histories: list[str] = field(default_factory=list)


@dataclass
class ComponentUsageProgramMatch:
    main_part_number: str
    pcb_part_number: str
    source_folder: str
    source_file: str
    found_rows: list[int] = field(default_factory=list)


@dataclass
class ComponentUsagePlanRow:
    component_part_number: str
    line: str
    wo_supply: str
    dms_part_number: str
    pcb_part_number: str
    sheet_name: str
    row_number: int
    main_part_number: str


@dataclass
class ComponentUsagePlanResult:
    component_part_number: str
    plan_file: str
    source_folder: str
    rows: list[ComponentUsagePlanRow]
    matched_programs: list[ComponentUsageProgramMatch]
    plan_entries: list[PlanHistoryEntry]
    unique_target_count: int
    pcb_folder_count: int
    candidate_file_count: int
    read_file_count: int
    skipped_files: list[str] = field(default_factory=list)


@dataclass
class BomSearchData:
    found_rows: list[int]


def find_component_usage_on_excel_plan(config: ComponentUsagePlanConfig, progress_callback=None):
    _validate_config(config)

    component_text = _value_text(config.component_part_number)
    target_key = _match_key(component_text)
    plan_path = Path(config.plan_file)
    source_folder = Path(config.source_folder)

    _emit_progress(progress_callback, 0, "Reading Excel plan...")
    plan_entries = _read_plan_history_entries(plan_path)
    targets = _build_program_targets(plan_entries)
    skipped_files = []

    if not targets:
        _emit_progress(progress_callback, 100, "No valid history found in column V")
        return ComponentUsagePlanResult(
            component_part_number=component_text,
            plan_file=str(plan_path),
            source_folder=str(source_folder),
            rows=[],
            matched_programs=[],
            plan_entries=plan_entries,
            unique_target_count=0,
            pcb_folder_count=0,
            candidate_file_count=0,
            read_file_count=0,
            skipped_files=[],
        )

    _emit_progress(progress_callback, 8, f"Indexing PCB folders for {len(targets)} target(s)...")
    pcb_folder_map, folder_errors = _map_pcb_folders(
        source_folder,
        [target.pcb_part_number for target in targets],
    )
    skipped_files.extend(folder_errors)

    rows = []
    matched_programs = []
    folder_file_cache = {}
    candidate_file_count = 0
    read_file_count = 0
    seen_plan_rows = set()
    total_targets = len(targets)

    for index, target in enumerate(targets, start=1):
        pcb_display = format_pcb_part_number(target.pcb_part_number, target.revision)
        percent = 10 + int((index - 1) / total_targets * 88)
        _emit_progress(
            progress_callback,
            percent,
            f"Checking {index}/{total_targets}: {target.main_part_number} / {pcb_display}",
        )

        folder_key = _match_key(target.pcb_part_number)
        matched_folders = pcb_folder_map.get(folder_key, [])
        if not matched_folders:
            skipped_files.append(f"{target.main_part_number} / {pcb_display}: folder PCB tidak ditemukan")
            continue

        target_had_candidate = False
        target_found_component = False
        for pcb_folder in matched_folders:
            excel_files = _excel_files_under_folder_cached(pcb_folder, folder_file_cache)
            candidate_files = [
                file_path
                for file_path in excel_files
                if _filename_matches_main_part(file_path, target.main_part_number)
            ]

            if not candidate_files:
                skipped_files.append(
                    f"{target.main_part_number} / {pcb_display}: file Excel program tidak ditemukan di {_relative_path(pcb_folder, source_folder)}"
                )
                continue

            target_had_candidate = True
            for file_path in candidate_files:
                candidate_file_count += 1
                try:
                    bom_data = _read_bom_file(file_path, target_key)
                except Exception as exc:
                    skipped_files.append(
                        f"{target.main_part_number} / {pcb_display} / {file_path.name}: {_error_message(exc)}"
                    )
                    continue

                read_file_count += 1
                if not bom_data.found_rows:
                    continue

                matched_programs.append(
                    ComponentUsageProgramMatch(
                        main_part_number=target.main_part_number,
                        pcb_part_number=pcb_display,
                        source_folder=_relative_path(file_path.parent, source_folder),
                        source_file=file_path.name,
                        found_rows=bom_data.found_rows,
                    )
                )
                target_found_component = True
                break

            if target_found_component:
                break

        if target_found_component:
            for plan_entry in target.plan_entries:
                result_key = (plan_entry.sheet_name, plan_entry.row_number)
                if result_key in seen_plan_rows:
                    continue
                seen_plan_rows.add(result_key)
                rows.append(
                    ComponentUsagePlanRow(
                        component_part_number=component_text,
                        line=plan_entry.line,
                        wo_supply=plan_entry.wo_supply,
                        dms_part_number=plan_entry.dms_part_number,
                        pcb_part_number=pcb_display,
                        sheet_name=plan_entry.sheet_name,
                        row_number=plan_entry.row_number,
                        main_part_number=plan_entry.main_part_number,
                    )
                )

        if not target_had_candidate:
            continue

    final_message = f"Search complete: {len(rows)} result(s) found" if rows else "No result found"
    _emit_progress(progress_callback, 100, final_message)
    return ComponentUsagePlanResult(
        component_part_number=component_text,
        plan_file=str(plan_path),
        source_folder=str(source_folder),
        rows=rows,
        matched_programs=matched_programs,
        plan_entries=plan_entries,
        unique_target_count=len(targets),
        pcb_folder_count=sum(len(folders) for folders in pcb_folder_map.values()),
        candidate_file_count=candidate_file_count,
        read_file_count=read_file_count,
        skipped_files=skipped_files,
    )


def suggest_plan_export_name(component_part_number):
    component_text = _value_text(component_part_number)
    safe_component = re.sub(r"[^A-Za-z0-9_.-]+", "_", component_text).strip("_") or "Component_Usage"
    return f"{safe_component}_Plan_Usage_{datetime.now().strftime('%y%m%d')}.xlsx"


def export_component_usage_plan_result(result, output_path):
    if result is None:
        raise ServiceError("Belum ada hasil pencarian untuk diexport.", title="Data kosong")

    output = _normalize_output_path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    preview_sheet = workbook.active
    preview_sheet.title = "Preview Result"
    _write_preview_sheet(preview_sheet, result.rows)

    plan_sheet = workbook.create_sheet("Plan Targets")
    _write_plan_targets_sheet(plan_sheet, result.plan_entries)

    log_sheet = workbook.create_sheet("Scan Log")
    _write_log_sheet(log_sheet, result)

    workbook.save(output)
    return str(output)


def parse_history_program(
    value,
    sheet_name="",
    row_number=0,
    line="",
    wo_supply="",
    dms_part_number="",
    plan_pcb_part_number="",
):
    text = _value_text(value)
    if not text:
        return None

    pcb_match = PCB_PATTERN.search(text)
    if not pcb_match:
        return None

    before_pcb = text[: pcb_match.start()]
    model_match = MODEL_PATTERN.search(before_pcb)
    if not model_match:
        return None

    history_pcb = pcb_match.group(1).upper()
    plan_pcb = _extract_pcb_part_number(plan_pcb_part_number)
    if not plan_pcb or _match_key(plan_pcb) != _match_key(history_pcb):
        return None

    return PlanHistoryEntry(
        sheet_name=sheet_name,
        row_number=row_number,
        line=_value_text(line),
        wo_supply=_value_text(wo_supply),
        dms_part_number=_value_text(dms_part_number),
        plan_pcb_part_number=plan_pcb,
        history_text=text,
        main_part_number=model_match.group(0).upper(),
        pcb_part_number=history_pcb,
        revision=_value_text(pcb_match.group(2)),
    )


def format_pcb_part_number(pcb_part_number, revision):
    pcb_text = _value_text(pcb_part_number) or "-"
    revision_text = _value_text(revision)
    return f"{pcb_text}({revision_text})" if revision_text and pcb_text != "-" else pcb_text


def _read_plan_history_entries(plan_path):
    suffix = plan_path.suffix.lower()
    if suffix in OPENPYXL_EXTENSIONS:
        return _read_plan_history_entries_openpyxl(plan_path)
    if suffix == ".xls":
        return _read_plan_history_entries_pandas(plan_path)
    raise ServiceError(
        f"Format Excel plan belum didukung: {plan_path.suffix}",
        title="Format Excel tidak valid",
    )


def _read_plan_history_entries_openpyxl(plan_path):
    workbook = load_workbook(plan_path, read_only=True, data_only=True, keep_links=False)
    try:
        entries = []
        for worksheet in workbook.worksheets:
            if worksheet.max_column and worksheet.max_column < PLAN_HISTORY_COLUMN:
                continue
            for row_number, row_values in enumerate(
                worksheet.iter_rows(
                    min_col=1,
                    max_col=PLAN_HISTORY_COLUMN,
                    values_only=True,
                ),
                start=1,
            ):
                entry = parse_history_program(
                    row_values[PLAN_HISTORY_COLUMN - 1] if len(row_values) >= PLAN_HISTORY_COLUMN else None,
                    worksheet.title,
                    row_number,
                    line=row_values[PLAN_LINE_COLUMN - 1] if len(row_values) >= PLAN_LINE_COLUMN else "",
                    wo_supply=row_values[PLAN_WO_SUPPLY_COLUMN - 1] if len(row_values) >= PLAN_WO_SUPPLY_COLUMN else "",
                    dms_part_number=row_values[PLAN_DMS_PART_COLUMN - 1] if len(row_values) >= PLAN_DMS_PART_COLUMN else "",
                    plan_pcb_part_number=row_values[PLAN_PCB_COLUMN - 1] if len(row_values) >= PLAN_PCB_COLUMN else "",
                )
                if entry:
                    entries.append(entry)
        return entries
    finally:
        workbook.close()


def _read_plan_history_entries_pandas(plan_path):
    try:
        import pandas as pd
    except ImportError as exc:
        raise ServiceError(
            "File .xls membutuhkan pandas dan xlrd. Jalankan install dependency dari requirements.txt.",
            title="Dependency belum lengkap",
        ) from exc

    try:
        excel = pd.ExcelFile(plan_path)
    except ImportError as exc:
        raise ServiceError(
            "File .xls membutuhkan library xlrd. Jalankan install dependency dari requirements.txt.",
            title="Dependency belum lengkap",
        ) from exc

    try:
        entries = []
        for sheet_name in excel.sheet_names:
            try:
                dataframe = pd.read_excel(
                    excel,
                    sheet_name=sheet_name,
                    header=None,
                    usecols=[
                        PLAN_LINE_COLUMN - 1,
                        PLAN_WO_SUPPLY_COLUMN - 1,
                        PLAN_DMS_PART_COLUMN - 1,
                        PLAN_PCB_COLUMN - 1,
                        PLAN_HISTORY_COLUMN - 1,
                    ],
                    dtype=object,
                    na_filter=False,
                )
            except ValueError:
                continue

            for dataframe_index, row in enumerate(dataframe.itertuples(index=False, name=None), start=1):
                entry = parse_history_program(
                    row[4],
                    sheet_name,
                    dataframe_index,
                    line=row[0],
                    wo_supply=row[1],
                    dms_part_number=row[2],
                    plan_pcb_part_number=row[3],
                )
                if entry:
                    entries.append(entry)
        return entries
    finally:
        excel.close()


def _build_program_targets(plan_entries):
    targets = OrderedDict()
    for entry in plan_entries:
        key = (
            _match_key(entry.main_part_number),
            _match_key(entry.pcb_part_number),
            _match_key(entry.revision),
        )
        if key not in targets:
            targets[key] = PlanProgramTarget(
                main_part_number=entry.main_part_number,
                pcb_part_number=entry.pcb_part_number,
                revision=entry.revision,
            )
        target = targets[key]
        target.plan_entries.append(entry)
        if entry.history_text not in target.histories:
            target.histories.append(entry.history_text)
    return list(targets.values())


def _map_pcb_folders(main_folder, pcb_numbers):
    targets = OrderedDict((_match_key(pcb), pcb) for pcb in pcb_numbers if _value_text(pcb))
    folder_map = {key: [] for key in targets}
    errors = []
    seen_paths = set()

    def add_if_match(folder):
        name_upper = folder.name.upper()
        matched_keys = set()
        for match in PCB_PATTERN.finditer(folder.name):
            pcb_key = _match_key(match.group(1))
            if pcb_key in targets:
                matched_keys.add(pcb_key)

        if not matched_keys:
            matched_keys.update(key for key in targets if key in name_upper)

        for key in matched_keys:
            path_key = str(folder.resolve()).upper()
            unique_key = (key, path_key)
            if unique_key in seen_paths:
                continue
            seen_paths.add(unique_key)
            folder_map[key].append(folder)

    def on_error(error):
        errors.append(f"{getattr(error, 'filename', '')}: {getattr(error, 'strerror', error)}")

    add_if_match(main_folder)
    for current_folder, dir_names, _file_names in os.walk(main_folder, onerror=on_error):
        dir_names.sort(key=lambda name: name.upper())
        for dir_name in dir_names:
            add_if_match(Path(current_folder) / dir_name)

    return {key: folders for key, folders in folder_map.items() if folders}, errors


def _excel_files_under_folder_cached(folder, cache):
    folder_key = str(Path(folder).resolve())
    if folder_key not in cache:
        cache[folder_key] = _excel_files_under_folder(Path(folder))
    return cache[folder_key]


def _excel_files_under_folder(folder):
    files = []

    def on_error(_error):
        return None

    for current_folder, dir_names, file_names in os.walk(folder, onerror=on_error):
        dir_names.sort(key=lambda name: name.upper())
        for file_name in sorted(file_names, key=lambda name: name.upper()):
            if file_name.startswith("~$"):
                continue
            path = Path(current_folder) / file_name
            if path.suffix.lower() in EXCEL_EXTENSIONS:
                files.append(path)
    return files


def _filename_matches_main_part(file_path, main_part_number):
    return _match_key(main_part_number) in _match_key(Path(file_path).stem)


def _read_bom_file(file_path, target_key):
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix in OPENPYXL_EXTENSIONS:
        return _read_bom_openpyxl(path, target_key)
    if suffix == ".xls":
        return _read_bom_pandas(path, target_key)
    raise ServiceError(f"Format file tidak didukung: {path.suffix}", title="Format tidak valid")


def _read_bom_openpyxl(path, target_key):
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        sheet_name = _find_sheet_case_insensitive(workbook.sheetnames, TARGET_SHEET_NAME)
        if sheet_name is None:
            raise ServiceError('Sheet "BOM" tidak ditemukan.', title="Format Excel tidak valid")

        found_rows = OrderedDict()
        for row_number, row_values in enumerate(workbook[sheet_name].iter_rows(values_only=True), start=1):
            for value in row_values:
                text = _value_text(value)
                if text and _cell_matches(text, target_key):
                    found_rows[row_number] = row_number
                    break
        return BomSearchData(list(found_rows.keys()))
    finally:
        workbook.close()


def _read_bom_pandas(path, target_key):
    try:
        import pandas as pd
    except ImportError as exc:
        raise ServiceError(
            "File .xls membutuhkan pandas dan xlrd. Jalankan install dependency dari requirements.txt.",
            title="Dependency belum lengkap",
        ) from exc

    try:
        excel = pd.ExcelFile(path)
    except ImportError as exc:
        raise ServiceError(
            "File .xls membutuhkan library xlrd. Jalankan install dependency dari requirements.txt.",
            title="Dependency belum lengkap",
        ) from exc

    try:
        sheet_name = _find_sheet_case_insensitive(excel.sheet_names, TARGET_SHEET_NAME)
        if sheet_name is None:
            raise ServiceError('Sheet "BOM" tidak ditemukan.', title="Format Excel tidak valid")

        dataframe = pd.read_excel(excel, sheet_name=sheet_name, header=None, dtype=object, na_filter=False)
        found_rows = OrderedDict()
        for dataframe_index, row in dataframe.iterrows():
            for value in row.tolist():
                text = _value_text(value)
                if text and _cell_matches(text, target_key):
                    row_number = int(dataframe_index) + 1
                    found_rows[row_number] = row_number
                    break

        return BomSearchData(list(found_rows.keys()))
    finally:
        excel.close()


def _write_preview_sheet(worksheet, rows):
    worksheet.append(
        [
            "No",
            "LINE",
            "WO SUPPLY",
            "DMS P/N",
            "PCB",
        ]
    )
    for index, row in enumerate(rows, start=1):
        worksheet.append(
            [
                index,
                row.line,
                row.wo_supply,
                row.dms_part_number,
                row.pcb_part_number,
            ]
        )
    _style_sheet(worksheet)


def _write_plan_targets_sheet(worksheet, plan_entries):
    worksheet.append(
        [
            "No",
            "Sheet",
            "Row",
            "LINE",
            "WO SUPPLY",
            "DMS P/N",
            "PCB Column I",
            "Main Part Number",
            "History PCB",
            "Revision",
            "History Program",
        ]
    )
    for index, entry in enumerate(plan_entries, start=1):
        worksheet.append(
            [
                index,
                entry.sheet_name,
                entry.row_number,
                entry.line,
                entry.wo_supply,
                entry.dms_part_number,
                entry.plan_pcb_part_number,
                entry.main_part_number,
                entry.pcb_part_number,
                entry.revision,
                entry.history_text,
            ]
        )
    _style_sheet(worksheet)
    worksheet.column_dimensions["K"].width = 80


def _write_log_sheet(worksheet, result):
    worksheet.append(["Field", "Value"])
    worksheet.append(["Component Part Number", result.component_part_number])
    worksheet.append(["Plan File", result.plan_file])
    worksheet.append(["Folder Induk PCB", result.source_folder])
    worksheet.append(["History Rows Parsed", len(result.plan_entries)])
    worksheet.append(["Unique Program Targets", result.unique_target_count])
    worksheet.append(["PCB Folders Found", result.pcb_folder_count])
    worksheet.append(["Candidate Program Files", result.candidate_file_count])
    worksheet.append(["Program Files Read", result.read_file_count])
    worksheet.append(["Preview Result Count", len(result.rows)])
    worksheet.append(["Skipped/Error Files", len(result.skipped_files)])
    worksheet.append([])

    worksheet.append(["Matched Program Files"])
    worksheet.append(["No", "Main Part Number", "PCB Part Number", "PCB Folder", "Program Excel", "BOM Row"])
    for index, match in enumerate(result.matched_programs, start=1):
        worksheet.append(
            [
                index,
                match.main_part_number,
                match.pcb_part_number,
                match.source_folder,
                match.source_file,
                _join_text(match.found_rows),
            ]
        )
    worksheet.append([])

    worksheet.append(["Skipped/Error"])
    worksheet.append(["No", "Message"])
    for index, skipped in enumerate(result.skipped_files, start=1):
        worksheet.append([index, skipped])

    _style_sheet(worksheet)
    worksheet.column_dimensions["B"].width = 80


def _style_sheet(worksheet):
    header_font = Font(name="Calibri", size=11, bold=True)
    body_font = Font(name="Calibri", size=11)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = header_font if cell.row == 1 else body_font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for row in worksheet.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        if cell.value in {"Matched Program Files", "Skipped/Error"}:
            cell.font = header_font

    for col_index in range(1, worksheet.max_column + 1):
        letter = get_column_letter(col_index)
        max_length = 10
        for row_index in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row_index, col_index).value
            if value not in (None, ""):
                max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[letter].width = min(max_length + 2, 60)

    worksheet.freeze_panes = "A2"


def _find_sheet_case_insensitive(sheet_names, target_name):
    target = target_name.lower()
    for sheet_name in sheet_names:
        if sheet_name.lower() == target:
            return sheet_name
    return None


def _cell_matches(value, target_key):
    text = _match_key(value)
    if not text:
        return False
    if text == target_key:
        return True
    return re.search(rf"(?<![A-Z0-9_.-]){re.escape(target_key)}(?![A-Z0-9_.-])", text) is not None


def _extract_pcb_part_number(value):
    match = PCB_PATTERN.search(_value_text(value))
    return match.group(1).upper() if match else ""


def _match_key(value):
    text = _value_text(value).upper()
    return re.sub(r"\s+", " ", text).strip()


def _value_text(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
    text = str(value)
    text = text.replace("\r", " ").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def _join_text(values):
    return ", ".join(str(value) for value in values) if values else ""


def _relative_path(path, root):
    try:
        return str(Path(path).resolve().relative_to(Path(root).resolve()))
    except ValueError:
        return str(path)


def _validate_config(config):
    if not _value_text(config.component_part_number):
        raise ServiceError("Component Part Number belum diisi.", title="Input belum lengkap")
    if not config.plan_file:
        raise ServiceError("File Excel plan belum dipilih.", title="Input belum lengkap")
    plan_path = Path(config.plan_file)
    if not plan_path.is_file():
        raise ServiceError(f"File Excel plan tidak ditemukan:\n{config.plan_file}", title="File tidak ditemukan")
    if plan_path.suffix.lower() not in PLAN_EXTENSIONS:
        raise ServiceError(f"Format Excel plan belum didukung: {plan_path.suffix}", title="Format Excel tidak valid")
    if not config.source_folder:
        raise ServiceError("Folder Induk PCB belum dipilih.", title="Input belum lengkap")
    if not Path(config.source_folder).is_dir():
        raise ServiceError(f"Folder Induk PCB tidak ditemukan:\n{config.source_folder}", title="Folder tidak ditemukan")


def _normalize_output_path(path):
    output_path = Path(path)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")
    return output_path


def _emit_progress(progress_callback, percent, message):
    if progress_callback:
        progress_callback(percent, message)


def _error_message(exc):
    return getattr(exc, "message", str(exc))
