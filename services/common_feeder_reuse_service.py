import math
import re
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services import feeder_mapping_service
from services.component_usage_finder_service import (
    format_pcb_part_number,
    parse_model_part_numbers,
    parse_pcb_part_number,
)
from services.errors import ServiceError
from services.io_helpers import open_pandas_excel_file, scan_recursive_files


EXCEL_EXTENSIONS = (".xls", ".xlsx", ".xlsm")
OPENPYXL_EXTENSIONS = (".xlsx", ".xlsm")
TARGET_SHEET_NAME = "BOM"

STATUS_SAFE = "SAFE"
STATUS_CONFLICT = "CONFLICT"
STATUS_CHECK = "CHECK"
STATUS_OPTIONS = ["SHOW ALL", STATUS_SAFE, STATUS_CONFLICT, STATUS_CHECK]

COMPATIBILITY_COLUMNS = [
    ("status", "Status"),
    ("candidate_part_number", "Candidate P/N"),
    ("main_part_number", "Main Feeder P/N"),
    ("location_code", "Location Code"),
    ("table", "Table"),
    ("slot", "Slot"),
    ("position", "Position"),
    ("candidate_usage_count", "Candidate Used In"),
    ("main_usage_count", "Main Used In"),
    ("conflict_count", "Conflict Count"),
    ("conflict_programs", "Conflict PCB / Model"),
    ("candidate_programs", "Candidate Usage"),
    ("main_programs", "Main Usage"),
]


@dataclass
class CommonFeederReuseConfig:
    source_folder: str
    feeder_source_path: str
    candidate_part_numbers: str = ""


@dataclass
class ProgramInfo:
    key: str
    display_name: str
    model_part_numbers: list[str]
    pcb_part_number: str
    revision: str
    source_folder: str
    source_file: str


@dataclass
class ComponentUsage:
    part_number: str
    programs: OrderedDict = field(default_factory=OrderedDict)

    @property
    def program_count(self):
        return len(self.programs)

    @property
    def total_rows(self):
        return sum(self.programs.values())


@dataclass
class CommonFeederReuseResult:
    rows: list[dict]
    matrix_rows: list[dict]
    program_infos: list[ProgramInfo]
    feeder_records: list[dict]
    total_files: int
    read_files: int
    skipped_files: list[str]
    candidate_count: int
    component_count: int
    safe_count: int
    conflict_count: int
    check_count: int


def analyze_common_feeder_reuse(config: CommonFeederReuseConfig, progress_callback=None):
    _validate_config(config)

    _emit_progress(progress_callback, 0, "Loading fixed feeder data...")
    feeder_records = _load_feeder_records(config.feeder_source_path)
    if not feeder_records:
        raise ServiceError("Fixed feeder data kosong atau tidak punya Part Number.", title="Data kosong")

    usage_map, program_lookup, total_files, read_files, skipped_files = _scan_component_usage(
        config.source_folder,
        progress_callback,
    )

    if not usage_map:
        raise ServiceError('Tidak ada component P/N yang terbaca dari sheet "BOM".', title="Data kosong")

    candidate_map = parse_candidate_part_numbers(config.candidate_part_numbers)
    if not candidate_map:
        fixed_keys = {_part_key(record.get("part_number")) for record in feeder_records if _part_key(record.get("part_number"))}
        candidate_map = OrderedDict(
            (key, usage.part_number)
            for key, usage in sorted(usage_map.items(), key=lambda item: item[1].part_number.upper())
            if key not in fixed_keys
        )

    if not candidate_map:
        raise ServiceError("Candidate component kosong. Isi Candidate P/N atau cek folder BOM.", title="Data kosong")

    rows = []
    total_pairs = max(1, len(feeder_records) * len(candidate_map))
    pair_index = 0

    _emit_progress(progress_callback, 96, "Checking feeder compatibility...")
    for feeder_record in feeder_records:
        main_part = _part_text(feeder_record.get("part_number"))
        main_key = _part_key(main_part)
        if not main_key:
            continue

        main_usage = usage_map.get(main_key)
        main_program_keys = set(main_usage.programs.keys()) if main_usage else set()

        for candidate_key, candidate_part in candidate_map.items():
            pair_index += 1
            if pair_index % 1000 == 0:
                percent = 96 + min(3, int(pair_index / total_pairs * 3))
                _emit_progress(progress_callback, percent, f"Checking pair {pair_index}/{total_pairs}...")

            if candidate_key == main_key:
                continue

            candidate_usage = usage_map.get(candidate_key)
            candidate_program_keys = set(candidate_usage.programs.keys()) if candidate_usage else set()
            conflict_keys = sorted(
                main_program_keys & candidate_program_keys,
                key=lambda key: program_lookup.get(key).display_name if key in program_lookup else key,
            )

            if conflict_keys:
                status = STATUS_CONFLICT
            elif not main_program_keys or not candidate_program_keys:
                status = STATUS_CHECK
            else:
                status = STATUS_SAFE

            rows.append(
                {
                    "status": status,
                    "candidate_part_number": candidate_part,
                    "main_part_number": main_part,
                    "location_code": _part_text(feeder_record.get("location_code")),
                    "table": _part_text(feeder_record.get("table")),
                    "slot": _part_text(feeder_record.get("slot")),
                    "position": _part_text(feeder_record.get("position")),
                    "candidate_usage_count": len(candidate_program_keys),
                    "main_usage_count": len(main_program_keys),
                    "conflict_count": len(conflict_keys),
                    "conflict_programs": _format_programs(conflict_keys, program_lookup),
                    "candidate_programs": _format_programs(candidate_program_keys, program_lookup),
                    "main_programs": _format_programs(main_program_keys, program_lookup),
                }
            )

    rows.sort(key=_compatibility_sort_key)
    matrix_rows = _build_matrix_rows(usage_map, program_lookup)
    safe_count = sum(1 for row in rows if row["status"] == STATUS_SAFE)
    conflict_count = sum(1 for row in rows if row["status"] == STATUS_CONFLICT)
    check_count = sum(1 for row in rows if row["status"] == STATUS_CHECK)

    _emit_progress(progress_callback, 100, "Analysis complete")
    return CommonFeederReuseResult(
        rows=rows,
        matrix_rows=matrix_rows,
        program_infos=list(program_lookup.values()),
        feeder_records=feeder_records,
        total_files=total_files,
        read_files=read_files,
        skipped_files=skipped_files,
        candidate_count=len(candidate_map),
        component_count=len(usage_map),
        safe_count=safe_count,
        conflict_count=conflict_count,
        check_count=check_count,
    )


def parse_candidate_part_numbers(text):
    candidates = OrderedDict()
    for token in re.split(r"[\r\n,;|\t]+", str(text or "")):
        clean = _part_text(token)
        key = _part_key(clean)
        if key and key not in candidates:
            candidates[key] = clean
    return candidates


def suggest_export_name():
    return f"Common_Feeder_Reuse_{datetime.now().strftime('%y%m%d')}.xlsx"


def export_common_feeder_reuse_result(result, output_path):
    if result is None:
        raise ServiceError("Belum ada hasil analisa untuk diexport.", title="Data kosong")

    output = _normalize_output_path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    compatibility_sheet = workbook.active
    compatibility_sheet.title = "Compatibility"
    _write_records_sheet(compatibility_sheet, result.rows, COMPATIBILITY_COLUMNS)

    conflict_sheet = workbook.create_sheet("Do Not Pair List")
    conflict_rows = [row for row in result.rows if row.get("status") == STATUS_CONFLICT]
    _write_records_sheet(conflict_sheet, conflict_rows, COMPATIBILITY_COLUMNS)

    summary_sheet = workbook.create_sheet("Component Usage")
    _write_records_sheet(
        summary_sheet,
        result.matrix_rows,
        [
            ("component_part_number", "Component P/N"),
            ("used_model_count", "Used In"),
            ("total_bom_rows", "BOM Rows"),
            ("programs", "PCB / Model Usage"),
        ],
    )

    matrix_sheet = workbook.create_sheet("Usage Matrix")
    _write_usage_matrix(matrix_sheet, result.matrix_rows, result.program_infos)

    log_sheet = workbook.create_sheet("Scan Log")
    _write_scan_log(log_sheet, result)

    workbook.save(output)
    return str(output)


def _scan_component_usage(source_folder, progress_callback=None):
    folder = Path(source_folder)
    excel_files, walk_errors = _find_excel_files(folder)
    usage_map = OrderedDict()
    program_lookup = OrderedDict()
    skipped_files = list(walk_errors)
    read_files = 0
    total_files = len(excel_files)

    for index, file_path in enumerate(excel_files, start=1):
        percent = max(1, min(95, int((index - 1) / max(1, total_files) * 95)))
        _emit_progress(progress_callback, percent, f"Reading BOM {index}/{total_files}: {file_path.name}")

        try:
            parts = _read_bom_parts(file_path)
        except Exception as exc:
            skipped_files.append(f"{file_path}: {_error_message(exc)}")
            continue

        read_files += 1
        if not parts:
            continue

        program_info = _program_info(file_path)
        program_lookup[program_info.key] = program_info
        for part in parts:
            key = _part_key(part)
            if not key:
                continue
            if key not in usage_map:
                usage_map[key] = ComponentUsage(part_number=part)
            usage_map[key].programs[program_info.key] = usage_map[key].programs.get(program_info.key, 0) + 1

    usage_map = OrderedDict(sorted(usage_map.items(), key=lambda item: item[1].part_number.upper()))
    return usage_map, program_lookup, total_files, read_files, skipped_files


def _read_bom_parts(path):
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in OPENPYXL_EXTENSIONS:
        return _read_bom_parts_openpyxl(path)
    if suffix == ".xls":
        return _read_bom_parts_pandas(path)
    raise ServiceError(f"Format file tidak didukung: {path.suffix}", title="Format tidak valid")


def _read_bom_parts_openpyxl(path):
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = _find_sheet_case_insensitive(workbook.sheetnames, TARGET_SHEET_NAME)
        if worksheet is None:
            raise ServiceError('Sheet "BOM" tidak ditemukan.', title="Format Excel tidak valid")

        values = []
        for row in workbook[worksheet].iter_rows(min_col=3, max_col=6, values_only=True):
            part_value = row[0] if row else None
            marker_value = row[3] if len(row) > 3 else None
            if _is_na_marker(marker_value):
                continue
            values.append(part_value)
        return _clean_part_values(values)
    finally:
        workbook.close()


def _read_bom_parts_pandas(path):
    try:
        import pandas as pd
    except ImportError as exc:
        raise ServiceError(
            "File .xls membutuhkan pandas dan xlrd. Jalankan install dependency dari requirements.txt.",
            title="Dependency belum lengkap",
        ) from exc

    excel = open_pandas_excel_file(pd, path)
    try:
        sheet_name = _find_sheet_case_insensitive(excel.sheet_names, TARGET_SHEET_NAME)
        if sheet_name is None:
            raise ServiceError('Sheet "BOM" tidak ditemukan.', title="Format Excel tidak valid")
        try:
            dataframe = pd.read_excel(
                excel,
                sheet_name=sheet_name,
                header=None,
                dtype=object,
                na_filter=False,
                usecols="C:F",
            )
            compact_columns = True
        except ValueError:
            dataframe = pd.read_excel(excel, sheet_name=sheet_name, header=None, dtype=object, na_filter=False)
            compact_columns = False

        values = []
        for _, row in dataframe.iterrows():
            if compact_columns:
                part_value = row.iloc[0] if len(row) > 0 else None
                marker_value = row.iloc[3] if len(row) > 3 else None
            else:
                part_value = row.iloc[2] if len(row) > 2 else None
                marker_value = row.iloc[5] if len(row) > 5 else None
            if _is_na_marker(marker_value):
                continue
            values.append(part_value)
        return _clean_part_values(values)
    finally:
        excel.close()


def _load_feeder_records(file_path):
    path = Path(file_path)
    if not path.is_file():
        raise ServiceError(f"Fixed feeder source tidak ditemukan:\n{path}", title="File tidak ditemukan")

    suffix = path.suffix.lower()
    if suffix == ".txt":
        result = feeder_mapping_service.load_feeder_mapping(str(path))
        return result.records
    if suffix in OPENPYXL_EXTENSIONS:
        return _read_feeder_records_openpyxl(path)
    if suffix == ".xls":
        return _read_feeder_records_pandas(path)

    raise ServiceError("Fixed feeder source harus berupa export NPM .txt atau feeder mapping Excel.", title="Format tidak valid")


def _read_feeder_records_openpyxl(path):
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        sheet_name = "Detailed Feeder Setup" if "Detailed Feeder Setup" in workbook.sheetnames else workbook.sheetnames[0]
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        return _parse_feeder_table_rows(rows)
    finally:
        workbook.close()


def _read_feeder_records_pandas(path):
    try:
        import pandas as pd
    except ImportError as exc:
        raise ServiceError(
            "File .xls membutuhkan pandas dan xlrd. Jalankan install dependency dari requirements.txt.",
            title="Dependency belum lengkap",
        ) from exc

    excel = open_pandas_excel_file(pd, path)
    try:
        sheet_name = "Detailed Feeder Setup" if "Detailed Feeder Setup" in excel.sheet_names else excel.sheet_names[0]
        dataframe = pd.read_excel(excel, sheet_name=sheet_name, header=None, dtype=object, na_filter=False)
        return _parse_feeder_table_rows(dataframe.values.tolist())
    finally:
        excel.close()


def _parse_feeder_table_rows(rows):
    header = None
    header_index = -1
    for index, row in enumerate(rows[:30]):
        normalized = [_normalize_header(value) for value in row]
        if "partnumber" in normalized:
            header = normalized
            header_index = index
            break

    if header is None:
        raise ServiceError('Kolom "Part Number" tidak ditemukan di feeder mapping Excel.', title="Format tidak valid")

    def column(name):
        return header.index(name) if name in header else -1

    part_idx = column("partnumber")
    table_idx = column("table")
    slot_idx = column("slot")
    position_idx = column("position")
    location_idx = column("locationcode")

    records = []
    for row in rows[header_index + 1 :]:
        part_number = _row_value(row, part_idx)
        if not part_number:
            continue
        table = _row_value(row, table_idx)
        slot = _row_value(row, slot_idx)
        position = _row_value(row, position_idx)
        location_code = _row_value(row, location_idx)
        records.append(
            {
                "table": table,
                "slot": slot,
                "position": position,
                "location_code": location_code,
                "part_number": part_number,
            }
        )

    return records


def _program_info(file_path):
    path = Path(file_path)
    model_part_numbers = parse_model_part_numbers(path.name)
    pcb_part_number, revision = parse_pcb_part_number(path)
    pcb_display = format_pcb_part_number(pcb_part_number, revision)
    model_display = ", ".join(model_part_numbers) if model_part_numbers else "-"

    if pcb_display != "-" and model_display != "-":
        display_name = f"{pcb_display} | {model_display}"
    elif pcb_display != "-":
        display_name = f"{pcb_display} | {path.stem}"
    elif model_display != "-":
        display_name = f"{model_display} | {path.stem}"
    else:
        display_name = path.stem

    return ProgramInfo(
        key=str(path.resolve()).upper(),
        display_name=display_name,
        model_part_numbers=model_part_numbers,
        pcb_part_number=pcb_part_number,
        revision=revision,
        source_folder=path.parent.name,
        source_file=path.name,
    )


def _build_matrix_rows(usage_map, program_lookup):
    rows = []
    for usage in usage_map.values():
        rows.append(
            {
                "component_part_number": usage.part_number,
                "used_model_count": usage.program_count,
                "total_bom_rows": usage.total_rows,
                "programs": _format_programs(usage.programs.keys(), program_lookup, limit=30),
                "_program_counts": dict(usage.programs),
            }
        )
    rows.sort(key=lambda row: (-row["used_model_count"], row["component_part_number"].upper()))
    return rows


def _compatibility_sort_key(row):
    status_order = {STATUS_SAFE: 0, STATUS_CHECK: 1, STATUS_CONFLICT: 2}
    return (
        status_order.get(row.get("status"), 9),
        -int(row.get("candidate_usage_count") or 0),
        str(row.get("candidate_part_number", "")).upper(),
        str(row.get("main_part_number", "")).upper(),
        str(row.get("location_code", "")).upper(),
    )


def _write_records_sheet(worksheet, rows, columns):
    worksheet.append([header for _, header in columns])
    for row in rows:
        worksheet.append([row.get(key, "") for key, _ in columns])
    _style_sheet(worksheet)


def _write_usage_matrix(worksheet, matrix_rows, program_infos):
    headers = ["Component P/N", "Used In", "BOM Rows"] + [program.display_name for program in program_infos]
    worksheet.append(headers)
    for row in matrix_rows:
        counts = row.get("_program_counts", {})
        worksheet.append(
            [
                row.get("component_part_number", ""),
                row.get("used_model_count", 0),
                row.get("total_bom_rows", 0),
                *[counts.get(program.key, "") for program in program_infos],
            ]
        )
    _style_sheet(worksheet, freeze="D2")


def _write_scan_log(worksheet, result):
    rows = [
        ("Excel files found", result.total_files),
        ("Files read", result.read_files),
        ("Components found", result.component_count),
        ("Fixed feeder rows", len(result.feeder_records)),
        ("Candidate parts", result.candidate_count),
        ("SAFE pairs", result.safe_count),
        ("CONFLICT pairs", result.conflict_count),
        ("CHECK pairs", result.check_count),
        ("Skipped/error files", len(result.skipped_files)),
    ]
    worksheet.append(["Item", "Value"])
    for item, value in rows:
        worksheet.append([item, value])

    if result.skipped_files:
        worksheet.append([])
        worksheet.append(["Skipped/error detail", ""])
        for skipped in result.skipped_files:
            worksheet.append([skipped, ""])

    _style_sheet(worksheet)


def _style_sheet(worksheet, freeze="A2"):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color="1F2937")

    if worksheet.max_row >= 1:
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    worksheet.freeze_panes = freeze
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"

    for column_index in range(1, worksheet.max_column + 1):
        letter = get_column_letter(column_index)
        max_length = 10
        for row_index in range(1, min(worksheet.max_row, 200) + 1):
            value = worksheet.cell(row_index, column_index).value
            if value not in (None, ""):
                max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[letter].width = min(max_length + 2, 42)


def _find_excel_files(source_folder):
    return scan_recursive_files(source_folder, EXCEL_EXTENSIONS, skip_prefixes=("~$",))


def _clean_part_values(values):
    parts = []
    for value in values:
        part = _part_text(value)
        if not part:
            continue
        if _is_header_text(part):
            continue
        if _is_excluded_part(part):
            continue
        parts.append(part)
    return parts


def _format_programs(program_keys, program_lookup, limit=8):
    keys = sorted(
        [key for key in program_keys if key in program_lookup],
        key=lambda key: program_lookup[key].display_name,
    )
    if not keys:
        return "-"
    names = [program_lookup[key].display_name for key in keys[:limit]]
    if len(keys) > limit:
        names.append(f"... +{len(keys) - limit} more")
    return "; ".join(names)


def _find_sheet_case_insensitive(sheet_names, target_name):
    target = target_name.lower()
    for sheet_name in sheet_names:
        if str(sheet_name).lower() == target:
            return sheet_name
    return None


def _row_value(row, index):
    if index < 0 or index >= len(row):
        return ""
    return _part_text(row[index])


def _part_text(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
    text = str(value)
    text = text.replace("\r", " ").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def _part_key(value):
    return _part_text(value).upper()


def _normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "", _part_text(value).lower())


def _is_header_text(value):
    return _part_key(value) in {
        "PART",
        "PART NAME",
        "PART NO",
        "PART NO.",
        "PART NUMBER",
        "PARTNUMBER",
        "PART_NAME",
        "P/N",
        "P/N COMPONENT",
    }


def _is_excluded_part(value):
    text = _part_key(value)
    if text.startswith("(") and text.endswith(")") and len(text) > 2:
        text = text[1:-1].strip()
    if len(text) >= 4 and text[:4] in {"EBU ", "EBT ", "EBR "}:
        return True
    return len(text) >= 3 and text[:3] in {"EAX", "EBU", "EBT", "EBR"}


def _is_na_marker(value):
    return _part_key(value) in {"#N/A", "#NA", "N/A"}


def _validate_config(config):
    if not config.source_folder:
        raise ServiceError("Folder Induk PCB belum dipilih.", title="Input belum lengkap")
    if not Path(config.source_folder).is_dir():
        raise ServiceError(f"Folder Induk PCB tidak ditemukan:\n{config.source_folder}", title="Folder tidak ditemukan")
    if not config.feeder_source_path:
        raise ServiceError("Fixed feeder source belum dipilih.", title="Input belum lengkap")
    if not Path(config.feeder_source_path).is_file():
        raise ServiceError(f"Fixed feeder source tidak ditemukan:\n{config.feeder_source_path}", title="File tidak ditemukan")


def _normalize_output_path(path):
    output_path = Path(path)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")
    return output_path


def _emit_progress(progress_callback, percent, message):
    if progress_callback:
        progress_callback(percent, message)


def _error_message(exc):
    return getattr(exc, "message", str(exc))


def _read_cm602_parts(path, mc_filter="3"):
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in OPENPYXL_EXTENSIONS:
        return _read_cm602_parts_openpyxl(path, mc_filter)
    if suffix == ".xls":
        return _read_cm602_parts_pandas(path, mc_filter)
    raise ServiceError(f"Format file tidak didukung: {path.suffix}", title="Format tidak valid")


def _read_cm602_parts_openpyxl(path, mc_filter="3"):
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = _find_sheet_case_insensitive(workbook.sheetnames, "CM602")
        if worksheet is None:
            raise ServiceError('Sheet "CM602" tidak ditemukan.', title="Format Excel tidak valid")

        values = []
        for row in workbook[worksheet].iter_rows(min_row=2, min_col=11, max_col=14, values_only=True):
            part_value = row[0] if row else None
            mc_value = row[3] if len(row) > 3 else None
            if mc_value is not None and str(mc_value).strip() == mc_filter:
                if part_value and str(part_value).strip():
                    values.append(part_value)
        return _clean_part_values(values)
    finally:
        workbook.close()


def _read_cm602_parts_pandas(path, mc_filter="3"):
    try:
        import pandas as pd
    except ImportError as exc:
        raise ServiceError(
            "File .xls membutuhkan pandas dan xlrd. Jalankan install dependency dari requirements.txt.",
            title="Dependency belum lengkap",
        ) from exc

    excel = open_pandas_excel_file(pd, path)
    try:
        sheet_name = _find_sheet_case_insensitive(excel.sheet_names, "CM602")
        if sheet_name is None:
            raise ServiceError('Sheet "CM602" tidak ditemukan.', title="Format Excel tidak valid")
        
        try:
            dataframe = pd.read_excel(
                excel,
                sheet_name=sheet_name,
                header=None,
                dtype=object,
                na_filter=False,
                usecols="K,N",
            )
            values = []
            for i in range(1, len(dataframe)):
                part_value = dataframe.iloc[i, 0]
                mc_value = dataframe.iloc[i, 1]
                if str(mc_value).strip() == mc_filter:
                    if part_value and str(part_value).strip():
                        values.append(part_value)
            return _clean_part_values(values)
        except ValueError:
            dataframe = pd.read_excel(excel, sheet_name=sheet_name, header=None, dtype=object, na_filter=False)
            values = []
            for i in range(1, len(dataframe)):
                if len(dataframe.columns) >= 14:
                    part_value = dataframe.iloc[i, 10]
                    mc_value = dataframe.iloc[i, 13]
                    if str(mc_value).strip() == mc_filter:
                        if part_value and str(part_value).strip():
                            values.append(part_value)
            return _clean_part_values(values)
    finally:
        excel.close()
