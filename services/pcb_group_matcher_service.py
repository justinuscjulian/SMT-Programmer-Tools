import csv
import math
import re
from collections import OrderedDict, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from services.common_feeder_reuse_service import _clean_part_values, _part_key, _part_text, _read_bom_parts
from services.errors import ServiceError
from services.io_helpers import scan_recursive_files
from services.all_table_feeder_group_service import VirtualMachine, get_master_mapping, GroupResult, export_all_table_groups, PcbInfo
import services.feeder_mapping_service as fms

MODE_BOM_FILE = "bom_file"
MODE_PROGRAM_FOLDER = "program_folder"
MODE_PROGRAM_EXCEL = "program_excel"

EXCEL_EXTENSIONS = (".xls", ".xlsx", ".xlsm", ".xlsb")
TEXT_EXTENSIONS = (".csv", ".tsv", ".txt")


@dataclass
class PcbGroupMatcherConfig:
    import_mode: str
    pcb_source_path: str
    fix_feeder_group_path: str
    line_type: str = "Line 1-5"
    output_path: str = ""


@dataclass
class GroupMatchItem:
    rank: int
    group_name: str
    member_pcbs: list[str]
    total_pcb_parts: int
    matched_count: int
    missing_count: int
    match_rate_percent: float
    status_recommendation: str
    recommendation_note: str
    matched_details: list[dict] = field(default_factory=list)
    missing_details: list[str] = field(default_factory=list)


@dataclass
class PcbGroupMatcherResult:
    output_path: str
    fix_feeder_group_path: str
    pcb_name: str
    import_mode: str
    pcb_parts_count: int
    unique_parts: list[str]
    group_matches: list[GroupMatchItem]
    best_match: GroupMatchItem


def suggest_output_name(pcb_name="NEW_PCB"):
    clean_name = re.sub(r'[\\/*?:"<>|]', '_', pcb_name).strip() or "NEW_PCB"
    return f"PCB_Group_Matcher_{clean_name}_{datetime.now().strftime('%y%m%d')}.xlsx"


def analyze_pcb_group_matcher(config: PcbGroupMatcherConfig, progress_callback=None):
    _validate_config(config)

    _emit_progress(progress_callback, 5, "Membaca komponen PCB baru...")
    pcb_parts, pcb_name = _extract_new_pcb_parts(config.import_mode, config.pcb_source_path, config.line_type)
    if not pcb_parts:
        raise ServiceError(
            "Tidak ditemukan Part Number komponen yang valid dari input PCB baru yang diberikan.",
            title="Komponen kosong",
        )

    _emit_progress(progress_callback, 25, "Membaca data Fix Feeder Groups dari Excel...")
    group_parts_map, group_pcbs_map = _load_fix_feeder_groups(config.fix_feeder_group_path)
    if not group_parts_map:
        raise ServiceError(
            "File Fix Feeder Group Excel tidak berisi data sheet group yang valid.",
            title="Data Fix Feeder kosong",
        )

    _emit_progress(progress_callback, 55, "Melakukan komparasi kecocokan PCB ke seluruh Group...")
    raw_matches = []
    total_groups = len(group_parts_map)
    total_parts = len(pcb_parts)

    for idx, (group_name, group_parts) in enumerate(group_parts_map.items(), start=1):
        percent = 55 + int(idx / max(1, total_groups) * 35)
        _emit_progress(progress_callback, percent, f"Menganalisis kecocokan {group_name} ({idx}/{total_groups})...")

        matched_details = []
        missing_details = []

        for part in pcb_parts:
            key = _part_key(part)
            if key in group_parts:
                g_info = group_parts[key]
                matched_details.append({
                    "part_number": part,
                    "location_code": g_info["location_code"],
                    "type": g_info["type"],
                })
            else:
                missing_details.append(part)

        matched_count = len(matched_details)
        missing_count = len(missing_details)
        match_rate = round((matched_count / total_parts) * 100.0, 1)

        raw_matches.append({
            "group_name": group_name,
            "member_pcbs": group_pcbs_map.get(group_name, []),
            "total_pcb_parts": total_parts,
            "matched_count": matched_count,
            "missing_count": missing_count,
            "match_rate_percent": match_rate,
            "matched_details": matched_details,
            "missing_details": missing_details,
        })

    # Sort matches by match_rate desc, missing_count asc, group_name asc
    raw_matches.sort(key=lambda x: (-x["match_rate_percent"], x["missing_count"], x["group_name"]))

    group_matches = []
    for rank, item in enumerate(raw_matches, start=1):
        rate = item["match_rate_percent"]
        matched_cnt = item["matched_count"]
        missing_cnt = item["missing_count"]
        tot_cnt = item["total_pcb_parts"]

        if rank == 1:
            if rate >= 90.0:
                status = "SANGAT DIREKOMENDASIKAN (BEST MATCH)"
                note = f"Paling efisien! {matched_cnt} dari {tot_cnt} komponen ({rate}%) sudah terpasang. Hanya butuh {missing_cnt} feeder ekstra di slot kosong."
            elif rate >= 70.0:
                status = "DIREKOMENDASIKAN"
                note = f"Sangat cocok! {matched_cnt} dari {tot_cnt} komponen ({rate}%) ter-cover. Butuh {missing_cnt} feeder tambahan."
            else:
                status = "KAPASITAS TERBATAS"
                note = f"Kecocokan tertinggi saat ini {rate}%. Perlu {missing_cnt} feeder tambahan dipasang manual."
        else:
            if rate >= 85.0:
                status = "HIGH MATCH"
                note = f"Kecocokan tinggi ({rate}%). Butuh {missing_cnt} feeder tambahan."
            elif rate >= 65.0:
                status = "MODERATE MATCH"
                note = f"Kecocokan sedang ({rate}%). Butuh {missing_cnt} feeder tambahan."
            else:
                status = "LOW MATCH"
                note = f"Kecocokan rendah ({rate}%). {missing_cnt} komponen perlu pasang baru."

        group_matches.append(GroupMatchItem(
            rank=rank,
            group_name=item["group_name"],
            member_pcbs=item["member_pcbs"],
            total_pcb_parts=tot_cnt,
            matched_count=matched_cnt,
            missing_count=missing_cnt,
            match_rate_percent=rate,
            status_recommendation=status,
            recommendation_note=note,
            matched_details=item["matched_details"],
            missing_details=item["missing_details"],
        ))

    best_match = group_matches[0] if group_matches else None

    _emit_progress(progress_callback, 100, "Analisis kecocokan selesai!")
    return PcbGroupMatcherResult(
        output_path=config.output_path,
        fix_feeder_group_path=config.fix_feeder_group_path,
        pcb_name=pcb_name,
        import_mode=config.import_mode,
        pcb_parts_count=len(pcb_parts),
        unique_parts=pcb_parts,
        group_matches=group_matches,
        best_match=best_match,
    )


def export_pcb_group_matcher_result(result: PcbGroupMatcherResult, output_path: str):
    if not result or not result.group_matches:
        raise ServiceError("Tidak ada hasil analisis kecocokan untuk diexport.", title="Data kosong")

    out_path = Path(output_path)
    if out_path.suffix.lower() != ".xlsx":
        out_path = out_path.with_suffix(".xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Sheet 1: Ranking Summary
    ws_rank = wb.active
    ws_rank.title = "Group Match Ranking"

    ws_rank.append([
        "Rank",
        "Group Name",
        "Match Rate (%)",
        "Matched Parts",
        "Extra Feeders Needed",
        "Total PCB Parts",
        "Member PCBs in Group",
        "Status & Recommendation",
        "Note",
    ])

    for item in result.group_matches:
        ws_rank.append([
            item.rank,
            item.group_name,
            f"{item.match_rate_percent}%",
            f"{item.matched_count} / {item.total_pcb_parts}",
            item.missing_count,
            item.total_pcb_parts,
            ", ".join(item.member_pcbs) if item.member_pcbs else "-",
            item.status_recommendation,
            item.recommendation_note,
        ])

    _style_summary_sheet(ws_rank)

    # Sheet 2: Matched Components Detail
    ws_matched = wb.create_sheet("Matched Components")
    ws_matched.append(["Group Name", "Location Code", "Component Part Number", "Fix Type"])
    for item in result.group_matches:
        for m in item.matched_details:
            ws_matched.append([
                item.group_name,
                m["location_code"],
                m["part_number"],
                m["type"],
            ])
    _style_detail_sheet(ws_matched)

    # Sheet 3: Extra Components Needed (Change-Over)
    ws_extra = wb.create_sheet("Extra Components Needed")
    ws_extra.append(["Group Name", "Extra Component Part Number", "Action Needed"])
    for item in result.group_matches:
        for p in item.missing_details:
            ws_extra.append([
                item.group_name,
                p,
                "Butuh Feeder Tambahan (Pasang di Slot Kosong)",
            ])
    _style_detail_sheet(ws_extra)

    wb.save(out_path)
    return str(out_path)


def _extract_new_pcb_parts(mode, source_path, line_type="Line 1-5"):
    path = Path(source_path)
    parts_set = set()
    pcb_name = path.stem

    def _read_parts_smart(file_path):
        if line_type == "Line 9":
            try:
                from services.model_feeder_group_service import _read_cm602_parts
                return _read_cm602_parts(file_path, mc_filter="3")
            except Exception:
                pass
        elif line_type == "CM602":
            try:
                from services.model_feeder_group_service import _read_cm602_parts
                return _read_cm602_parts(file_path, mc_filter="0")
            except Exception:
                pass
        return _read_parts_from_excel(file_path)

    if mode == MODE_BOM_FILE:
        if path.suffix.lower() in EXCEL_EXTENSIONS:
            part_list = _read_parts_from_excel(path)
        elif path.suffix.lower() in TEXT_EXTENSIONS:
            part_list = _read_parts_from_text_bom(path)
        else:
            raise ServiceError(f"Ekstensi file BOM tidak didukung: {path.suffix}", title="Format file tidak didukung")
        for p in part_list:
            key = _part_key(p)
            if key:
                parts_set.add(p)

    elif mode == MODE_PROGRAM_EXCEL:
        if path.suffix.lower() not in EXCEL_EXTENSIONS:
            raise ServiceError("File program PCB harus berformat Excel (.xlsx, .xls, .xlsm).", title="Format file tidak valid")
        part_list = _read_parts_smart(path)
        for p in part_list:
            key = _part_key(p)
            if key:
                parts_set.add(p)

    elif mode == MODE_PROGRAM_FOLDER:
        if not path.is_dir():
            raise ServiceError(f"Folder PCB tidak ditemukan:\n{path}", title="Folder tidak ditemukan")
        pcb_name = path.name
        
        from services.model_feeder_group_service import _scan_models
        # We pass target_pcb_list=[pcb_name] to filter exactly this folder if it's within a parent folder.
        # However, _scan_models's source_folder is `path`. If `path` is the PCB folder itself,
        # _pcb_folders(path) will just return `[path]`, so we can pass target_pcb_list=None.
        models, _, _, skipped = _scan_models(source_folder=str(path), target_pcb_list=None, progress_callback=None, line_type=line_type)
        
        if not models:
            raise ServiceError(f"Tidak ada part number yang ditemukan dalam folder program:\n{path}\nSkipped: {skipped}", title="Komponen kosong")
            
        for model in models.values():
            for part in model.components.values():
                if part:
                    parts_set.add(part)
        # Use the primary model name if possible
        if models:
            pcb_name = list(models.values())[0].pcb_part_number

    else:
        raise ServiceError(f"Mode import tidak dikenal: {mode}", title="Mode tidak valid")

    # Sort parts alphabetically
    sorted_parts = sorted(list(parts_set), key=lambda x: x.upper())
    return sorted_parts, pcb_name


def _read_parts_from_excel(path):
    try:
        part_values = _read_bom_parts(path)
        if part_values:
            return part_values
    except Exception:
        pass

    # Fallback openpyxl scan across worksheets
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        parts = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    val = _part_text(cell)
                    if val and not val.isdigit() and len(val) >= 4 and not val.upper().startswith(("PART", "SLOT", "LOCATION")):
                        parts.append(val)
        wb.close()
        return _clean_part_values(parts)
    except Exception:
        return []


def _read_parts_from_text_bom(path):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
    except Exception as exc:
        raise ServiceError(f"Gagal membaca file text BOM:\n{path}", title="File Error") from exc

    lines = [line.strip() for line in content.splitlines() if line.strip()]
    if not lines:
        return []

    sample = lines[0]
    delim = "\t" if "\t" in sample else (";" if ";" in sample else ",")
    reader = csv.reader(lines, delimiter=delim)
    header = None
    part_col = -1
    raw_parts = []

    for row in reader:
        if not row:
            continue
        clean_row = [_part_text(c) for c in row]
        if header is None:
            upper_row = [c.upper() for c in clean_row]
            for idx, c in enumerate(upper_row):
                if "PART" in c or "P/N" in c or "COMPONENT" in c:
                    header = upper_row
                    part_col = idx
                    break
            if header is not None:
                continue

        if part_col >= 0 and len(clean_row) > part_col:
            val = clean_row[part_col]
        else:
            val = clean_row[0] if clean_row else ""

        if val and not val.upper().startswith(("PART", "P/N", "LOCATION", "SKIPPED")):
            raw_parts.append(val)

    return _clean_part_values(raw_parts)


def _load_fix_feeder_groups(fix_feeder_group_path):
    path = Path(fix_feeder_group_path)
    if not path.is_file():
        raise ServiceError(f"File Fix Feeder Group Excel tidak ditemukan:\n{path}", title="File tidak ditemukan")

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise ServiceError("File Excel Fix Feeder Group tidak dapat dibaca.", title="Excel Error") from exc

    group_pcbs_map = defaultdict(list)
    if "Summary" in wb.sheetnames:
        ws_sum = wb["Summary"]
        header = None
        group_idx = -1
        pcb_idx = -1
        for row in ws_sum.iter_rows(values_only=True):
            vals = [_part_text(c) for c in row]
            if not any(vals):
                continue
            if header is None:
                header_upper = [v.upper() for v in vals]
                if "GROUP NAME" in header_upper and "PCB NAME" in header_upper:
                    header = header_upper
                    group_idx = header_upper.index("GROUP NAME")
                    pcb_idx = header_upper.index("PCB NAME")
            else:
                if group_idx >= 0 and pcb_idx >= 0 and len(vals) > max(group_idx, pcb_idx):
                    g_name = vals[group_idx]
                    p_name = vals[pcb_idx]
                    if g_name and p_name:
                        if p_name not in group_pcbs_map[g_name]:
                            group_pcbs_map[g_name].append(p_name)

    group_parts_map = OrderedDict()

    for worksheet in wb.worksheets:
        title = worksheet.title
        if title.lower() in {"summary", "summary sheet"}:
            continue

        parts_in_group = OrderedDict()
        header_map = {}
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            raw_cells = [_part_text(c) for c in row]
            while raw_cells and not raw_cells[-1]:
                raw_cells.pop()
            if not any(raw_cells):
                continue

            cells_upper = [c.upper() for c in raw_cells]
            if "PART NUMBER" in cells_upper or "LOCATION CODE" in cells_upper or "TYPE" in cells_upper:
                for idx, c in enumerate(cells_upper):
                    if "PART NUMBER" in c or "PART" in c:
                        header_map["part"] = idx
                    elif "LOCATION" in c or "SLOT" in c:
                        header_map["loc"] = idx
                    elif "TYPE" in c:
                        header_map["type"] = idx
                continue

            part_num = ""
            loc_code = ""
            part_type = "FIXED"

            if "loc" in header_map and "part" in header_map:
                if len(raw_cells) > max(header_map["loc"], header_map["part"]):
                    loc_code = raw_cells[header_map["loc"]]
                    part_num = raw_cells[header_map["part"]]
                    if "type" in header_map and len(raw_cells) > header_map["type"]:
                        part_type = raw_cells[header_map["type"]].upper() or "FIXED"

            if not loc_code or not part_num:
                # Fallback check row for location code & part number
                for idx, c in enumerate(raw_cells):
                    if re.match(r"^\[(\d+)\](\d+)", c):
                        loc_code = c
                        break
                if loc_code:
                    for c in raw_cells:
                        if c != loc_code and c.upper() not in {"FIXED", "SUBSTITUTE", "DYNAMIC", "L", "R"} and not re.match(r"^\[(\d+)\](\d+)", c):
                            part_num = c
                            break

            if loc_code and part_num and part_type in {"FIXED", "SUBSTITUTE"}:
                key = _part_key(part_num)
                if key:
                    parts_in_group[key] = {
                        "part_number": part_num,
                        "location_code": loc_code,
                        "type": part_type,
                    }

        if parts_in_group:
            group_parts_map[title] = parts_in_group

    wb.close()
    return group_parts_map, group_pcbs_map


def _validate_config(config: PcbGroupMatcherConfig):
    if not config.import_mode:
        raise ServiceError("Metode import PCB belum dipilih.", title="Input belum lengkap")
    if not config.pcb_source_path:
        raise ServiceError("File / Folder PCB baru belum dipilih.", title="Input belum lengkap")
    if not config.fix_feeder_group_path:
        raise ServiceError("File Fix Feeder Group Excel belum dipilih.", title="Input belum lengkap")


def _emit_progress(progress_callback, percent, message):
    if progress_callback:
        progress_callback(percent, message)


def _style_summary_sheet(ws):
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx in {1, 3, 4, 5, 6}:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)


def _style_detail_sheet(ws):
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = left_align

    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)



import re

def _load_ohm_ini_library():
    import win32com.client
    
    file_path = r"C:\Users\User\Documents\PROJECT\SMT Programmer Tools\SMT-Programmer-Tools\≪ OHM_INI ≫.xlsb"
    if not Path(file_path).is_file():
        return {}
        
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    
    library_map = {}
    try:
        wb = excel.Workbooks.Open(file_path, ReadOnly=True)
        ws = wb.Worksheets(1)
        
        data = ws.UsedRange.Value
        if data:
            for i, row in enumerate(data):
                if i == 0: continue # Skip header
                if len(row) > 4:
                    part = str(row[0] or "").strip()
                    fd = str(row[4] or "").strip()
                    if part and fd:
                        library_map[part] = fd
    except Exception as e:
        print(f"Failed to read OHM_INI library: {e}")
    finally:
        try:
            wb.Close(False)
        except:
            pass
        excel.Quit()
        
    return library_map


def generate_merged_fix_feeder_group(
    pcb_result: PcbGroupMatcherResult,
    selected_group_name: str,
    line_type: str,
    master_excel_path: str,
    output_excel_path: str,
    base_npm_path: str = None,
    progress_callback=None
):
    _emit_progress(progress_callback, 10, "Membaca data Base Fix Feeder Group...")
    group_parts_map, group_pcbs_map = _load_fix_feeder_groups(pcb_result.fix_feeder_group_path)
    if selected_group_name not in group_parts_map:
        raise ServiceError(f"Group '{selected_group_name}' tidak ditemukan di file base.", title="Group tidak ditemukan")

    base_parts = group_parts_map[selected_group_name]
    base_pcbs = group_pcbs_map.get(selected_group_name, [])

    _emit_progress(progress_callback, 30, "Membaca Master Mapping Excel...")
    master = get_master_mapping(master_excel_path, line_type)

    # Initialize VirtualMachine and Mappings
    vm = VirtualMachine(line_type)
    slot_mapping = {}
    part_mapping = {}
    substitute_mapping = {}

    for key, info in base_parts.items():
        loc = info["location_code"]
        part = info["part_number"]
        part_type = info.get("type", "FIXED").upper()
        
        if part_type == "FIXED":
            if vm.can_add(loc):
                vm.add(loc)
            slot_mapping[loc] = part
            part_mapping[part] = loc
        else:
            if loc not in substitute_mapping:
                substitute_mapping[loc] = []
            substitute_mapping[loc].append((part, ["BASE"]))
            part_mapping[part] = loc

    # Get missing parts that need to be merged
    missing_parts = []
    for match in pcb_result.group_matches:
        if match.group_name == selected_group_name:
            missing_parts = match.missing_details
            break

    _emit_progress(progress_callback, 50, f"Menempatkan {len(missing_parts)} komponen baru ke slot kosong...")
    unassigned_parts = []
    
    # Sort missing parts based on master frequency / options
    def part_sort_key(p):
        loc_list = master.get(p, [])
        num_options = 0
        for item in loc_list:
            if item["location"]:
                num_options += 1
            num_options += len(item.get("alternatives", []))
        if num_options == 0:
            num_options = 999
        master_freq = max([item["frequency"] for item in loc_list], default=0)
        return (num_options, -master_freq)

    sorted_missing = sorted(list(set(missing_parts)), key=part_sort_key)

    ohm_ini_library = None
    ohm_ini_loaded = False

    for part in sorted_missing:
        loc_list = master.get(part, [])
        if not loc_list:
            pass # We will try fallback below
        else:
            loc_list = sorted(loc_list, key=lambda x: x["frequency"], reverse=True)
            primary_loc = loc_list[0]["location"]
            if primary_loc:
                candidates = [primary_loc]
                for alt in loc_list[0].get("alternatives", []):
                    if alt and alt not in candidates:
                        candidates.append(alt)
                        
                fallback_tables = None
                if line_type == "Line 9":
                    allowed_t_set = set()
                    for item in master.get(part, []):
                        l_val = item.get("location")
                        if l_val:
                            p_l = vm._parse_loc(l_val)
                            if p_l: allowed_t_set.add(p_l[0])
                        for alt in item.get("alternatives", []):
                            if alt:
                                p_a = vm._parse_loc(alt)
                                if p_a: allowed_t_set.add(p_a[0])
                    if allowed_t_set:
                        fallback_tables = sorted(list(allowed_t_set))
        
                placed = False
                for loc in candidates:
                    if vm.can_add(loc):
                        vm.add(loc)
                        slot_mapping[loc] = part
                        part_mapping[part] = loc
                        placed = True
                        break
                    else:
                        fallback = vm.find_fallback(loc, fallback_tables=fallback_tables)
                        if fallback:
                            vm.add(fallback)
                            slot_mapping[fallback] = part
                            part_mapping[part] = fallback
                            placed = True
                            break
                
                if placed:
                    continue
        
        # Fallback to OHM_INI library if part not in master or no slot found
        if line_type in ("Line 1-5", "Line 6-7"):
            if not ohm_ini_loaded:
                _emit_progress(progress_callback, 60, "Membaca OHM INI Library...")
                ohm_ini_library = _load_ohm_ini_library()
                ohm_ini_loaded = True
                
            fd_val = ohm_ini_library.get(part, "")
            if fd_val and "TRAY" not in fd_val.upper():
                m = re.search(r'(?:E|P|e|p)[^\d]*(\d{2})', fd_val)
                if m:
                    size = int(m.group(1))
                    dummy_loc = None
                    allowed_tables = None
                    if size == 8:
                        dummy_loc = "[1]1L"
                        allowed_tables = [1, 2, 3, 4, 5, 6, 8]
                    elif 12 <= size <= 24:
                        dummy_loc = "[7]1" if size < 24 else "[7]1-2"
                        allowed_tables = [7]
                    elif size >= 32:
                        dummy_loc = "[9]1-2"
                        allowed_tables = [9]
                        
                    if dummy_loc and allowed_tables:
                        fallback = vm.find_fallback(dummy_loc, fallback_tables=allowed_tables)
                        if fallback:
                            vm.add(fallback)
                            slot_mapping[fallback] = part
                            part_mapping[part] = fallback
                            continue
                            
        # Fallback to Substitute Logic
        valid_tables = None
        part_span = None
        
        # 1. Constraints from master
        loc_list = master.get(part, [])
        if loc_list:
            valid_tables = set()
            for item in loc_list:
                l_val = item.get("location")
                if l_val:
                    p = vm._parse_loc(l_val)
                    if p:
                        valid_tables.add(p[0])
                        part_span = p[2] - p[1] + 1
                for alt in item.get("alternatives", []):
                    if alt:
                        p = vm._parse_loc(alt)
                        if p:
                            valid_tables.add(p[0])
                            part_span = p[2] - p[1] + 1
            if not valid_tables:
                valid_tables = None

        # 2. Constraints from OHM INI
        if not valid_tables and line_type in ("Line 1-5", "Line 6-7") and ohm_ini_library:
            fd_val = ohm_ini_library.get(part, "")
            if fd_val and "TRAY" not in fd_val.upper():
                m = re.search(r'(?:E|P|e|p)[^\d]*(\d{2})', fd_val)
                if m:
                    size = int(m.group(1))
                    if size == 8:
                        valid_tables = {1, 2, 3, 4, 5, 6, 8}
                        part_span = 1
                    elif 12 <= size <= 24:
                        valid_tables = {7}
                        part_span = 1 if size < 24 else 2
                    elif size >= 32:
                        valid_tables = {9}
                        part_span = 2

        if valid_tables:
            placed_as_sub = False
            for loc, occupant in slot_mapping.items():
                parsed_loc = vm._parse_loc(loc)
                if not parsed_loc or parsed_loc[0] not in valid_tables:
                    continue
                loc_span = parsed_loc[2] - parsed_loc[1] + 1
                if part_span and loc_span != part_span:
                    continue
                    
                if occupant in pcb_result.unique_parts:
                    continue
                
                subs = substitute_mapping.get(loc, [])
                conflict = False
                for sub_part, _ in subs:
                    if sub_part in pcb_result.unique_parts:
                        conflict = True
                        break
                if conflict:
                    continue
                    
                if loc not in substitute_mapping:
                    substitute_mapping[loc] = []
                substitute_mapping[loc].append((part, [pcb_result.pcb_name]))
                part_mapping[part] = loc
                placed_as_sub = True
                break
                
            if placed_as_sub:
                continue
                            
        unassigned_parts.append(part)

    _emit_progress(progress_callback, 70, "Mengekspor hasil ke Excel...")
    
    # Combine PCBs
    combined_pcb_names = list(base_pcbs)
    if pcb_result.pcb_name not in combined_pcb_names:
        combined_pcb_names.append(pcb_result.pcb_name)
        
    group_pcbs = [PcbInfo(name, "", set()) for name in combined_pcb_names]

    res = GroupResult(
        group_name=selected_group_name + "_Merged",
        pcbs=group_pcbs,
        slot_mapping=slot_mapping,
        part_mapping=part_mapping,
        unassigned_parts=unassigned_parts,
        substitute_mapping=substitute_mapping,
        master_excel_path=master_excel_path
    )

    saved_path = export_all_table_groups([res], output_excel_path, master_excel_path)

    _emit_progress(progress_callback, 90, "Generate file mesin (NPM/CRB)...")
    if base_npm_path and Path(base_npm_path).is_file():
        out_dir = Path(saved_path).parent / f"{Path(saved_path).stem}_NPM_Files"
        out_dir.mkdir(parents=True, exist_ok=True)
        if line_type == "Line 8":
            fms.generate_npm_feeder_import_batch_from_groups_line8(saved_path, base_npm_path, str(out_dir))
        elif line_type == "Line 9":
            fms.generate_npm_feeder_import_batch_from_groups_line9(saved_path, base_npm_path, str(out_dir))
        elif line_type == "CM602":
            fms.generate_cm602_feeder_import_batch_from_groups(saved_path, base_npm_path, str(out_dir))
        else:
            fms.generate_npm_feeder_import_batch_from_groups(saved_path, base_npm_path, str(out_dir))

    _emit_progress(progress_callback, 100, "Proses Merge selesai!")
    return saved_path, len(unassigned_parts)
