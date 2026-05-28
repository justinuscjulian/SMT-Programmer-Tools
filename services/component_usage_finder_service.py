import math
import os
import re
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from services.errors import ServiceError


EXCEL_EXTENSIONS = (".xls", ".xlsx", ".xlsm")
OPENPYXL_EXTENSIONS = (".xlsx", ".xlsm")
TARGET_SHEET_NAME = "BOM"
PCB_PATTERN = re.compile(r"(?<![A-Za-z0-9])(EAX[A-Za-z0-9]{8})(?:\s*\(([^()]*)\))?", re.IGNORECASE)
MODEL_PATTERN = re.compile(r"(?<![A-Za-z0-9])(?:EBU|EBT|EBR)[A-Za-z0-9-]{5,}", re.IGNORECASE)


@dataclass
class ComponentUsageFinderConfig:
    component_part_number: str
    source_folder: str


@dataclass
class ComponentUsageRow:
    component_part_number: str
    model_part_number: str
    pcb_part_number: str
    revision: str
    source_folder: str
    source_file: str
    found_rows: list[int] = field(default_factory=list)


@dataclass
class ComponentUsageFileMatch:
    source_folder: str
    source_file: str
    pcb_part_number: str
    revision: str
    found_rows: list[int]
    model_part_numbers: list[str]


@dataclass
class ComponentUsageSearchResult:
    component_part_number: str
    rows: list[ComponentUsageRow]
    matched_files: list[ComponentUsageFileMatch]
    total_files: int
    read_files: int
    skipped_files: list[str] = field(default_factory=list)


@dataclass
class BomSearchData:
    found_rows: list[int]
    model_part_numbers: list[str]


def find_component_usage(config: ComponentUsageFinderConfig, progress_callback=None):
    _validate_config(config)
    component_text = _value_text(config.component_part_number)
    target_key = _match_key(component_text)
    source_folder = Path(config.source_folder)

    _emit_progress(progress_callback, 0, "Scanning folders...")
    excel_files, walk_errors = _find_excel_files(source_folder)
    if not excel_files:
        return ComponentUsageSearchResult(
            component_part_number=component_text,
            rows=[],
            matched_files=[],
            total_files=0,
            read_files=0,
            skipped_files=walk_errors,
        )

    rows = []
    matched_files = []
    skipped_files = list(walk_errors)
    seen_result_keys = set()
    read_files = 0
    total_files = len(excel_files)

    for index, file_path in enumerate(excel_files, start=1):
        percent = max(1, min(99, int((index - 1) / total_files * 100)))
        _emit_progress(progress_callback, percent, f"Reading file {index}/{total_files}: {file_path.name}")

        try:
            bom_data = _read_bom_file(file_path, target_key)
        except Exception as exc:
            skipped_files.append(f"{file_path}: {_error_message(exc)}")
            continue

        read_files += 1
        if not bom_data.found_rows:
            continue

        model_part_numbers = parse_model_part_numbers(file_path.name)
        if not model_part_numbers:
            model_part_numbers = bom_data.model_part_numbers
        if not model_part_numbers:
            model_part_numbers = ["-"]

        pcb_part_number, revision = parse_pcb_part_number(file_path)
        pcb_display = format_pcb_part_number(pcb_part_number, revision)
        source_folder_name = file_path.parent.name

        matched_files.append(
            ComponentUsageFileMatch(
                source_folder=source_folder_name,
                source_file=file_path.name,
                pcb_part_number=pcb_part_number,
                revision=revision,
                found_rows=bom_data.found_rows,
                model_part_numbers=model_part_numbers,
            )
        )

        for model_part_number in model_part_numbers:
            result_key = (_match_key(model_part_number), _match_key(pcb_display))
            if result_key in seen_result_keys:
                continue
            seen_result_keys.add(result_key)
            rows.append(
                ComponentUsageRow(
                    component_part_number=component_text,
                    model_part_number=model_part_number,
                    pcb_part_number=pcb_display,
                    revision=revision,
                    source_folder=source_folder_name,
                    source_file=file_path.name,
                    found_rows=bom_data.found_rows,
                )
            )

    final_message = (
        f"Search complete: {len(rows)} result(s) found"
        if rows
        else "No result found"
    )
    _emit_progress(progress_callback, 100, final_message)
    return ComponentUsageSearchResult(
        component_part_number=component_text,
        rows=rows,
        matched_files=matched_files,
        total_files=total_files,
        read_files=read_files,
        skipped_files=skipped_files,
    )


def suggest_export_name(component_part_number):
    component_text = _value_text(component_part_number)
    safe_component = re.sub(r"[^A-Za-z0-9_.-]+", "_", component_text).strip("_") or "Component_Usage"
    return f"{safe_component}_Usage_{datetime.now().strftime('%y%m%d')}.xlsx"


def export_component_usage_result(result, output_path):
    if result is None:
        raise ServiceError("Belum ada hasil pencarian untuk diexport.", title="Data kosong")

    output = _normalize_output_path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    preview_sheet = workbook.active
    preview_sheet.title = "Preview Result"
    _write_preview_sheet(preview_sheet, result.rows)

    log_sheet = workbook.create_sheet("Scan Log")
    _write_log_sheet(log_sheet, result)

    workbook.save(output)
    return str(output)


def parse_model_part_numbers(filename):
    stem = Path(filename).stem
    groups = re.findall(r"\(([^()]*)\)", stem)

    for group in groups:
        if _looks_like_revision(group) or PCB_PATTERN.search(group):
            continue
        models = _extract_models_from_text(group)
        if models:
            return models

    if len(groups) >= 2:
        candidate_group = groups[1]
        if not _looks_like_revision(candidate_group) and not PCB_PATTERN.search(candidate_group):
            tokens = _split_part_tokens(candidate_group)
            if tokens:
                return _unique_text(tokens)

    return _extract_models_from_text(stem)


def parse_pcb_part_number(file_path):
    path = Path(file_path)
    file_pcb = _parse_pcb_from_text(path.stem)
    folder_pcb = _parse_pcb_from_text(path.parent.name)

    if file_pcb:
        pcb_part_number, revision = file_pcb
        if not revision and folder_pcb and folder_pcb[0] == pcb_part_number:
            revision = folder_pcb[1]
        return pcb_part_number, revision
    if folder_pcb:
        return folder_pcb
    return "-", ""


def format_pcb_part_number(pcb_part_number, revision):
    pcb_text = _value_text(pcb_part_number) or "-"
    revision_text = _value_text(revision)
    return f"{pcb_text}({revision_text})" if revision_text and pcb_text != "-" else pcb_text


def _write_preview_sheet(worksheet, rows):
    worksheet.append(["No", "Model Part Number", "PCB Part Number"])
    for index, row in enumerate(rows, start=1):
        worksheet.append([index, row.model_part_number, row.pcb_part_number])
    _style_sheet(worksheet)


def _write_log_sheet(worksheet, result):
    worksheet.append(["Field", "Value"])
    worksheet.append(["Component Part Number", result.component_part_number])
    worksheet.append(["Excel Files Found", result.total_files])
    worksheet.append(["Files Read", result.read_files])
    worksheet.append(["Preview Result Count", len(result.rows)])
    worksheet.append(["Skipped/Error Files", len(result.skipped_files)])
    worksheet.append([])

    worksheet.append(["Matched Source Files"])
    worksheet.append(["No", "Source Folder", "Source File", "PCB Part Number", "Revision", "Found Row", "Model Part Number"])
    for index, match in enumerate(result.matched_files, start=1):
        worksheet.append(
            [
                index,
                match.source_folder,
                match.source_file,
                format_pcb_part_number(match.pcb_part_number, match.revision),
                match.revision,
                ", ".join(str(row) for row in match.found_rows),
                ", ".join(match.model_part_numbers),
            ]
        )

    worksheet.append([])
    worksheet.append(["Skipped/Error"])
    worksheet.append(["No", "Message"])
    for index, skipped in enumerate(result.skipped_files, start=1):
        worksheet.append([index, skipped])

    _style_sheet(worksheet)
    worksheet.column_dimensions["G"].width = 44


def _style_sheet(worksheet):
    header_font = Font(name="Calibri", size=11, bold=True)
    body_font = Font(name="Calibri", size=11)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = header_font if cell.row == 1 else body_font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    section_titles = {"Matched Source Files", "Skipped/Error"}
    for row in worksheet.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        if cell.value in section_titles:
            cell.font = header_font

    for col_index in range(1, worksheet.max_column + 1):
        letter = get_column_letter(col_index)
        max_length = 10
        for row_index in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row_index, col_index).value
            if value not in (None, ""):
                max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[letter].width = min(max_length + 2, 60)

    worksheet.freeze_panes = "A2"


def _parse_pcb_from_text(value):
    match = PCB_PATTERN.search(str(value or ""))
    if not match:
        return None
    pcb_part_number = match.group(1).upper()
    revision = _value_text(match.group(2))
    return pcb_part_number, revision


def _read_bom_file(file_path, target_key):
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix in OPENPYXL_EXTENSIONS:
        return _read_bom_openpyxl(path, target_key)
    if suffix == ".xls":
        return _read_bom_pandas(path, target_key)
    raise ServiceError(f"Format file tidak didukung: {path.suffix}", title="Format tidak valid")


def _read_bom_openpyxl(path, target_key):
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = _find_sheet_case_insensitive(workbook.sheetnames, TARGET_SHEET_NAME)
        if worksheet is None:
            raise ServiceError('Sheet "BOM" tidak ditemukan.', title="Format Excel tidak valid")

        found_rows = OrderedDict()
        model_part_numbers = OrderedDict()
        for row_number, row_values in enumerate(workbook[worksheet].iter_rows(values_only=True), start=1):
            row_has_match = False
            for value in row_values:
                text = _value_text(value)
                if not text:
                    continue
                if _cell_matches(text, target_key):
                    row_has_match = True
                for model_part_number in _extract_models_from_text(text):
                    model_key = _match_key(model_part_number)
                    if model_key != target_key and model_key not in model_part_numbers:
                        model_part_numbers[model_key] = model_part_number
            if row_has_match:
                found_rows[row_number] = row_number

        return BomSearchData(list(found_rows.keys()), list(model_part_numbers.values()))
    finally:
        workbook.close()


def _read_bom_pandas(path, target_key):
    try:
        import pandas as pd
    except ImportError as exc:
        raise ServiceError(
            "File .xls membutuhkan pandas dan xlrd. Jalankan install dependency dari requirements.txt.",
            title="Dependency belum lengkap",
        ) from exc

    try:
        excel = pd.ExcelFile(path)
    except ImportError as exc:
        raise ServiceError(
            "File .xls membutuhkan library xlrd. Jalankan install dependency dari requirements.txt.",
            title="Dependency belum lengkap",
        ) from exc

    try:
        sheet_name = _find_sheet_case_insensitive(excel.sheet_names, TARGET_SHEET_NAME)
        if sheet_name is None:
            raise ServiceError('Sheet "BOM" tidak ditemukan.', title="Format Excel tidak valid")

        dataframe = pd.read_excel(excel, sheet_name=sheet_name, header=None, dtype=object, na_filter=False)
        found_rows = OrderedDict()
        model_part_numbers = OrderedDict()

        for dataframe_index, row in dataframe.iterrows():
            row_has_match = False
            for value in row.tolist():
                text = _value_text(value)
                if not text:
                    continue
                if _cell_matches(text, target_key):
                    row_has_match = True
                for model_part_number in _extract_models_from_text(text):
                    model_key = _match_key(model_part_number)
                    if model_key != target_key and model_key not in model_part_numbers:
                        model_part_numbers[model_key] = model_part_number
            if row_has_match:
                found_rows[int(dataframe_index) + 1] = int(dataframe_index) + 1

        return BomSearchData(list(found_rows.keys()), list(model_part_numbers.values()))
    finally:
        excel.close()


def _find_excel_files(source_folder):
    files = []
    errors = []

    def on_error(error):
        errors.append(f"{getattr(error, 'filename', '')}: {getattr(error, 'strerror', error)}")

    for current_folder, dir_names, file_names in os.walk(source_folder, onerror=on_error):
        dir_names.sort(key=lambda name: name.upper())
        for file_name in sorted(file_names, key=lambda name: name.upper()):
            if file_name.startswith("~$"):
                continue
            path = Path(current_folder) / file_name
            if path.suffix.lower() in EXCEL_EXTENSIONS:
                files.append(path)

    return files, errors


def _find_sheet_case_insensitive(sheet_names, target_name):
    target = target_name.lower()
    for sheet_name in sheet_names:
        if sheet_name.lower() == target:
            return sheet_name
    return None


def _cell_matches(value, target_key):
    text = _match_key(value)
    if not text:
        return False
    if text == target_key:
        return True
    return re.search(rf"(?<![A-Z0-9_.-]){re.escape(target_key)}(?![A-Z0-9_.-])", text) is not None


def _extract_models_from_text(value):
    text = _value_text(value)
    if not text:
        return []
    return _unique_text(match.group(0).upper() for match in MODEL_PATTERN.finditer(text))


def _split_part_tokens(value):
    tokens = []
    for token in re.split(r"[+;,|\r\n]+", str(value or "")):
        clean = _value_text(token).strip("()[]{}")
        if clean:
            tokens.append(clean.upper())
    return tokens


def _unique_text(values):
    unique = OrderedDict()
    for value in values:
        clean = _value_text(value)
        if not clean:
            continue
        key = _match_key(clean)
        if key and key not in unique:
            unique[key] = clean.upper()
    return list(unique.values())


def _looks_like_revision(value):
    text = _value_text(value)
    return bool(re.fullmatch(r"[A-Za-z]?\d+(?:[._-]\d+)*", text))


def _match_key(value):
    text = _value_text(value).upper()
    return re.sub(r"\s+", " ", text).strip()


def _value_text(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
    text = str(value)
    text = text.replace("\r", " ").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def _validate_config(config):
    if not _value_text(config.component_part_number):
        raise ServiceError("Component Part Number belum diisi.", title="Input belum lengkap")
    if not config.source_folder:
        raise ServiceError("Folder Induk PCB belum dipilih.", title="Input belum lengkap")
    if not Path(config.source_folder).is_dir():
        raise ServiceError(f"Folder Induk PCB tidak ditemukan:\n{config.source_folder}", title="Folder tidak ditemukan")


def _normalize_output_path(path):
    output_path = Path(path)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")
    return output_path


def _emit_progress(progress_callback, percent, message):
    if progress_callback:
        progress_callback(percent, message)


def _error_message(exc):
    return getattr(exc, "message", str(exc))
