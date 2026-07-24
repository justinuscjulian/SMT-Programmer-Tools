import re
import shlex
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from services.errors import ServiceError
from utils.encoding import read_lines_with_fallback
from utils.paths import resource_path
from utils.sort import natural_sort_key


OUTPUT_HEADERS = ["Table", "Slot", "Position", "Location Code", "Part Number"]
SUMMARY_HEADERS = [
    "Part Number",
    "Feeder Paling Sering",
    "Jumlah File di Feeder Itu",
    "Total Muncul",
    "Feeder Lain",
    "Notes",
]
DEFAULT_BALANCING_PART_NUMBERS = [
    "0RJ1002C678",
    "0CK104BF56A",
    "0RJ0000C678",
    "0CK104BH56A",
    "0RJ1000C678",
    "EAE52158501",
    "EAE66302101",
]
FIXED_FEEDER_SIDES = "ABCDEFGHIJ"
VALID_NPM_EXPORT_SUFFIXES = {".txt", ".crb"}
CM602_FEEDER_SIDES = "ABCDEFGHIJ"
CM602_FEEDER_FIX_HEADER = [
    "MachineName",
    "BeamNo",
    "PU",
    "TrayFlag",
    "FeederA",
    "FeederB",
    "FeederC",
    "FeederD",
    "FeederE",
    "FeederF",
    "FeederG",
    "FeederH",
    "FeederI",
    "FeederJ",
    "PartsA",
    "PartsA_EX",
    "PartsB",
    "PartsB_EX",
    "PartsC",
    "PartsC_EX",
    "PartsD",
    "PartsD_EX",
    "PartsE",
    "PartsE_EX",
    "PartsF",
    "PartsF_EX",
    "PartsG",
    "PartsG_EX",
    "PartsH",
    "PartsH_EX",
    "PartsI",
    "PartsI_EX",
    "PartsJ",
    "PartsJ_EX",
]
CM602_FEEDER_PROTECT_HEADER = ["MachineName", "BeamNo", *[f"PU{index:02d}" for index in range(1, 71)]]
CM602_EXACT_FEEDER_IDS = {
    "0CC300BKF1A": "2481",
    "0CC475CD5DA": "2482",
    "0TRKE80046A": "2502",
    "EAE30281301": "2481",
    "EAE32166101": "2481",
    "EAE39487601": "2481",
    "EAE52158501": "2502",
    "EAE58111501": "2502",
    "EAE58399101": "2481",
    "EAE60663801": "2502",
    "EAE61081701": "2502",
    "EAE62161801": "2482",
    "EAE62303301": "2482",
    "EAF61530102": "2481",
    "EAF61630102": "2481",
    "EAH33945901": "2482",
    "EAH61714301": "2502",
    "EAH61872501": "2502",
    "EAH62134701": "2502",
    "EAM32500208": "2481",
    "EAM38769505": "2482",
    "EAM61130602": "2482",
    "EAM62091201": "2482",
    "EAN62794401": "2502",
    "EAN62871401": "2502",
    "EBC00892202": "2482",
    "EBC41607701": "2481",
    "EBC62582701": "2481",
    "EBC62582702": "2481",
    "EBC62582703": "2481",
    "EBC62582705": "2481",
    "EBC62582706": "2481",
    "EBC62582707": "2481",
    "EBC62582708": "2481",
    "EBC62582709": "2481",
    "EBC62798406": "2481",
    "EBK61913702": "2502",
    "ERHY0000124": "2481",
    "ERHZ0000219": "2481",
}


@dataclass
class FeederMappingResult:
    records: list
    source_file: str
    row_count: int
    table_count: int
    part_count: int


@dataclass
class FeederMappingBatchResult:
    mappings: list
    summary_records: list
    raw_summary_records: list
    source_count: int
    row_count: int
    part_count: int
    output_path: str = ""


class NpmFeederImportResult(dict):
    def __init__(
        self,
        output_path: str = "",
        mapping_file: str = "",
        template_file: str = "",
        mapping_row_count: int = 0,
        assigned_part_count: int = 0,
        assignment_count: int = 0,
        missing_part_rows: list = None,
        missing_location_rows: list = None,
        missing_feeder_rows: list = None,
        conflict_rows: list = None,
        duplicate_rows: list = None,
    ):
        data = {
            "output_path": output_path,
            "mapping_file": mapping_file,
            "template_file": template_file,
            "mapping_row_count": mapping_row_count,
            "assigned_part_count": assigned_part_count,
            "assignment_count": assignment_count,
            "missing_part_rows": missing_part_rows if missing_part_rows is not None else [],
            "missing_location_rows": missing_location_rows if missing_location_rows is not None else [],
            "missing_feeder_rows": missing_feeder_rows if missing_feeder_rows is not None else [],
            "conflict_rows": conflict_rows if conflict_rows is not None else [],
            "duplicate_rows": duplicate_rows if duplicate_rows is not None else [],
        }
        super().__init__(data)

    def __getattr__(self, name):
        try:
            return self[name]
        except KeyError:
            raise AttributeError(f"'NpmFeederImportResult' object has no attribute '{name}'")

    def __setattr__(self, name, value):
        self[name] = value


@dataclass
class NpmFeederImportBatchResult:
    output_dir: str
    mapping_file: str
    group_results: list
    total_groups: int
    successful_groups: int



@dataclass
class Cm602ProgramCmTxtResult:
    output_path: str
    source_file: str
    row_count: int
    part_count: int
    board_x: float
    board_y: float


@dataclass
class Cm602FeederFixImportResult:
    output_path: str
    mapping_file: str
    mapping_row_count: int
    assignment_count: int
    slot_count: int
    part_count: int
    duplicate_rows: list
    default_feeder_rows: list


def suggest_output_name(source_path):
    stem = _clean_filename_part(Path(source_path or "Feeder_Mapping").stem) or "Feeder_Mapping"
    return f"Feeder_Mapping_Result_{stem}.xlsx"


def suggest_multiple_output_name(source_paths=None):
    paths = list(source_paths or [])
    if len(paths) == 1:
        stem = _clean_filename_part(Path(paths[0]).stem) or "Multiple"
        return f"Feeder_Mapping_Multiple_{stem}.xlsx"
    return "Feeder_Mapping_Multiple_Result.xlsx"


def suggest_cm602_output_name(source_path):
    stem = _clean_filename_part(Path(source_path or "CM602_Feeder_Mapping").stem) or "CM602_Feeder_Mapping"
    return f"CM602_Feeder_Mapping_Result_{stem}.xlsx"


def suggest_cm602_program_cm_txt_output_name(source_path=None):
    return "CM.txt"


def suggest_cm602_feeder_fix_output_name(mapping_path):
    stem = _clean_filename_part(Path(mapping_path or "CM602_FeederFix").stem) or "CM602_FeederFix"
    return f"{stem}_FIX.txt"


def suggest_npm_import_output_name(mapping_path):
    stem = _clean_filename_part(Path(mapping_path or "Feeder_Mapping").stem) or "Feeder_Mapping"
    return f"{stem}_NPM_Feeder_Import.txt"


def default_balancing_part_numbers_text():
    return ""


def load_feeder_mapping(file_path):
    path = _clean_path(file_path)
    if not Path(path).is_file():
        raise ServiceError(f"File mesin NPM tidak ditemukan:\n{path}", title="File tidak ditemukan")

    try:
        lines, _ = read_lines_with_fallback(path)
    except Exception as exc:
        raise ServiceError("File mesin NPM tidak bisa dibaca.", title="Encoding error") from exc

    parts_rows = _read_first_available_section_rows(lines, ("PartsData", "PartsDataEx"))
    feeder_rows = _read_section_rows(lines, "FeederData")
    fixed_rows = _read_section_rows(lines, "FixedFeeder")

    part_lookup = _build_lookup(parts_rows, "IDNUM")
    feeder_lookup = _build_lookup(feeder_rows, "IDNUM")
    records = _build_mapping_records(fixed_rows, part_lookup, feeder_lookup)

    if not records:
        raise ServiceError("Tidak ada Fixed Feeder aktif yang bisa dikonversi.", title="Data kosong")

    table_count = len({record["table"] for record in records})
    part_count = len({record["part_number"] for record in records})
    return FeederMappingResult(
        records=records,
        source_file=Path(path).name,
        row_count=len(records),
        table_count=table_count,
        part_count=part_count,
    )


def generate_feeder_mapping_excel(source_path, output_path):
    result = load_feeder_mapping(source_path)
    exported_path = export_feeder_mapping(result.records, output_path)
    return result, exported_path


def load_cm602_feeder_mapping(file_path):
    path = _clean_path(file_path)
    if not Path(path).is_file():
        raise ServiceError(f"File CM602 tidak ditemukan:\n{path}", title="File tidak ditemukan")

    try:
        lines, _ = read_lines_with_fallback(path)
    except Exception as exc:
        raise ServiceError("File CM602 tidak bisa dibaca.", title="Encoding error") from exc

    if _has_section(lines, "FeederFix"):
        feeder_fix_rows = _read_section_rows(lines, "FeederFix", machine_label="CM602")
        records = _build_cm602_feeder_fix_records(feeder_fix_rows)
    elif _has_section(lines, "StockData") and _has_section(lines, "PartsData"):
        stock_rows = _read_section_rows(lines, "StockData", machine_label="CM602")
        part_lookup = _build_lookup(_read_section_rows(lines, "PartsData", machine_label="CM602"), "IDNUM")
        records = _build_cm602_stock_records(stock_rows, part_lookup)
    else:
        raise ServiceError(
            "File CM602 harus berisi section [FeederFix] atau kombinasi [StockData] dan [PartsData].",
            title="Format CM602 tidak valid",
        )

    if not records:
        raise ServiceError("Tidak ada feeder CM602 aktif yang bisa dikonversi.", title="Data kosong")

    table_count = len({record["table"] for record in records})
    part_count = len({record["part_number"] for record in records})
    return FeederMappingResult(
        records=records,
        source_file=Path(path).name,
        row_count=len(records),
        table_count=table_count,
        part_count=part_count,
    )


def generate_cm602_feeder_mapping_excel(source_path, output_path):
    result = load_cm602_feeder_mapping(source_path)
    exported_path = export_feeder_mapping(result.records, output_path)
    return result, exported_path


def generate_cm602_program_cm_txt(source_path, output_path):
    source = Path(_clean_path(source_path))
    if not source.is_file():
        raise ServiceError(f"File program CM602 tidak ditemukan:\n{source}", title="File tidak ditemukan")

    output = Path(output_path)
    if output.suffix.lower() != ".txt":
        output = output.with_suffix(".txt")
    output.parent.mkdir(parents=True, exist_ok=True)

    try:
        lines, _ = read_lines_with_fallback(source)
    except Exception as exc:
        raise ServiceError("File program CM602 tidak bisa dibaca.", title="Encoding error") from exc

    board_x, board_y = _read_cm602_board_size(lines)
    block_rows = _read_section_rows(lines, "BlockData", machine_label="CM602")
    part_lookup = _build_lookup(_read_section_rows(lines, "PartsData", machine_label="CM602"), "IDNUM")
    output_lines = _build_cm602_program_cm_txt_lines(block_rows, part_lookup, board_x, board_y)

    if not output_lines:
        raise ServiceError("Tidak ada data mounting di section [BlockData].", title="Data kosong")

    with output.open("w", encoding="utf-8", newline="") as handle:
        handle.write("\r\n".join(output_lines))
        handle.write("\r\n")

    part_count = len({_cm602_part_name(part_lookup.get(str(row.get("PARTS", "")).strip())) for row in block_rows if str(row.get("PARTS", "")).strip() in part_lookup})
    return Cm602ProgramCmTxtResult(
        output_path=str(output),
        source_file=source.name,
        row_count=len(output_lines),
        part_count=part_count,
        board_x=board_x,
        board_y=board_y,
    )


def generate_cm602_feeder_fix_import_file(mapping_path, output_path):
    mapping_file = Path(_clean_path(mapping_path))
    if not mapping_file.is_file():
        raise ServiceError(f"File Excel CM602 feeder mapping tidak ditemukan:\n{mapping_file}", title="File tidak ditemukan")
    if mapping_file.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ServiceError("Input CM602 feeder mapping harus file Excel .xlsx/.xlsm.", title="Format tidak valid")

    output = Path(output_path)
    if output.suffix.lower() != ".txt":
        output = output.with_suffix(".txt")
    output.parent.mkdir(parents=True, exist_ok=True)

    mapping_records, duplicate_rows = _load_cm602_feeder_fix_mapping_workbook(mapping_file)
    output_lines, default_feeder_rows = _build_cm602_feeder_fix_import_lines(mapping_records)
    with output.open("w", encoding="utf-8", newline="") as handle:
        handle.write("\r\n".join(output_lines))
        handle.write("\r\n")

    return Cm602FeederFixImportResult(
        output_path=str(output),
        mapping_file=mapping_file.name,
        mapping_row_count=len(mapping_records),
        assignment_count=len(mapping_records),
        slot_count=len(_cm602_feeder_fix_slots(mapping_records)),
        part_count=len({_part_key(record["part_number"]) for record in mapping_records}),
        duplicate_rows=duplicate_rows,
        default_feeder_rows=default_feeder_rows,
    )


def load_multiple_feeder_mappings(source_paths):
    paths = _clean_source_paths(source_paths)
    if not paths:
        raise ServiceError("Belum ada file feeder/program yang dipilih.", title="Input belum lengkap")

    mappings = [_load_feeder_mapping_auto(path) for path in paths]
    summary_records, raw_summary_records = _build_summary_records(mappings)
    return FeederMappingBatchResult(
        mappings=mappings,
        summary_records=summary_records,
        raw_summary_records=raw_summary_records,
        source_count=len(mappings),
        row_count=sum(mapping.row_count for mapping in mappings),
        part_count=len(summary_records),
    )


def generate_multiple_feeder_mapping_excel(source_paths, output_path):
    result = load_multiple_feeder_mappings(source_paths)
    result.output_path = export_multiple_feeder_mapping(result, output_path)
    return result


def _load_feeder_mapping_auto(file_path):
    try:
        return load_feeder_mapping(file_path)
    except Exception as npm_exc:
        try:
            return load_cm602_feeder_mapping(file_path)
        except Exception as cm602_exc:
            npm_message = _error_message(npm_exc)
            cm602_message = _error_message(cm602_exc)
            raise ServiceError(
                (
                    f"File tidak cocok sebagai NPM export ataupun CM602 program/feeder file:\n"
                    f"{file_path}\n\n"
                    f"NPM: {npm_message}\n"
                    f"CM602: {cm602_message}"
                ),
                title="Format feeder tidak valid",
            ) from cm602_exc


def generate_npm_feeder_import_file(mapping_path, template_path, output_path):
    mapping_file = Path(_clean_path(mapping_path))
    if not mapping_file.is_file():
        raise ServiceError(f"File Excel feeder mapping tidak ditemukan:\n{mapping_file}", title="File tidak ditemukan")
    if mapping_file.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ServiceError("Input feeder mapping harus file Excel .xlsx/.xlsm.", title="Format tidak valid")

    if template_path:
        template = Path(_clean_path(template_path))
        if not template.is_file():
            raise ServiceError(f"Template program NPM tidak ditemukan:\n{template}", title="File tidak ditemukan")
        if template.suffix.lower() not in {".txt", ".crb"}:
            raise ServiceError("Template NPM harus file .txt atau .crb.", title="Format tidak valid")

    output = Path(output_path)
    if output.suffix.lower() != ".txt":
        output = output.with_suffix(".txt")
    output.parent.mkdir(parents=True, exist_ok=True)

    mapping_records, duplicate_rows = _load_import_mapping_workbook(mapping_file)

    return _generate_npm_import_from_records(
        mapping_records=mapping_records,
        duplicate_rows=duplicate_rows,
        mapping_file_name=mapping_file.name,
        template_path=template_path,
        output_path=output,
    )


def generate_npm_feeder_import_batch_from_groups(mapping_path, template_path, output_dir_path):
    mapping_file = Path(_clean_path(mapping_path))
    if not mapping_file.is_file():
        raise ServiceError(f"File Excel feeder mapping tidak ditemukan:\n{mapping_file}", title="File tidak ditemukan")
    if mapping_file.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ServiceError("Input feeder mapping harus file Excel .xlsx/.xlsm.", title="Format tidak valid")

    if template_path:
        template = Path(_clean_path(template_path))
        if not template.is_file():
            raise ServiceError(f"Template program NPM tidak ditemukan:\n{template}", title="File tidak ditemukan")
        if template.suffix.lower() not in {".txt", ".crb"}:
            raise ServiceError("Template NPM harus file .txt atau .crb.", title="Format tidak valid")

    output_dir = Path(output_dir_path)
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        workbook = load_workbook(mapping_file, data_only=True, read_only=True)
    except Exception as exc:
        raise ServiceError("File Excel feeder mapping tidak bisa dibaca.", title="Excel tidak valid") from exc

    # Parse Summary sheet to get PCB names mapping for each group
    group_pcbs = defaultdict(list)
    if "Summary" in workbook.sheetnames:
        ws_sum = workbook["Summary"]
        header = None
        group_idx = -1
        pcb_idx = -1
        for row in ws_sum.iter_rows(values_only=True):
            vals = [_clean_excel_cell(c) for c in row]
            if not any(vals):
                continue
            if header is None:
                header_upper = [v.upper() for v in vals]
                if "GROUP NAME" in header_upper and "PCB NAME" in header_upper:
                    header = header_upper
                    group_idx = header_upper.index("GROUP NAME")
                    pcb_idx = header_upper.index("PCB NAME")
            else:
                if group_idx >= 0 and pcb_idx >= 0 and len(vals) > max(group_idx, pcb_idx):
                    g_name = vals[group_idx]
                    p_name = vals[pcb_idx]
                    if g_name and p_name:
                        if p_name not in group_pcbs[g_name]:
                            group_pcbs[g_name].append(p_name)

    group_results = []

    for worksheet in workbook.worksheets:
        if worksheet.title.lower() in {"summary", "summary sheet"}:
            continue

        records, duplicates = _read_import_mapping_sheet(worksheet)
        if not records:
            continue

        sheet_title = worksheet.title
        pcbs = group_pcbs.get(sheet_title, [])
        if pcbs:
            raw_filename = ", ".join(pcbs)
        else:
            raw_filename = sheet_title

        clean_filename = re.sub(r'[\\/*?:"<>|]', '_', raw_filename).strip()
        if not clean_filename:
            clean_filename = sheet_title

        out_txt_path = output_dir / f"{clean_filename}.txt"

        res = _generate_npm_import_from_records(
            mapping_records=records,
            duplicate_rows=duplicates,
            mapping_file_name=mapping_file.name,
            template_path=template_path,
            output_path=out_txt_path,
        )
        group_results.append(res)

    workbook.close()

    if not group_results:
        raise ServiceError(
            "Tidak ditemukan sheet group fix feeder yang memiliki data slot dan part number valid.",
            title="Data tidak ditemukan",
        )

    return NpmFeederImportBatchResult(
        output_dir=str(output_dir),
        mapping_file=mapping_file.name,
        group_results=group_results,
        total_groups=len(group_results),
        successful_groups=len(group_results),
    )


def _generate_npm_import_from_records(mapping_records, duplicate_rows, mapping_file_name, template_path, output_path):
    output = Path(output_path)
    if output.suffix.lower() != ".txt":
        output = output.with_suffix(".txt")
    output.parent.mkdir(parents=True, exist_ok=True)

    if not template_path:
        template = _select_auto_template(mapping_records)
    else:
        template = Path(_clean_path(template_path))

    template_lines, encoding = read_lines_with_fallback(template)
    fixed_header = _fixed_feeder_header(template_lines)
    fixed_rows = _read_section_rows(template_lines, "FixedFeeder")
    part_rows = _read_first_available_section_rows(template_lines, ("PartsData", "PartsDataEx"))
    part_lookup = _build_part_lookup_by_name(part_rows)
    part_lookup_by_id = _build_lookup(part_rows, "IDNUM")
    feeder_lookup = _build_lookup(_read_section_rows(template_lines, "FeederData"), "IDNUM")
    template_assignments = _existing_fixed_assignments(fixed_rows, part_lookup_by_id, feeder_lookup)

    is_auto = not bool(template_path)
    if is_auto:
        max_idnum = max([int(float(str(r.get("IDNUM", 0)))) for r in part_rows if str(r.get("IDNUM", "")).replace(".", "", 1).isdigit()], default=0)
        for record in mapping_records:
            part_key = _part_key(record["part_number"])
            if part_key not in part_lookup:
                max_idnum += 1
                new_row = {
                    "IDNUM": str(max_idnum),
                    "NAME": f'"{record["part_number"]}"',
                    "LNAME": '"ohm"',
                    "REELS": "1",
                    "SKIP": "0",
                    "NoAutoDivide": "0",
                    "Alt": "0",
                    "AltNum": "0",
                    "NoArrange": "0",
                }
                part_rows.append(new_row)
                part_lookup[part_key] = new_row

    new_fixed_rows = [_empty_fixed_import_row(row) for row in fixed_rows]
    fixed_by_pu = {str(row.get("PU", "")).strip(): row for row in new_fixed_rows}

    assigned_part_keys = set()
    assignment_count = 0
    missing_part_rows = []
    missing_location_rows = []
    missing_feeder_rows = []
    conflict_rows = []
    occupied = {}
    blocked_pus = set()

    for record in sorted(mapping_records, key=_import_mapping_priority):
        part_number = record["part_number"]
        part_key = _part_key(part_number)
        location_code = record["location_code"]
        part_row = part_lookup.get(part_key)
        if part_row is None:
            missing_part_rows.append(f"Row {record['row_number']}: {part_number}")
            continue

        fixed_row = fixed_by_pu.get(str(record["pu"]))
        if fixed_row is None:
            max_idnum = max([int(float(str(r.get("IDNUM", 0)))) for r in new_fixed_rows if str(r.get("IDNUM", "")).replace(".", "", 1).isdigit()], default=0)
            base_row = {col: "0" for col in fixed_header}
            base_row["IDNUM"] = str(max_idnum + 1)
            base_row["Group"] = "0"
            base_row["PU"] = str(record["pu"])
            fixed_row = _empty_fixed_import_row(base_row)
            new_fixed_rows.append(fixed_row)
            fixed_by_pu[str(record["pu"])] = fixed_row

        conflict = _first_import_location_conflict(record, occupied)
        if conflict:
            conflict_rows.append(f"Row {record['row_number']}: {part_number} @ {location_code} nabrak dengan {conflict}")
            continue

        feeder_id = _feeder_id_for_import_location(
            part_key,
            part_row,
            feeder_lookup,
            record["uses_lr_position"],
            template_assignments,
            is_auto,
            spans_slots=record.get("spans_slots", 1),
        )
        if not feeder_id:
            label = "compact L/R" if record["uses_lr_position"] else "non L/R"
            missing_feeder_rows.append(f"Row {record['row_number']}: {part_number} tidak punya feeder {label}")
            continue

        _set_fixed_import_assignment(
            fixed_row,
            record["side"],
            feeder_id,
            part_row["IDNUM"],
            record["uses_lr_position"],
        )
        _occupy_import_location(record, occupied)
        assigned_part_keys.add(part_key)
        assignment_count += 1

        spans_slots = record.get("spans_slots", 1)
        if spans_slots > 1:
            for s in range(1, spans_slots):
                blocked_pu = record["pu"] + s
                blocked_pus.add(blocked_pu)

    # Filter out rows that are blocked by multi-slot feeders
    new_fixed_rows = [row for row in new_fixed_rows if int(float(str(row.get("PU", 0)))) not in blocked_pus]
    new_fixed_rows.sort(key=lambda r: int(float(str(r.get("PU", 0)))))
    output_lines = _replace_feeder_import_sections(template_lines, fixed_header, new_fixed_rows, part_rows if is_auto else None)
    with output.open("w", encoding=encoding, newline="") as handle:
        handle.writelines(output_lines)

    return NpmFeederImportResult(
        output_path=str(output),
        mapping_file=mapping_file_name,
        template_file=template.name,
        mapping_row_count=len(mapping_records),
        assigned_part_count=len(assigned_part_keys),
        assignment_count=assignment_count,
        missing_part_rows=missing_part_rows,
        missing_location_rows=missing_location_rows,
        missing_feeder_rows=missing_feeder_rows,
        conflict_rows=conflict_rows,
        duplicate_rows=duplicate_rows,
    )


def _load_import_mapping_workbook(mapping_file):
    try:
        workbook = load_workbook(mapping_file, data_only=True, read_only=True)
    except Exception as exc:
        raise ServiceError("File Excel feeder mapping tidak bisa dibaca.", title="Excel tidak valid") from exc

    best_records = []
    best_duplicates = []
    best_sheet = ""
    for worksheet in workbook.worksheets:
        records, duplicates = _read_import_mapping_sheet(worksheet)
        if len(records) > len(best_records):
            best_records = records
            best_duplicates = duplicates
            best_sheet = worksheet.title

    workbook.close()
    if not best_records:
        raise ServiceError(
            "Excel feeder mapping harus berisi minimal dua kolom: Part Number dan Location Code.",
            title="Data kosong",
        )

    for record in best_records:
        record["sheet_name"] = best_sheet
    return best_records, best_duplicates


def _load_cm602_feeder_fix_mapping_workbook(mapping_file):
    try:
        workbook = load_workbook(mapping_file, data_only=True, read_only=True)
    except Exception as exc:
        raise ServiceError("File Excel CM602 feeder mapping tidak bisa dibaca.", title="Excel tidak valid") from exc

    best_records = []
    best_duplicates = []
    best_sheet = ""
    for worksheet in workbook.worksheets:
        records, duplicates = _read_cm602_feeder_fix_mapping_sheet(worksheet)
        if len(records) > len(best_records):
            best_records = records
            best_duplicates = duplicates
            best_sheet = worksheet.title

    workbook.close()
    if not best_records:
        raise ServiceError(
            "Excel CM602 harus berisi minimal dua kolom: Location Code dan Part Number, contoh [1]7L.",
            title="Data kosong",
        )

    for record in best_records:
        record["sheet_name"] = best_sheet
    return best_records, best_duplicates


def _read_cm602_feeder_fix_mapping_sheet(worksheet):
    records = []
    duplicates = []
    seen_locations = {}
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = [_clean_excel_cell(value) for value in row[:4]]
        while len(values) < 4:
            values.append("")
        first, second, third, fourth = values
        if not any(values):
            continue
        if _looks_like_import_header(first, second):
            continue

        part_number, location_code = first, second
        if _parse_import_location_code(first) and not _parse_import_location_code(second):
            part_number, location_code = second, first

        parsed_location = _parse_import_location_code(location_code)
        if not part_number or not parsed_location:
            continue

        feeder_id = _first_cm602_feeder_id_value(third, fourth)
        location_key = parsed_location["normalized_location"].upper()
        existing = seen_locations.get(location_key)
        if existing:
            duplicates.append(f"Row {row_number}: {part_number} @ {location_code} duplicate dengan row {existing}")
            continue
        seen_locations[location_key] = row_number

        records.append(
            {
                "row_number": row_number,
                "part_number": part_number,
                "feeder_id": feeder_id,
                "location_code": parsed_location["normalized_location"],
                **parsed_location,
            }
        )

    records.sort(key=_import_mapping_priority)
    return records, duplicates


def _first_cm602_feeder_id_value(*values):
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        if re.fullmatch(r"\d+(?:\.0+)?", text):
            return str(int(float(text)))
    return ""


def _cm602_feeder_fix_slots(records):
    return {(record["table"], record["slot"]) for record in records}


def _build_cm602_feeder_fix_import_lines(mapping_records):
    rows_by_slot = {}
    default_feeder_rows = []

    for record in sorted(mapping_records, key=_import_mapping_priority):
        slot_key = (record["table"], record["slot"])
        slot = rows_by_slot.setdefault(slot_key, {})
        side = record["side"]
        part_number = record["part_number"]
        feeder_id = str(record.get("feeder_id", "")).strip()
        if not feeder_id:
            feeder_id, is_default = _infer_cm602_feeder_id(part_number)
            if is_default:
                default_feeder_rows.append(f"Row {record['row_number']}: {part_number} @ {record['location_code']} pakai default feeder {feeder_id}")

        slot[side] = {
            "part_number": part_number,
            "feeder_id": feeder_id,
        }

    output_lines = [
        "[FeederFix]",
        f"{' '.join(CM602_FEEDER_FIX_HEADER)} ",
    ]

    for table, slot in sorted(rows_by_slot, key=lambda item: (item[0], item[1])):
        output_lines.append(_format_cm602_feeder_fix_row(table, slot, rows_by_slot[(table, slot)]))

    output_lines.extend(
        [
            "",
            "[PaletteData]",
            "IDNUM N X Y ROT ",
            "",
            "[FeederProtect]",
            f"{' '.join(CM602_FEEDER_PROTECT_HEADER)} ",
            "",
        ]
    )
    return output_lines, default_feeder_rows


def _format_cm602_feeder_fix_row(table, slot, assignments):
    values = [_quote_cm602_text("CM602-1"), str(table), str(slot), "0"]
    for side in CM602_FEEDER_SIDES:
        assignment = assignments.get(side, {})
        values.append(str(assignment.get("feeder_id", "0") or "0"))

    for side in CM602_FEEDER_SIDES:
        assignment = assignments.get(side, {})
        part_number = str(assignment.get("part_number", "") or "").strip()
        values.append(_quote_cm602_text(part_number) if part_number else '""')
        values.append("0")
    return " ".join(values)


def _quote_cm602_text(value):
    text = str(value or "").replace('"', '\\"')
    return f'"{text}"'


def _infer_cm602_feeder_id(part_number):
    key = _part_key(part_number)
    if not key:
        return "2481", True
    if key in CM602_EXACT_FEEDER_IDS:
        return CM602_EXACT_FEEDER_IDS[key], False

    if key.startswith("0TRK"):
        return "2502", False
    if key.startswith("0CZZ"):
        return "2481", False
    if key.startswith("0RH"):
        return "2482", False
    if key.startswith("0RJ"):
        if re.match(r"^0RJ\d+D", key):
            return "2482", False
        return "2481", False
    if key.startswith("0R"):
        return "2481", False

    ceramic_match = re.match(r"^0C[CK]\d+([A-Z]{2})", key)
    if ceramic_match:
        package_code = ceramic_match.group(1)
        if package_code in {"DC", "DD", "DK"}:
            return "2502", False
        if package_code in {"CC", "CD", "CK"}:
            return "2482", False
        return "2481", False

    if key.startswith(("EAN", "EBK")):
        return "2502", False
    if key.startswith("EAH"):
        return "2502", False
    if key.startswith("EAM"):
        return "2482", False
    if key.startswith(("EAF", "EBC", "ERH")):
        return "2481", False
    if key.startswith("EAE"):
        return "2481", False

    return "2481", True


def _read_import_mapping_sheet(worksheet):
    records = []
    duplicates = []
    seen_exact_rows = set()
    header_map = {}
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        raw_cells = [_clean_excel_cell(c) for c in row]
        while raw_cells and not raw_cells[-1]:
            raw_cells.pop()
        if not any(raw_cells):
            continue

        cells_upper = [c.upper() for c in raw_cells]
        if "PART NUMBER" in cells_upper or "LOCATION CODE" in cells_upper or "LOCATION" in cells_upper:
            for idx, c in enumerate(cells_upper):
                if "PART NUMBER" in c or "PART" in c:
                    header_map["part"] = idx
                elif "LOCATION" in c or "SLOT" in c:
                    header_map["loc"] = idx
            continue

        part_number = ""
        location_code = ""

        if "loc" in header_map and "part" in header_map:
            if len(raw_cells) > max(header_map["loc"], header_map["part"]):
                c_loc = raw_cells[header_map["loc"]]
                c_part = raw_cells[header_map["part"]]
                if _parse_import_location_code(c_loc):
                    location_code = c_loc
                    part_number = c_part

        if not location_code or not part_number:
            loc_idx = -1
            for idx, c in enumerate(raw_cells):
                if _parse_import_location_code(c):
                    location_code = c
                    loc_idx = idx
                    break
            if location_code and loc_idx >= 0:
                for idx, c in enumerate(raw_cells):
                    if (
                        idx != loc_idx
                        and c
                        and c.upper() not in {"FIXED", "DYNAMIC", "L", "R", "LEFT", "RIGHT", "SKIPPED"}
                        and not _parse_import_location_code(c)
                    ):
                        part_number = c
                        if not c.isdigit():
                            break

        parsed_location = _parse_import_location_code(location_code)
        if not part_number or not parsed_location:
            continue

        exact_key = (_part_key(part_number), parsed_location["normalized_location"])
        if exact_key in seen_exact_rows:
            duplicates.append(f"Row {row_number}: {part_number} @ {location_code}")
            continue
        seen_exact_rows.add(exact_key)

        records.append(
            {
                "row_number": row_number,
                "part_number": part_number,
                "location_code": parsed_location["normalized_location"],
                **parsed_location,
            }
        )
    return records, duplicates


def _clean_excel_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _looks_like_import_header(first, second):
    text = f"{first} {second}".strip().lower()
    return "part" in text and ("location" in text or "feeder" in text)


def _parse_import_location_code(location_code):
    match = re.match(r"^\[(\d+)\](\d+)(?:-(\d+))?[-_]?([LR])?$", str(location_code or "").strip(), flags=re.IGNORECASE)
    if not match:
        return None

    table = int(match.group(1))
    slot = int(match.group(2))
    end_slot = int(match.group(3)) if match.group(3) else slot
    spans_slots = end_slot - slot + 1
    suffix = str(match.group(4) or "").upper()
    side = "B" if suffix == "R" else "A"
    
    normalized_location = f"[{table}]{slot:02d}" if slot < 10 else f"[{table}]{slot}"
    if end_slot > slot:
        normalized_location += f"-{end_slot}"
    if suffix:
        normalized_location += f"-{suffix}" if "-" in str(location_code) else suffix

    return {
        "table": table,
        "slot": slot,
        "end_slot": end_slot,
        "pu": table * 10000 + slot,
        "pu_side": 2 if suffix == "R" else 1,
        "side": side,
        "side_char": suffix,
        "uses_lr_position": suffix in {"L", "R"},
        "normalized_location": normalized_location,
        "spans_slots": spans_slots,
    }


def _select_auto_template(mapping_records):
    has_large_slot_table_1_4 = False
    has_table_8 = False
    has_table_9_or_10 = False
    has_table_7_large = False

    for record in mapping_records:
        table = record.get("table", record.get("parsed_location", {}).get("table", 0))
        slot = record.get("slot", record.get("parsed_location", {}).get("slot", 0))
        spans = record.get("spans_slots", 1)
        uses_lr = record.get("uses_lr_position", False)

        if table in (1, 2, 3, 4) and slot > 17:
            has_large_slot_table_1_4 = True
        if table == 8:
            has_table_8 = True
        if table in (9, 10):
            has_table_9_or_10 = True
        if table == 7 and (spans >= 2 or not uses_lr):
            has_table_7_large = True

    if has_large_slot_table_1_4:
        template_name = "assets/npm_base_template_line67.txt"
    elif (has_table_8 or has_table_7_large) and not has_table_9_or_10:
        template_name = "assets/npm_base_template_line8.txt"
    else:
        template_name = "assets/npm_base_template.txt"

    return Path(resource_path(template_name))


def _import_mapping_priority(record):
    return (
        record["pu"],
        1 if record["uses_lr_position"] else 0,
        0 if record["side"] == "A" else 1,
        record["row_number"],
    )


def _build_part_lookup_by_name(parts_rows):
    lookup = {}
    for row in parts_rows:
        key = _part_key(row.get("NAME", ""))
        if key and key not in lookup:
            lookup[key] = row
    return lookup


def _feeder_id_for_import_location(part_key, part_row, feeder_lookup, uses_lr_position, template_assignments, is_auto=False, spans_slots=1):
    candidates = _part_feeder_candidates(part_row, feeder_lookup)
    if not candidates:
        candidates = _template_feeder_candidates(part_key, feeder_lookup, template_assignments)

    if not candidates and is_auto:
        inferred = _infer_npm_feeder_id(part_key, uses_lr_position=uses_lr_position, spans_slots=spans_slots)
        if inferred in feeder_lookup:
            candidates = [(inferred, feeder_lookup[inferred])]

    if uses_lr_position:
        for feeder_id, feeder_row in candidates:
            if _feeder_kind(feeder_row) == 2:
                return feeder_id
        return ""

    for feeder_id, feeder_row in candidates:
        if _feeder_kind(feeder_row) != 2:
            return feeder_id
    return ""

KNOWN_LINE8_PART_FEEDERS = {
    "6630V93270J": "306825",
    "6630V93270K": "306825",
    "6630VK19605": "304425",
    "6212AB2015C": "304424",
    "EBJ30065001": "302983",
    "EBJ30065002": "302983",
    "EBJ30065003": "302983",
    "EBJ30065004": "302983",
    "EAG61090215": "306025",
    "EAG61090216": "306025",
    "EAG61090218": "306825",
    "EAG61090219": "306825",
    "EAG62571101": "305227",
    "EAG64089801": "304424",
    "EAG64089806": "304424",
    "EAG65010001": "306824",
    "EAG66129804": "305226",
    "EAG66129806": "305226",
    "EAG66129807": "305225",
    "EAG66129808": "305225",
    "EAG66129903": "305225",
    "EAG66129904": "305225",
    "EAG66736401": "305224",
    "EAG66765701": "306825",
    "EAG66854501": "304425",
    "EAG66854503": "304426",
    "EBF61874701": "304424",
    "EBF61874702": "304425",
    "MDS62110242": "304424",
    "EAG00473601": "306824",
    "EAG00473603": "306024",
    "EAG00474101": "304426",
    "EAG00474301": "305227",
    "EAG00474302": "305227",
    "EAG58732903": "304424",
    "EAG61008202": "305227",
    "EAG61008203": "304427",
    "EAG61030009": "304424",
    "EAG61030010": "304424",
    "EAG65031501": "302823",
    "EAP65276405": "304425",
    "EAP65297001": "304425",
}

def _infer_npm_feeder_id(part_number: str, uses_lr_position: bool = True, spans_slots: int = 1) -> str:
    # 1. Multi-slot Feeder Check
    if spans_slots == 2:
        return "304421"  # Double-slot tape feeder (16/24/32mm)
    if spans_slots >= 3:
        return "305221"  # Multi-slot tape feeder (44mm+)

    key = _part_key(part_number)  # clean & uppercase string

    if key in KNOWN_LINE8_PART_FEEDERS:
        return KNOWN_LINE8_PART_FEEDERS[key]

    # 2. Non-L/R Position (Single Lane 8mm Feeder)
    if not uses_lr_position:
        if not key:
            return "302001"
        if key.startswith(("EAN", "EAH", "EBK", "EAP", "EAV", "EAG", "MDS")):
            return "303623"
        if key.startswith(("EAM", "EAF", "EBC", "ERH", "EAE")):
            return "302983"
        return "302001"

    # 3. L/R Position (Double Lane 8mm Feeder)
    if not key:
        return "302481"
    if key.startswith("0TRK"):
        return "302502"
    if key.startswith("0CZZ"):
        return "302481"
    if key.startswith("0RH"):
        return "302482"
    if key.startswith("0RJ"):
        if re.match(r"^0RJ\d+D", key):
            return "302482"
        return "302481"
    if key.startswith("0R"):
        return "302481"

    # Ceramic Capacitor Matching via Regex
    ceramic_match = re.match(r"^0C[CK]\d+([A-Z]{2})", key)
    if ceramic_match:
        package_code = ceramic_match.group(1)
        if package_code in {"DK", "DC", "DD"}:
            return "302502"
        if package_code in {"CD", "CC", "CK"}:
            return "302482"
        return "302481"

    # Component Prefix Checks for Double Lane Feeder
    if key.startswith(("EAN", "EBK", "EAH")):
        return "302502"
    if key.startswith("EAM"):
        return "302482"
    if key.startswith(("EAF", "EBC", "ERH", "EAE")):
        return "302481"

    return "302481"  # Default fallback for Double Lane 8mm Feeder


def _template_feeder_candidates(part_key, feeder_lookup, template_assignments):
    candidates = []
    seen = set()
    for assignment in template_assignments.get(part_key, []):
        feeder_id = str(assignment.get("feeder_id", "")).strip()
        if feeder_id in seen or feeder_id not in feeder_lookup:
            continue
        seen.add(feeder_id)
        candidates.append((feeder_id, feeder_lookup[feeder_id]))
    return candidates


def _part_feeder_candidates(part_row, feeder_lookup):
    candidates = []
    for key in ("FA", "FB", "FC", "FD", "FE", "FF", "FG", "FH", "FI", "FJ"):
        feeder_id = str(part_row.get(key, "")).strip()
        if feeder_id in {"", "0", "-1"} or feeder_id not in feeder_lookup:
            continue
        candidates.append((feeder_id, feeder_lookup[feeder_id]))
    return candidates


def _empty_fixed_import_row(row):
    output = dict(row)
    for side in FIXED_FEEDER_SIDES:
        output[f"Feeder{side}"] = "-1" if side in {"A", "B"} else "0"
        output[f"Parts{side}"] = "0"
    return output


def _set_fixed_import_assignment(row, side, feeder_id, part_id, uses_lr_position):
    for extra_side in FIXED_FEEDER_SIDES[2:]:
        row[f"Feeder{extra_side}"] = "0"
        row[f"Parts{extra_side}"] = "0"

    if not uses_lr_position:
        row["FeederA"] = str(feeder_id)
        row["PartsA"] = str(part_id)
        row["FeederB"] = "0"
        row["PartsB"] = "0"
        return

    row[f"Feeder{side}"] = str(feeder_id)
    row[f"Parts{side}"] = str(part_id)


def _import_occupied_keys(record):
    pu_val = int(record["pu"])
    spans_slots = record.get("spans_slots", 1)
    keys = []
    for s in range(spans_slots):
        pu = str(pu_val + s)
        if record["uses_lr_position"]:
            keys.append((pu, record["side"]))
        else:
            keys.append((pu, "A"))
            keys.append((pu, "B"))
    return keys


def _first_import_location_conflict(record, occupied):
    for key in _import_occupied_keys(record):
        if key in occupied:
            return occupied[key]
    return ""


def _occupy_import_location(record, occupied):
    label = f"{record['part_number']} @ {record['location_code']}"
    for key in _import_occupied_keys(record):
        occupied[key] = label


def _replace_feeder_import_sections(lines, fixed_header, fixed_rows, part_rows=None):
    newline = _detect_newline(lines)
    output_lines = _replace_section_lines(
        lines,
        "FixedFeeder",
        _build_section_lines("FixedFeeder", fixed_header, fixed_rows, newline),
    )

    if part_rows is not None and _has_section(output_lines, "PartsDataEx"):
        part_header = _section_header(output_lines, "PartsDataEx")
        output_lines = _replace_section_lines(
            output_lines,
            "PartsDataEx",
            _build_section_lines("PartsDataEx", part_header, part_rows, newline),
        )

    if _has_section(output_lines, "StockData"):
        stock_header = _section_header(output_lines, "StockData")
        stock_rows = _build_stock_rows_from_fixed_rows(fixed_rows, stock_header)
        output_lines = _replace_section_lines(
            output_lines,
            "StockData",
            _build_section_lines("StockData", stock_header, stock_rows, newline),
        )
    return output_lines


def _build_stock_rows_from_fixed_rows(fixed_rows, stock_header):
    stock_rows = []
    for fixed_row in fixed_rows:
        if not _is_fixed_import_row_active(fixed_row):
            continue

        stock_row = {}
        for column in stock_header:
            if column == "IDNUM":
                stock_row[column] = str(len(stock_rows) + 1)
            elif column == "N":
                stock_row[column] = str(fixed_row.get("PU", ""))
            elif len(column) == 2 and column[0] == "P" and column[1] in FIXED_FEEDER_SIDES:
                stock_row[column] = _stock_value(fixed_row.get(f"Parts{column[1]}", "0"))
            elif len(column) == 2 and column[0] == "T" and column[1] in FIXED_FEEDER_SIDES:
                stock_row[column] = _stock_value(fixed_row.get(f"Feeder{column[1]}", "0"))
            elif column == "Force":
                stock_row[column] = "0"
            elif column == "LdFrmW":
                stock_row[column] = "0.000"
            elif column == "WaferRot":
                stock_row[column] = "0"
            else:
                stock_row[column] = ""
        stock_rows.append(stock_row)
    return stock_rows


def _is_fixed_import_row_active(row):
    for side in FIXED_FEEDER_SIDES:
        feeder_id = str(row.get(f"Feeder{side}", "")).strip()
        part_id = str(row.get(f"Parts{side}", "")).strip()
        if feeder_id not in {"", "0", "-1"} and part_id not in {"", "0", "-1"}:
            return True
    return False


def _stock_value(value):
    text = str(value or "").strip()
    return "0" if text in {"", "-1"} else text


def _build_section_lines(section_name, header, rows, newline):
    section_lines = [
        f"[{section_name}]" + newline,
        " ".join(header) + newline,
    ]
    for row in rows:
        section_lines.append(" ".join(str(row.get(column, "")) for column in header) + newline)
    return section_lines


def _replace_section_lines(lines, section_name, section_lines):
    start, end = _section_bounds(lines, section_name)
    return list(lines[:start]) + section_lines + list(lines[end:])


def export_feeder_mapping(records, output_path):
    output = Path(output_path)
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Detailed Feeder Setup"
    _append_mapping_sheet(worksheet, records)
    workbook.save(output)
    return str(output)


def export_multiple_feeder_mapping(result, output_path):
    output = Path(output_path)
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _append_summary_sheet(summary_sheet, result.summary_records)

    raw_summary_sheet = workbook.create_sheet("Summary (Raw)")
    _append_summary_sheet(raw_summary_sheet, result.raw_summary_records)

    used_titles = {"summary", "summary (raw)"}
    for mapping in result.mappings:
        worksheet = workbook.create_sheet(_unique_sheet_title(mapping.source_file, used_titles))
        _append_mapping_sheet(worksheet, mapping.records)

    workbook.save(output)
    return str(output)


def _build_mapping_records(fixed_rows, part_lookup, feeder_lookup):
    records = []
    for row in fixed_rows:
        table, slot = _decode_pickup_unit(row.get("PU", ""))
        if not table or not slot:
            continue

        feeder_a = row.get("FeederA", "")
        feeder_b = row.get("FeederB", "")
        part_a = row.get("PartsA", "")
        part_b = row.get("PartsB", "")
        active_a = _is_active(feeder_a, part_a, part_lookup)
        active_b = _is_active(feeder_b, part_b, part_lookup)
        if not active_a and not active_b:
            continue

        feeder_for_kind = feeder_a if active_a else feeder_b
        feeder_row = feeder_lookup.get(feeder_for_kind, {})
        kind = _feeder_kind(feeder_row)
        span = _feeder_slot_span(feeder_row)

        if kind == 2:
            if active_a:
                records.append(_record(table, slot, "L", f"[{table}]{slot}L", part_lookup[part_a]["NAME"], feeder_a))
            if active_b:
                records.append(_record(table, slot, "R", f"[{table}]{slot}R", part_lookup[part_b]["NAME"], feeder_b))
            continue

        if not active_a:
            records.append(_record(table, slot, f"Kind {kind or 'Unknown'} R", f"[{table}]{slot}R", part_lookup[part_b]["NAME"], feeder_b))
            continue

        if span == 2:
            records.append(_record(table, slot, "Large (2-Rel)", f"[{table}]{slot}-{slot + 1}", part_lookup[part_a]["NAME"], feeder_a))
            continue

        if span == 3:
            records.append(
                _record(
                    table,
                    slot,
                    "Extra Large (3-Rel)",
                    f"[{table}]{slot}-{slot + 2}",
                    part_lookup[part_a]["NAME"],
                    feeder_a,
                )
            )
            continue

        if span > 3:
            records.append(
                _record(
                    table,
                    slot,
                    f"Multi-Rel ({span}-Rel)",
                    f"[{table}]{slot}-{slot + span - 1}",
                    part_lookup[part_a]["NAME"],
                    feeder_a,
                )
            )
            continue

        records.append(_record(table, slot, f"Kind {kind or 'Unknown'}", f"[{table}]{slot}", part_lookup[part_a]["NAME"], feeder_a))

    position_order = {"L": 0, "R": 1, "Large (2-Rel)": 0, "Extra Large (3-Rel)": 0}
    records.sort(key=lambda item: (item["table"], item["slot"], position_order.get(item["position"], 9)))
    return records


def _build_cm602_feeder_fix_records(feeder_fix_rows):
    records = []
    for row in feeder_fix_rows:
        table = _safe_int(row.get("BeamNo", ""))
        slot = _safe_int(row.get("PU", ""))
        if not table or not slot:
            continue

        for side in CM602_FEEDER_SIDES:
            part_number = _clean_cm602_part(row.get(f"Parts{side}", ""))
            feeder_id = str(row.get(f"Feeder{side}", "")).strip()
            if not _is_cm602_assignment_active(feeder_id, part_number):
                continue

            position = _cm602_position_label(side)
            records.append(_record(table, slot, position, f"[{table}]{slot}{position}", part_number, feeder_id))

    records.sort(key=_cm602_record_sort_key)
    return records


def _build_cm602_stock_records(stock_rows, part_lookup):
    records = []
    for row in stock_rows:
        table, slot = _decode_pickup_unit(row.get("N", ""))
        if not table or not slot:
            continue

        for side in CM602_FEEDER_SIDES:
            part_id = str(row.get(f"P{side}", "")).strip()
            feeder_id = str(row.get(f"T{side}", "")).strip()
            if part_id in {"", "0", "-1"} or part_id not in part_lookup:
                continue
            part_number = _clean_cm602_part(part_lookup[part_id].get("NAME", ""))
            if not _is_cm602_assignment_active(feeder_id, part_number):
                continue

            position = _cm602_position_label(side)
            records.append(_record(table, slot, position, f"[{table}]{slot}{position}", part_number, feeder_id))

    records.sort(key=_cm602_record_sort_key)
    return records


def _read_cm602_board_size(lines):
    section_lines = _extract_section(lines, "BoardData")
    values = {}
    for line_number, line in section_lines[1:]:
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip().upper()] = value.strip()

    try:
        return float(values["X"]), float(values["Y"])
    except KeyError as exc:
        raise ServiceError(
            "Section [BoardData] harus punya nilai X dan Y.",
            title="Format CM602 tidak valid",
        ) from exc
    except ValueError as exc:
        raise ServiceError(
            f"Nilai BoardData X/Y tidak valid: X={values.get('X', '')}, Y={values.get('Y', '')}",
            title="Format CM602 tidak valid",
        ) from exc


def _build_cm602_program_cm_txt_lines(block_rows, part_lookup, board_x, board_y):
    lines = []
    missing_part_ids = []
    for index, row in enumerate(block_rows, start=1):
        part_id = str(row.get("PARTS", "")).strip()
        part_number = _cm602_part_name(part_lookup.get(part_id))
        if not part_number:
            missing_part_ids.append(part_id or f"row {index}")
            continue

        x = _required_float(row.get("X", ""), "X", index)
        y = _required_float(row.get("Y", ""), "Y", index)
        angle = _format_cm602_program_angle(row.get("A", ""), index)
        lines.append(
            "\t".join(
                [
                    str(row.get("O", "")).strip(),
                    _clean_cm602_part(row.get("C", "")),
                    str(row.get("BRM", "")).strip(),
                    _format_cm602_program_decimal(x),
                    _format_cm602_program_decimal(y),
                    _format_cm602_program_decimal(board_x - x),
                    _format_cm602_program_decimal(board_y - y),
                    angle,
                    "2",
                    "1",
                    part_number,
                ]
            )
        )

    if missing_part_ids:
        preview = ", ".join(missing_part_ids[:10])
        if len(missing_part_ids) > 10:
            preview = f"{preview}, +{len(missing_part_ids) - 10} lagi"
        raise ServiceError(
            f"Ada PARTS ID di [BlockData] yang tidak ditemukan di [PartsData]: {preview}",
            title="Format CM602 tidak valid",
        )
    return lines


def _cm602_part_name(part_row):
    if not part_row:
        return ""
    return _clean_cm602_part(part_row.get("NAME", ""))


def _required_float(value, column_name, row_number):
    try:
        return float(str(value).strip())
    except (TypeError, ValueError) as exc:
        raise ServiceError(
            f"Nilai {column_name} tidak valid di [BlockData] row {row_number}: {value}",
            title="Format CM602 tidak valid",
        ) from exc


def _format_cm602_program_decimal(value):
    return f"{float(value):.3f}"


def _format_cm602_program_angle(value, row_number):
    angle = _required_float(value, "A", row_number) % 360
    return str(int(round(angle)) % 360)


def _read_section_rows(lines, section_name, machine_label="NPM"):
    section_lines = _extract_section(lines, section_name)
    if len(section_lines) < 2:
        raise ServiceError(f"Section [{section_name}] kosong atau tidak punya header.", title=f"Format {machine_label} tidak valid")

    header_line_number, header_line = section_lines[1]
    header = _split_line(header_line, section_name, header_line_number, machine_label)
    rows = []
    for line_number, line in section_lines[2:]:
        if not line.strip():
            continue
        values = _split_line(line, section_name, line_number, machine_label)
        if len(values) != len(header):
            raise ServiceError(
                f"Jumlah kolom tidak sesuai di [{section_name}] line {line_number}.\n"
                f"Expected {len(header)}, got {len(values)}.",
                title=f"Format {machine_label} tidak valid",
            )
        rows.append(dict(zip(header, values)))
    return rows


def _read_first_available_section_rows(lines, section_names):
    for section_name in section_names:
        if _has_section(lines, section_name):
            return _read_section_rows(lines, section_name)

    names = ", ".join(f"[{section_name}]" for section_name in section_names)
    raise ServiceError(
        f"Section {names} tidak ditemukan.\n"
        "File export NPM harus berisi data part number.",
        title="Format NPM tidak valid",
    )


def _extract_section(lines, section_name):
    start_index = None
    section_marker = f"[{section_name}]"
    for index, line in enumerate(lines):
        if line.strip().lower() == section_marker.lower():
            start_index = index
            break

    if start_index is None:
        raise ServiceError(f"Section [{section_name}] tidak ditemukan.", title="Format NPM tidak valid")

    end_index = len(lines)
    for index in range(start_index + 1, len(lines)):
        stripped = lines[index].strip()
        if stripped.startswith("[") and stripped.endswith("]"):
            end_index = index
            break

    return [(line_number, lines[line_number - 1].strip()) for line_number in range(start_index + 1, end_index + 1)]


def _has_section(lines, section_name):
    section_marker = f"[{section_name}]"
    return any(line.strip().lower() == section_marker.lower() for line in lines)


def _split_line(line, section_name, line_number, machine_label="NPM"):
    try:
        return shlex.split(line, posix=True)
    except ValueError as exc:
        raise ServiceError(
            f"Gagal membaca [{section_name}] line {line_number}:\n{exc}",
            title=f"Format {machine_label} tidak valid",
        ) from exc


def _build_lookup(rows, key):
    lookup = {}
    for row in rows:
        value = str(row.get(key, "")).strip()
        if value:
            lookup[value] = row
    return lookup


def _decode_pickup_unit(value):
    try:
        pickup_unit = int(float(str(value).strip()))
    except (TypeError, ValueError):
        return 0, 0
    if pickup_unit <= 0:
        return 0, 0
    return pickup_unit // 10000, pickup_unit % 10000


def _is_active(feeder_id, part_id, part_lookup):
    return (
        str(feeder_id).strip() not in {"", "0", "-1"}
        and str(part_id).strip() not in {"", "0", "-1"}
        and str(part_id).strip() in part_lookup
    )


def _feeder_kind(feeder_row):
    try:
        return int(float(str(feeder_row.get("Kind", "0")).strip()))
    except (TypeError, ValueError):
        return 0


def _feeder_slot_span(feeder_row):
    if not feeder_row:
        return 1
    kind = _feeder_kind(feeder_row)
    name = str(feeder_row.get("NAME", "")).lower()

    if "56mm" in name:
        return 4
    if "32mm" in name or "44mm" in name:
        return 3
    if "24mm" in name:
        return 2

    if kind in (1, 2, 3):
        return 1
    if kind == 4:
        return 2
    if kind == 5:
        return 3
    if kind >= 6:
        return kind - 2

    return 1


def _safe_int(value):
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return 0


def _clean_cm602_part(value):
    return str(value or "").strip().strip('"').strip()


def _is_cm602_assignment_active(feeder_id, part_number):
    return (
        str(feeder_id).strip() not in {"", "0", "-1"}
        and str(part_number).strip() not in {"", "0", "-1"}
    )


def _cm602_position_label(side):
    side = str(side or "").upper()
    if side == "A":
        return "L"
    if side == "B":
        return "R"
    return side


def _cm602_record_sort_key(record):
    position_order = {"L": 0, "R": 1}
    position = str(record.get("position", ""))
    return (
        record.get("table", 0),
        record.get("slot", 0),
        position_order.get(position, 2 + CM602_FEEDER_SIDES.find(position) if position in CM602_FEEDER_SIDES else 99),
    )


def _location_sort_key(location_code):
    text = str(location_code or "").strip()
    match = re.match(r"^\[(\d+)\](\d+)(?:-\d+)?([A-Za-z]+)?$", text)
    if not match:
        return (999999, 999999, 999999, natural_sort_key(text))

    table = int(match.group(1))
    slot = int(match.group(2))
    side = str(match.group(3) or "").upper()
    side_order = {
        "": 0,
        "L": 0,
        "A": 0,
        "R": 1,
        "B": 1,
    }.get(side, 2 + (CM602_FEEDER_SIDES.find(side) if side in CM602_FEEDER_SIDES else 99))
    return (table, slot, side_order, natural_sort_key(text))


def _record(table, slot, position, location_code, part_number, feeder_id=""):
    return {
        "table": table,
        "slot": slot,
        "position": position,
        "location_code": location_code,
        "part_number": part_number,
        "feeder_id": feeder_id,
    }


def _append_mapping_sheet(worksheet, records):
    worksheet.append(OUTPUT_HEADERS)

    for record in records:
        worksheet.append(
            [
                record["table"],
                record["slot"],
                record["position"],
                record["location_code"],
                record["part_number"],
            ]
        )

    _style_mapping_sheet(worksheet)


def _append_summary_sheet(worksheet, summary_records):
    header_row = 1
    worksheet.append(SUMMARY_HEADERS)

    for record in summary_records:
        worksheet.append(
            [
                record["part_number"],
                record["top_location"],
                record["top_count"],
                record["total_count"],
                record["other_locations"],
                record["notes"],
            ]
        )

    _style_summary_sheet(worksheet, header_row)


def _most_frequent_location_set(item):
    if not item.get("set_counts"):
        return ()
    best_sets = []
    for loc_set, count in item["set_counts"].items():
        total_inserts = sum(item["location_counts"].get(loc, 0) for loc in loc_set)
        best_sets.append((loc_set, count, total_inserts))
    
    # Sort by file count (desc), total inserts (desc), natural sort
    best_sets.sort(
        key=lambda x: (-x[1], -x[2], [natural_sort_key(loc) for loc in x[0]])
    )
    return best_sets[0][0]


def _build_summary_records(mappings):
    stats = {}

    for mapping in mappings:
        per_file_locations = {}
        for record in mapping.records:
            part_number = str(record.get("part_number", "")).strip()
            if not part_number:
                continue
            key = part_number.upper()
            item = stats.setdefault(
                key,
                {
                    "part_number": part_number,
                    "location_counts": Counter(),
                    "location_file_counts": Counter(),
                    "set_counts": Counter(),
                    "file_count": 0,
                    "multi_file_count": 0,
                },
            )
            location_code = str(record.get("location_code", "")).strip()
            if location_code:
                item["location_counts"][location_code] += 1
                per_file_locations.setdefault(key, set()).add(location_code)

        for key, locations in per_file_locations.items():
            item = stats[key]
            location_set = tuple(sorted(locations, key=natural_sort_key))
            item["set_counts"][location_set] += 1
            item["file_count"] += 1
            if len(location_set) > 1:
                item["multi_file_count"] += 1
            for location_code in locations:
                item["location_file_counts"][location_code] += 1

    summary_records = []
    for key, item in stats.items():
        top_set = _most_frequent_location_set(item)
        if not top_set:
            continue
            
        location_items = _sorted_location_counts(item["location_file_counts"])
        candidates = [
            {
                "location": location,
                "file_count": file_count,
                "total_count": item["location_counts"].get(location, 0),
            }
            for location, file_count in location_items
        ]
        
        balancing_locations = list(top_set) if len(top_set) > 1 else []
        
        # Get the file count for this exact top set
        top_set_file_count = item["set_counts"].get(top_set, 0)
        
        summary_records.append(
            {
                "part_number": item["part_number"],
                "file_count": item["file_count"],
                "total_count": sum(item["location_counts"].values()),
                "candidates": candidates,
                "balancing_locations": balancing_locations,
                "top_set_file_count": top_set_file_count,
            }
        )

    raw_records = []
    for row in summary_records:
        candidates = row["candidates"]
        top_candidate = candidates[0]
        if row.get("balancing_locations"):
            selected_candidates = [
                candidate for candidate in candidates if candidate["location"] in row["balancing_locations"]
            ]
        else:
            selected_candidates = [top_candidate]
        
        for candidate in selected_candidates:
            selected_locations = {c["location"] for c in selected_candidates}
            other_candidates = [c for c in candidates if c["location"] not in selected_locations]
            other_locations = ", ".join(f"{c['location']} ({c['file_count']} file)" for c in other_candidates)
            notes = []
            if len(selected_candidates) > 1:
                selected_location_text = ", ".join(sorted(selected_locations, key=_location_sort_key))
                notes.append(f"BALANCING SLOTS: {selected_location_text}")
            
            raw_records.append({
                "part_number": row["part_number"],
                "top_location": candidate["location"],
                "top_count": str(candidate["file_count"]),
                "sort_count": candidate["file_count"],
                "file_count": row["file_count"],
                "total_count": row["total_count"],
                "other_locations": other_locations,
                "notes": "; ".join(notes),
            })
            
    raw_records.sort(
        key=lambda row: (
            _location_sort_key(row["top_location"]),
            -row["sort_count"],
            natural_sort_key(row["part_number"]),
        )
    )
    for row in raw_records:
        row.pop("sort_count", None)

    return _resolve_summary_location_conflicts(summary_records), raw_records


def _resolve_summary_location_conflicts(summary_records):
    ordered_records = sorted(
        summary_records,
        key=lambda row: (
            0 if row.get("balancing_locations") else 1,
            len(row["balancing_locations"] or row["candidates"]),
            -row.get("top_set_file_count", row["candidates"][0]["file_count"]),
            -row["file_count"],
            natural_sort_key(row["part_number"]),
        ),
    )

    occupied_locations = {}
    resolved_records = []
    for row in ordered_records:
        candidates = row["candidates"]
        top_candidate = candidates[0]
        selected_candidates = []
        notes = []

        if row.get("balancing_locations"):
            selected_candidates = [
                candidate
                for candidate in candidates
                if candidate["location"] in row["balancing_locations"]
            ]
            notes.append("BALANCING: part sama dipakai di beberapa feeder; bukan conflict.")
        else:
            for candidate in candidates:
                if candidate["location"] not in occupied_locations:
                    selected_candidates = [candidate]
                    break

            if not selected_candidates:
                selected_candidates = [top_candidate]
            elif selected_candidates[0]["location"] != top_candidate["location"]:
                top_blocker = occupied_locations.get(top_candidate["location"], "komponen lain")
                notes.append(f"Feeder top {top_candidate['location']} nabrak dengan {top_blocker}; pakai alternatif paling sering yang kosong.")

        for candidate in selected_candidates:
            current_part = occupied_locations.get(candidate["location"])
            if not current_part:
                occupied_locations[candidate["location"]] = row["part_number"]

        selected_locations = {candidate["location"] for candidate in selected_candidates}
        selected_location_text = ", ".join(sorted(selected_locations, key=_location_sort_key))
        if row.get("balancing_locations"):
            notes.append(f"BALANCING SLOTS: {selected_location_text}.")
        other_candidates = [candidate for candidate in candidates if candidate["location"] not in selected_locations]
        other_locations = ", ".join(f"{candidate['location']} ({candidate['file_count']} file)" for candidate in other_candidates)
        for candidate in sorted(selected_candidates, key=lambda item: _location_sort_key(item["location"])):
            resolved_records.append(
                {
                    "part_number": row["part_number"],
                    "top_location": candidate["location"],
                    "top_count": str(candidate["file_count"]),
                    "sort_count": candidate["file_count"],
                    "file_count": row["file_count"],
                    "total_count": row["total_count"],
                    "other_locations": other_locations,
                    "notes": "; ".join(notes),
                    "_selected_locations": [candidate["location"]],
                }
            )

    records_by_location = defaultdict(list)
    for record in resolved_records:
        for location in record["_selected_locations"]:
            records_by_location[location].append(record)

    for location, records in records_by_location.items():
        if len(records) < 2:
            continue
        parts = [record["part_number"] for record in records]
        for record in records:
            other_parts = [part for part in parts if part != record["part_number"]]
            conflict_note = f"CONFLICT: feeder ini juga dipakai {', '.join(other_parts)}."
            record["notes"] = f"{record['notes']}; {conflict_note}" if record["notes"] else conflict_note

    resolved_records.sort(
        key=lambda row: (
            _location_sort_key(row["top_location"]),
            -row["sort_count"],
            natural_sort_key(row["part_number"]),
        )
    )
    for row in resolved_records:
        row.pop("sort_count", None)
        row.pop("_selected_locations", None)
    return resolved_records


def _summary_balancing_locations(item):
    return _most_common_multi_location_set(item)


def _sorted_location_counts(location_counts):
    return sorted(
        location_counts.items(),
        key=lambda item: (
            -item[1],
            natural_sort_key(item[0]),
        ),
    )


def _collect_reference_files(reference_folder, exclude_path=None):
    folder = Path(_clean_path(reference_folder))
    if not folder.is_dir():
        raise ServiceError(f"Folder reference tidak ditemukan:\n{folder}", title="Folder tidak ditemukan")

    exclude_key = str(Path(exclude_path).resolve()).lower() if exclude_path else ""
    paths = []
    for path in folder.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in VALID_NPM_EXPORT_SUFFIXES:
            continue
        if str(path.resolve()).lower() == exclude_key:
            continue
        paths.append(str(path))
    paths.sort(key=lambda value: natural_sort_key(Path(value).name))
    return paths


def _parse_balancing_part_numbers(value):
    if value is None:
        values = []
    elif isinstance(value, str):
        values = re.split(r"[\s,;]+", value)
    else:
        values = value
    return {_part_key(item) for item in values if _part_key(item)}


def _build_npm_feeder_recommendations(mappings, balancing_parts):
    stats = {}
    for mapping in mappings:
        per_file_locations = {}
        for record in mapping.records:
            part_number = str(record.get("part_number", "")).strip()
            location_code = str(record.get("location_code", "")).strip()
            if not part_number or not location_code:
                continue

            key = _part_key(part_number)
            item = stats.setdefault(
                key,
                {
                    "part_number": part_number,
                    "location_counts": Counter(),
                    "location_file_counts": Counter(),
                    "feeder_counts": {},
                    "set_counts": Counter(),
                    "file_count": 0,
                    "multi_file_count": 0,
                },
            )
            item["location_counts"][location_code] += 1
            item["feeder_counts"].setdefault(location_code, Counter())[str(record.get("feeder_id", "")).strip()] += 1
            per_file_locations.setdefault(key, set()).add(location_code)

        for key, locations in per_file_locations.items():
            item = stats[key]
            location_set = tuple(sorted(locations, key=natural_sort_key))
            item["set_counts"][location_set] += 1
            item["file_count"] += 1
            if len(location_set) > 1:
                item["multi_file_count"] += 1
            for location_code in locations:
                item["location_file_counts"][location_code] += 1

    balancing_parts = set(balancing_parts or set()) | set(DEFAULT_BALANCING_PART_NUMBERS)

    recommendations = {}
    for key, item in stats.items():
        locations = _recommended_locations_for_part(key, item, balancing_parts)
        if not locations:
            continue
        feeder_ids = {}
        for location_code in locations:
            feeder_counts = item["feeder_counts"].get(location_code, Counter())
            feeder_ids[location_code] = _most_common_non_empty(feeder_counts)
        recommendations[key] = {
            "part_number": item["part_number"],
            "locations": locations,
            "feeder_ids": feeder_ids,
        }
    return recommendations


def _recommended_locations_for_part(part_key, item, balancing_parts):
    top_set = _most_frequent_location_set(item)
    is_balancing = (part_key in balancing_parts) or (len(top_set) > 1)
    
    if is_balancing:
        # If it's a balancing part but top set is single-slot, try to find a multi-slot set
        if len(top_set) <= 1:
            multi_set = _most_common_multi_location_set(item)
            if multi_set:
                return list(multi_set)
        if top_set:
            return list(top_set)
            
    if top_set:
        return list(top_set)
        
    counts_dict = item.get("location_file_counts") or item["location_counts"]
    location_items = _sorted_location_counts(counts_dict)
    if location_items:
        return [location_items[0][0]]
    return []


def _most_common_multi_location_set(item):
    multi_sets = [
        (location_set, count)
        for location_set, count in item["set_counts"].items()
        if len(location_set) > 1
    ]
    if not multi_sets:
        return []

    multi_sets.sort(
        key=lambda pair: (
            -pair[1],
            -sum(item["location_counts"].get(location, 0) for location in pair[0]),
            [natural_sort_key(location) for location in pair[0]],
        )
    )
    return list(multi_sets[0][0])


def _most_common_non_empty(counter):
    for value, _ in counter.most_common():
        if str(value).strip() not in {"", "0", "-1"}:
            return str(value).strip()
    return ""


def _target_used_part_ids(lines, part_lookup):
    section_names = []
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("[PositionData") and stripped.endswith("]"):
            section_names.append(stripped[1:-1])

    part_ids = set()
    for section_name in section_names:
        try:
            rows = _read_section_rows(lines, section_name)
        except ServiceError:
            continue
        for row in rows:
            part_id = str(row.get("PARTS", "")).strip()
            if part_id in {"", "0", "-1"} or part_id not in part_lookup:
                continue
            part_name = str(part_lookup.get(part_id, {}).get("NAME", "")).strip()
            if part_name.upper() in {"", "OHM", "BLANK"}:
                continue
            part_ids.add(part_id)
    return part_ids


def _unique_target_parts(part_ids, part_lookup):
    parts = {}
    for part_id in sorted(part_ids, key=natural_sort_key):
        part_row = part_lookup.get(part_id, {})
        part_number = str(part_row.get("NAME", "")).strip()
        part_key = _part_key(part_number)
        if not part_key or part_key in parts:
            continue
        parts[part_key] = {
            "part_id": part_id,
            "part_number": part_number,
            "part_row": part_row,
        }
    return parts


def _fixed_feeder_header(lines):
    return _section_header(lines, "FixedFeeder")


def _section_header(lines, section_name):
    start, _ = _section_bounds(lines, section_name)
    if start + 1 >= len(lines):
        raise ServiceError(f"Header [{section_name}] tidak ditemukan.", title="Format NPM tidak valid")
    return _split_line(lines[start + 1].strip(), section_name, start + 2)


def _existing_fixed_assignments(fixed_rows, part_lookup, feeder_lookup):
    assignments = {}
    for row in fixed_rows:
        for assignment in _fixed_row_assignments(row, part_lookup, feeder_lookup):
            assignments.setdefault(_part_key(assignment["part_number"]), []).append(assignment)
    return assignments


def _existing_feeder_id_for_location(existing_assignments, part_key, location_code):
    for assignment in existing_assignments.get(part_key, []):
        if assignment.get("location_code") == location_code:
            return assignment.get("feeder_id", "")
    return ""


def _fixed_row_assignments(row, part_lookup, feeder_lookup):
    table, slot = _decode_pickup_unit(row.get("PU", ""))
    if not table or not slot:
        return []

    assignments = []
    for side in ("A", "B"):
        feeder_id = str(row.get(f"Feeder{side}", "")).strip()
        part_id = str(row.get(f"Parts{side}", "")).strip()
        if not _is_active(feeder_id, part_id, part_lookup):
            continue

        feeder_row = feeder_lookup.get(feeder_id, {})
        kind = _feeder_kind(feeder_row)
        span = _feeder_slot_span(feeder_row)

        if kind == 2:
            position = "L" if side == "A" else "R"
            location_code = f"[{table}]{slot}{position}"
        elif span == 2:
            location_code = f"[{table}]{slot}-{slot + 1}"
        elif span == 3:
            location_code = f"[{table}]{slot}-{slot + 2}"
        elif span > 3:
            location_code = f"[{table}]{slot}-{slot + span - 1}"
        else:
            location_code = f"[{table}]{slot}"

        assignments.append(
            {
                "part_number": part_lookup[part_id]["NAME"],
                "location_code": location_code,
                "feeder_id": feeder_id,
            }
        )
    return assignments


def _empty_fixed_row(row):
    output = dict(row)
    for side in FIXED_FEEDER_SIDES:
        output[f"Feeder{side}"] = "-1"
        output[f"Parts{side}"] = "0"
    return output


def _set_fixed_assignment(row, side, feeder_id, part_id):
    for extra_side in FIXED_FEEDER_SIDES[2:]:
        row[f"Feeder{extra_side}"] = "0"
        row[f"Parts{extra_side}"] = "0"
    row[f"Feeder{side}"] = str(feeder_id)
    row[f"Parts{side}"] = str(part_id)


def _parse_location_code(location_code):
    match = re.match(r"^\[(\d+)\](\d+)(?:-\d+)?([LR])?$", str(location_code or "").strip(), flags=re.IGNORECASE)
    if not match:
        return None
    table = int(match.group(1))
    slot = int(match.group(2))
    side = "B" if str(match.group(3) or "").upper() == "R" else "A"
    return table * 10000 + slot, side


def _default_feeder_id(part_row, feeder_lookup):
    for key in ("FA", "FB", "FC", "FD", "FE", "FF", "FG", "FH", "FI", "FJ"):
        feeder_id = str(part_row.get(key, "")).strip()
        if feeder_id not in {"", "0", "-1"} and feeder_id in feeder_lookup:
            return feeder_id
    return ""


def _replace_fixed_feeder_section(lines, fixed_header, fixed_rows):
    start, end = _section_bounds(lines, "FixedFeeder")
    newline = _detect_newline(lines)
    section_lines = [
        _strip_newline(lines[start]) + newline,
        " ".join(fixed_header) + newline,
    ]
    for row in fixed_rows:
        section_lines.append(" ".join(str(row.get(column, "")) for column in fixed_header) + newline)
    return list(lines[:start]) + section_lines + list(lines[end:])


def _section_bounds(lines, section_name):
    marker = f"[{section_name}]"
    start = None
    for index, line in enumerate(lines):
        if line.strip().lower() == marker.lower():
            start = index
            break
    if start is None:
        raise ServiceError(f"Section [{section_name}] tidak ditemukan.", title="Format NPM tidak valid")

    end = len(lines)
    for index in range(start + 1, len(lines)):
        stripped = lines[index].strip()
        if stripped.startswith("[") and stripped.endswith("]"):
            end = index
            break
    return start, end


def _detect_newline(lines):
    for line in lines:
        if line.endswith("\r\n"):
            return "\r\n"
        if line.endswith("\n"):
            return "\n"
    return "\n"


def _strip_newline(line):
    return str(line).rstrip("\r\n")


def _part_key(value):
    return str(value or "").strip().upper()


def _style_mapping_sheet(worksheet):
    header_fill = PatternFill("solid", fgColor="FF2B3A4C")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    default_font = Font(name="Calibri", size=11)
    border_side = Side(style="thin", color="FFD3D3D3")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    widths = {"A": 9, "B": 8, "C": 23, "D": 17, "E": 15}
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = default_font
            cell.border = border
            horizontal = "left" if cell.column in (4, 5) else "center"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def _style_summary_sheet(worksheet, header_row):
    header_fill = PatternFill("solid", fgColor="FF2B3A4C")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    default_font = Font(name="Calibri", size=11)
    border_side = Side(style="thin", color="FFD3D3D3")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    widths = {
        "A": 28,
        "B": 22,
        "C": 24,
        "D": 14,
        "E": 44,
        "F": 72,
    }
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width

    for cell in worksheet[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in worksheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.font = default_font
            cell.border = border
            horizontal = "left" if cell.column in (1, 2, 5, 6) else "center"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center")

    worksheet.freeze_panes = f"A{header_row + 1}"
    worksheet.auto_filter.ref = f"A{header_row}:F{worksheet.max_row}"


def _unique_sheet_title(source_file, used_titles):
    base = _clean_sheet_title(Path(source_file or "Feeder Setup").stem) or "Feeder Setup"
    if base.lower() == "summary":
        base = "Summary Detail"

    title = base[:31]
    counter = 2
    while title.lower() in used_titles:
        suffix = f" ({counter})"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1

    used_titles.add(title.lower())
    return title


def _clean_sheet_title(value):
    text = str(value or "").strip()
    text = re.sub(r"[\[\]:*?/\\]+", "_", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" .'")


def _clean_source_paths(source_paths):
    if isinstance(source_paths, (str, Path)):
        raw_paths = [source_paths]
    else:
        raw_paths = list(source_paths or [])

    paths = []
    seen = set()
    for raw_path in raw_paths:
        path = _clean_path(raw_path)
        if not path:
            continue
        normalized = str(Path(path))
        key = normalized.lower()
        if key in seen:
            continue
        seen.add(key)
        paths.append(normalized)
    return paths


def _clean_filename_part(value):
    text = str(value or "").strip()
    text = re.sub(r'[<>:"/\\|?*]+', "_", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" .")


def _clean_path(file_path):
    return str(file_path or "").strip().replace('"', "").replace("'", "")
