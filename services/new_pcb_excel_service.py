import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font

from services import bom_service
from services.errors import ServiceError


CAD_ENCODINGS = ("utf-8-sig", "utf-8", "cp949", "euc-kr", "cp1252", "latin-1")
LIBRARY_SHEET_NAME = "\uc774\ud615 Data Lib"


@dataclass
class NewPcbExcelConfig:
    cad_path: str
    bom_path: str
    library_path: str
    reference_path: str
    gerber_image_path: str
    model: str
    program_part_number: str
    equivalent_part_number: str
    pcb_part_number: str
    pcb_revision: str
    wo_supply: str
    creator: str
    line: str
    output_path: str


def suggest_output_name(config: NewPcbExcelConfig):
    date_code = datetime.now().strftime("%y%m%d")
    model = _clean_filename_part(config.model) or "MODEL"
    program = _clean_filename_part(_combined_program_part_number(config)) or "PROGRAM"
    pcb = _clean_filename_part(config.pcb_part_number) or "PCB"
    revision = _clean_filename_part(config.pcb_revision) or "REV"
    return f"(INI_OHM)_{model}({program})_{pcb}({revision})_{date_code}.xlsx"


def generate_new_pcb_excel(config: NewPcbExcelConfig):
    _validate_config(config)

    cad_rows = _read_cad_rows(config.cad_path)
    pcb_width, pcb_height = _extract_pcb_size(cad_rows)
    bom_result = bom_service.load_raw_bom(config.bom_path, check_duplicate_circuits=False)
    reference_lookup = _load_reference_mc_lookup(config.reference_path)

    output_path = Path(config.output_path)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    ws_memo = workbook.active
    ws_memo.title = "MEMO"
    ws_cad = workbook.create_sheet("CAD")
    ws_dx = workbook.create_sheet("DX")
    ws_bom = workbook.create_sheet("BOM")

    _build_memo_sheet(ws_memo, config)
    _build_cad_sheet(ws_cad, cad_rows, pcb_width, pcb_height)
    _insert_gerber_image(ws_cad, config.gerber_image_path)
    dx_last_row = _build_dx_sheet(ws_dx, cad_rows, pcb_width, pcb_height, config, bom_result, reference_lookup)
    _build_bom_sheet(ws_bom, bom_result, config, ws_dx)
    _apply_workbook_options(workbook)
    _apply_common_format(ws_memo, ws_cad, ws_dx, ws_bom, dx_last_row)

    workbook.save(output_path)
    return str(output_path)


def _validate_config(config):
    files = [
        (config.cad_path, "CAD Data"),
        (config.bom_path, "BOM File"),
        (config.library_path, "Excel Part Library"),
        (config.reference_path, "Excel Referensi"),
        (config.gerber_image_path, "Gerber PCB Image"),
    ]
    for path, label in files:
        if not path:
            raise ServiceError(f"{label} belum dipilih.", title="Input belum lengkap")
        if not Path(path).is_file():
            raise ServiceError(f"{label} tidak ditemukan:\n{path}", title="File tidak ditemukan")

    required_fields = [
        (config.model, "Model"),
        (config.program_part_number, "Part Number Program"),
        (config.pcb_part_number, "PCB Part Number"),
        (config.pcb_revision, "PCB Revision Number"),
        (config.wo_supply, "WO Supply"),
        (config.creator, "Nama pembuat"),
        (config.line, "LINE"),
        (config.output_path, "Lokasi output"),
    ]
    missing = [label for value, label in required_fields if not str(value or "").strip()]
    if missing:
        raise ServiceError("Field wajib belum diisi:\n" + ", ".join(missing), title="Input belum lengkap")


def _read_cad_rows(path):
    lines = None
    last_error = None
    for encoding in CAD_ENCODINGS:
        try:
            with open(path, "r", encoding=encoding) as handle:
                lines = handle.readlines()
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    if lines is None:
        raise ServiceError(f"CAD Data tidak bisa dibaca.\n{last_error}", title="Encoding error")

    rows = []
    for line in lines:
        raw = line.rstrip("\r\n")
        if "|" in raw:
            parts = raw.split("|")
        elif "\t" in raw:
            parts = raw.split("\t")
        else:
            parts = [raw]
        rows.append([_coerce_cad_value(part) for part in parts])
    return rows


def _coerce_cad_value(value):
    if value == "":
        return None
    stripped = value.strip()
    if re.fullmatch(r"[+-]?\d+(?:[.,]\d+)?", stripped):
        number = float(stripped.replace(",", "."))
        return int(number) if number.is_integer() else number
    return value


def _extract_pcb_size(cad_rows):
    if len(cad_rows) < 3:
        raise ServiceError("CAD Data terlalu pendek untuk membaca ukuran PCB.", title="Format CAD tidak valid")
    width = _extract_number_after_equal(cad_rows[1][0] if cad_rows[1] else "")
    height = _extract_number_after_equal(cad_rows[2][0] if cad_rows[2] else "")
    if width == 0 or height == 0:
        raise ServiceError("Gagal membaca WIDTH (X) dan HIGHT (Y) dari CAD Data.", title="Format CAD tidak valid")
    return width, height


def _extract_number_after_equal(value):
    text = str(value or "")
    if "=" not in text:
        return 0
    tail = text.split("=", 1)[1].strip()
    match = re.search(r"[+-]?\d+(?:[.,]\d+)?", tail)
    if not match:
        return 0
    return float(match.group(0).replace(",", "."))


def _build_memo_sheet(ws, config):
    ws["A1"] = datetime.now()
    ws["A1"].number_format = "m/dd/yyyy hh:mm:ss"
    ws["B1"] = _normalize_line(config.line)
    ws["C1"] = "NEW"
    ws["D1"] = _creator_memo_value(config.creator)
    ws["E1"] = config.wo_supply.strip()

    ws.column_dimensions["A"].width = 16.86
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 13.43

    ws["E1"].alignment = Alignment(vertical="center")


def _normalize_line(value):
    text = str(value or "").strip().upper()
    if not text:
        return ""
    if text.isdigit():
        return f"INI{text}"
    if re.fullmatch(r"LINE\s*\d+", text):
        return "INI" + re.search(r"\d+", text).group(0)
    return text


def _creator_memo_value(value):
    text = str(value or "").strip()
    if not text:
        return ""
    return text if text.startswith(">") else f">>{text}"


def _combined_program_part_number(config):
    program = str(config.program_part_number or "").strip()
    equivalent = str(config.equivalent_part_number or "").strip()
    if program and equivalent:
        return f"{program}+{equivalent}"
    return program or equivalent


def _build_cad_sheet(ws, cad_rows, pcb_width, pcb_height):
    _write_rows(ws, cad_rows)
    for column in range(1, 15):
        ws.column_dimensions[_column_letter(column)].width = 8 if column <= 10 else 13

    sma_row = _find_component_row(cad_rows, "SMA")
    smb_row = _find_component_row(cad_rows, "SMB")
    if not sma_row or not smb_row:
        raise ServiceError("Data SMA/SMB tidak ditemukan di CAD Data.", title="Format CAD tidak valid")

    header_row = min(sma_row, smb_row) - 1
    width_text = _format_formula_number(pcb_width)
    height_text = _format_formula_number(pcb_height)
    ws.cell(header_row, 12).value = f"XX={width_text}"
    ws.cell(header_row, 13).value = f"YY={height_text}"
    for cell in (ws.cell(header_row, 12), ws.cell(header_row, 13)):
        cell.number_format = "@"
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFF0000")

    for row in (sma_row, smb_row):
        ws.cell(row, 12).value = f"={width_text}-D{row}"
        ws.cell(row, 13).value = f"={height_text}-E{row}"
        ws.cell(row, 12).number_format = "0.000"
        ws.cell(row, 13).number_format = "0.000"
        for column in range(1, 14):
            ws.cell(row, column).font = Font(name="Calibri", size=10, bold=True, color="FFFF0000")


def _insert_gerber_image(ws, image_path):
    try:
        image = ExcelImage(image_path)
    except Exception as exc:
        raise ServiceError(f"Gerber PCB Image tidak bisa dibaca:\n{exc}", title="Image error") from exc

    max_width = 520
    max_height = 360
    scale = min(max_width / image.width, max_height / image.height, 1)
    image.width = int(image.width * scale)
    image.height = int(image.height * scale)
    image.anchor = "K7"
    ws.add_image(image)


def _build_dx_sheet(ws, cad_rows, pcb_width, pcb_height, config, bom_result, reference_lookup):
    source_rows = _dx_source_rows(cad_rows)
    bom_lookup = _build_bom_part_lookup(bom_result)
    source_rows = _sort_dx_rows_like_reprog(source_rows, bom_lookup, reference_lookup)
    header = ["S", "CODE", "T", "X", "Y", "XX", "YY", "A", "P", "B", "PART", "SPEC", "FD", "MC"]
    _write_rows(ws, [header])

    for target_row, source in enumerate(source_rows, start=2):
        expanded = _expand_dx_row(source)
        for column, value in enumerate(expanded, start=1):
            if value is not None:
                ws.cell(target_row, column).value = value

    width_text = _format_formula_number(pcb_width)
    height_text = _format_formula_number(pcb_height)
    library_ref = _external_sheet_ref(config.library_path, LIBRARY_SHEET_NAME)
    last_row = len(source_rows) + 1

    for row in range(2, last_row + 1):
        circuit = _key(ws.cell(row, 2).value)
        part_number = bom_lookup.get(circuit)
        ws.cell(row, 11).value = None
        ws.cell(row, 12).value = None
        ws.cell(row, 15).value = None
        ws.cell(row, 1).value = 1
        ws.cell(row, 6).value = f"={width_text}-D{row}"
        ws.cell(row, 7).value = f"={height_text}-E{row}"
        ws.cell(row, 9).value = 2
        if part_number:
            ws.cell(row, 11).value = f"=VLOOKUP(B{row},BOM!B:C,2,0)"
            ws.cell(row, 12).value = f"=VLOOKUP(K{row},{library_ref}!$A:$B,2,0)"
            ws.cell(row, 13).value = f"=VLOOKUP(K{row},{library_ref}!$A:$E,5,0)"
            ws.cell(row, 14).value = reference_lookup.get(_key(part_number))

    for column in ("A", "C", "I", "J", "N"):
        ws.column_dimensions[column].width = 4.5
    for column in ("B", "D", "E", "F", "G", "H", "M"):
        ws.column_dimensions[column].width = 9
    for column in ("O", "P", "Q", "R", "S", "T"):
        ws.column_dimensions[column].width = 13
    for column in ("A", "C", "I", "J", "N"):
        ws.column_dimensions[column].width = 4.5
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 70

    return last_row


def _expand_dx_row(row):
    values = list(row)
    if len(values) < 12:
        values.extend([None] * (12 - len(values)))
    expanded = values[:5] + [None, None] + values[5:]
    if len(expanded) < 20:
        expanded.extend([None] * (20 - len(expanded)))
    return expanded[:20]


def _build_bom_part_lookup(bom_result):
    lookup = {}
    for _, row in bom_result.dataframe.iterrows():
        circuit = _key(row.get("Circuit", ""))
        part_number = str(row.get("PartNo", "") or "").strip()
        if circuit and part_number and circuit not in lookup:
            lookup[circuit] = part_number
    return lookup


def _load_reference_mc_lookup(reference_path):
    try:
        workbook = load_workbook(reference_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ServiceError(f"Excel Referensi tidak bisa dibaca untuk lookup kolom N:\n{exc}") from exc

    if "DX" not in workbook.sheetnames:
        workbook.close()
        raise ServiceError('Sheet "DX" tidak ditemukan pada Excel Referensi.', title="Format Referensi tidak valid")

    try:
        worksheet = workbook["DX"]
        lookup = {}
        for row in worksheet.iter_rows(min_row=2, min_col=11, max_col=14, values_only=True):
            part_number = _key(row[0])
            machine_code = row[3]
            if part_number and machine_code not in (None, "") and part_number not in lookup:
                lookup[part_number] = machine_code
        return lookup
    finally:
        workbook.close()


def _sort_dx_rows_like_reprog(rows, bom_lookup, reference_lookup):
    def sort_key(row):
        expanded = _expand_dx_row(row)
        circuit = _key(expanded[1])
        part_number = bom_lookup.get(circuit, "")
        machine_code = reference_lookup.get(_key(part_number))
        return (
            _sort_value(expanded[2]),
            _sort_value(machine_code),
            _sort_value(""),
            _sort_value(part_number),
            _sort_value(circuit),
        )

    return sorted(rows, key=sort_key)


def _sort_value(value):
    if value is None or str(value).strip() == "":
        return (1, 0, "")
    if isinstance(value, (int, float)):
        return (0, 0, float(value))
    text = str(value).strip()
    try:
        return (0, 0, float(text))
    except ValueError:
        return (0, 1, text.upper())


def _dx_source_rows(cad_rows):
    sorting_headers = []
    for idx, row in enumerate(cad_rows):
        kind = _cad_sorting_kind(row)
        if kind:
            sorting_headers.append((idx, kind))

    source_rows = []
    for header_pos, kind in sorting_headers:
        if kind != "CHIP":
            continue
        next_header_pos = next(
            (idx for idx, _ in sorting_headers if idx > header_pos),
            len(cad_rows),
        )
        source_rows.extend(cad_rows[header_pos + 1 : next_header_pos])

    source_rows = [
        list(row)
        for row in source_rows
        if any(value is not None and str(value).strip() for value in row)
    ]

    if not source_rows:
        header_summary = ", ".join(f"{kind}@{idx + 1}" for idx, kind in sorting_headers) or "tidak ada"
        raise ServiceError(
            "Data CHIP Sorting tidak ditemukan di CAD Data.\n"
            f"Header Sorting terdeteksi: {header_summary}",
            title="Format CAD tidak valid",
        )

    return source_rows


def _cad_sorting_kind(row):
    label = str(row[0] if row else "")
    label_upper = label.upper()
    if "SORTING" not in label_upper or "###" not in label:
        return ""
    if "CHIP" in label_upper:
        return "CHIP"
    if "DIP" in label_upper:
        return "DIP"
    return ""


def _find_component_row(rows, name):
    for idx, row in enumerate(rows, start=1):
        value = row[1] if len(row) > 1 else None
        if str(value or "").strip().upper() == name:
            return idx
    return 0


def _build_bom_sheet(ws, bom_result, config, ws_dx):
    now = datetime.now()
    pcb_part_number = config.pcb_part_number.strip() or bom_result.pcb_pn
    ws.append([now, "Circuit No", pcb_part_number, "Spec", "Side"])
    ws["A1"].number_format = "m/d/yy h:mm"

    dx_insert_lookup = _build_dx_insert_lookup(ws_dx)
    bom_rows = []
    for _, row in bom_result.dataframe.iterrows():
        circuit = str(row.get("Circuit", "") or "")
        bom_rows.append(
            (
                _bom_sort_key(dx_insert_lookup.get(_key(circuit))),
                [
                    str(row.get("Chassis", "") or ""),
                    circuit,
                    str(row.get("PartNo", "") or ""),
                    str(row.get("Spec", "") or ""),
                    str(row.get("Side", "") or ""),
                ],
            )
        )

    for _, row_values in sorted(bom_rows, key=lambda item: item[0]):
        ws.append(
            row_values
        )

    last_row = ws.max_row
    for row in range(2, last_row + 1):
        ws.cell(row, 6).value = f"=VLOOKUP(B{row},DX!B:C,2,0)"

    if last_row >= 2:
        ws["G2"] = "=COUNTIF(F:F,0)"
        ws["H2"] = "TOP"
        ws["G3"] = "=COUNTIF(F:F,1)"
        ws["H3"] = "BOT"
        ws["G4"] = "=COUNTIF(F:F,#N/A)"
        ws["H4"] = "ORP/BPR"
        ws["G5"] = "=SUM(G2:G4)"
        ws["H5"] = "TOTAL"

    widths = {"A": 15.86, "B": 8.71, "C": 12.57, "D": 59.86, "E": 9, "F": 13, "G": 13, "H": 13}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for cell in ws[1]:
        cell.font = Font(name="Calibri", size=10, bold=True)


def _build_dx_insert_lookup(ws_dx):
    lookup = {}
    for row in range(2, ws_dx.max_row + 1):
        circuit = _key(ws_dx.cell(row, 2).value)
        insert_side = ws_dx.cell(row, 3).value
        if circuit and circuit not in lookup:
            lookup[circuit] = insert_side
    return lookup


def _bom_sort_key(insert_side):
    if insert_side is None or str(insert_side).strip() == "":
        return (0, "")
    try:
        return (1, -float(insert_side))
    except (TypeError, ValueError):
        return (1, str(insert_side).upper())


def _apply_workbook_options(workbook):
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.calcMode = "auto"


def _apply_common_format(ws_memo, ws_cad, ws_dx, ws_bom, dx_last_row):
    center = Alignment(horizontal="center", vertical="center")

    for ws in (ws_memo, ws_cad, ws_dx, ws_bom):
        ws.sheet_format.defaultRowHeight = 12.75
        for row_number in range(1, ws.max_row + 1):
            ws.row_dimensions[row_number].height = 12.75
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(
                    name="Calibri",
                    size=10,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    underline=cell.font.underline,
                    strike=cell.font.strike,
                    color="FF000000",
                )

    for row in range(1, ws_cad.max_row + 1):
        ws_cad[f"D{row}"].number_format = "0.000"
        ws_cad[f"E{row}"].number_format = "0.000"

    _highlight_cad_mount_mark_rows(ws_cad)
    _highlight_cad_pcb_size_labels(ws_cad)

    for row in ws_dx.iter_rows(min_row=1, max_row=max(dx_last_row, 1), min_col=1, max_col=20):
        for cell in row:
            cell.alignment = center
    for column in ("D", "E", "F", "G", "M"):
        for row in range(1, dx_last_row + 1):
            ws_dx[f"{column}{row}"].number_format = "0.000"


def _highlight_cad_mount_mark_rows(ws):
    red_font = Font(name="Calibri", size=10, bold=True, color="FFFF0000")
    for row in range(1, ws.max_row + 1):
        marker = str(ws.cell(row, 2).value or "").strip().upper()
        if marker in {"SMA", "SMB"}:
            for column in range(1, max(ws.max_column, 14) + 1):
                ws.cell(row, column).font = red_font


def _highlight_cad_pcb_size_labels(ws):
    red_font = Font(name="Calibri", size=10, bold=True, color="FFFF0000")
    for row in range(1, ws.max_row + 1):
        for column in (12, 13):
            value = str(ws.cell(row, column).value or "").strip().upper()
            if value.startswith("XX=") or value.startswith("YY="):
                ws.cell(row, column).font = red_font


def _write_rows(ws, rows):
    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            if value is not None:
                ws.cell(row_index, column_index).value = value


def _external_sheet_ref(path, sheet_name):
    file_path = Path(path).resolve()
    directory = str(file_path.parent).replace("/", "\\")
    if directory and not directory.endswith("\\"):
        directory += "\\"
    ref = f"{directory}[{file_path.name}]{sheet_name}"
    return "'" + ref.replace("'", "''") + "'"


def _key(value):
    return str(value or "").strip()


def _format_formula_number(value):
    if float(value).is_integer():
        return str(int(value))
    return f"{float(value):.6f}".rstrip("0").rstrip(".")


def _clean_filename_part(value):
    text = str(value or "").strip()
    text = re.sub(r'[<>:"/\\|?*]+', "_", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" .")


def _column_letter(index):
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters
