import os
import re
import shutil
import tempfile
import time
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from services.errors import ServiceError


EXCEL_EXTENSIONS = (".xls", ".xlsx", ".xlsm", ".xlsb")
OPENPYXL_EXTENSIONS = (".xlsx", ".xlsm")
MODE_PCB_LIST = "pcb_list"
MODE_PROGRAM_FOLDER = "program_folder"
XL_UP = -4162
XL_CALCULATION_MANUAL = -4135
XL_CALCULATION_AUTOMATIC = -4105
COM_BUSY_HRESULTS = {-2147418111, -2147417846, -2147417845}


@dataclass
class UsedPartComponentConfig:
    mode: str
    source_folder: str
    output_path: str
    pcb_part_numbers: str = ""


@dataclass
class UsedPartComponentResult:
    output_path: str
    mode: str
    group_count: int
    file_count: int
    part_count: int
    skipped_files: list[str] = field(default_factory=list)


def suggest_output_name(mode=MODE_PROGRAM_FOLDER):
    mode_label = "MODE_1" if mode == MODE_PCB_LIST else "MODE_2"
    return f"Used_Part_Component_{mode_label}_{datetime.now().strftime('%y%m%d')}.xlsx"


def generate_used_part_component(config: UsedPartComponentConfig):
    _validate_config(config)
    output_path = _normalize_output_path(config.output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with BomReader() as reader:
        if config.mode == MODE_PCB_LIST:
            collection = _collect_by_pcb_list(config, reader)
            _write_mode_1_workbook(collection, output_path)
        else:
            collection = _collect_by_program_folder(config, reader)
            _write_mode_2_workbook(collection, output_path)

    return UsedPartComponentResult(
        output_path=str(output_path),
        mode=config.mode,
        group_count=len(collection.groups),
        file_count=collection.file_count,
        part_count=collection.part_count,
        skipped_files=collection.skipped_files,
    )


@dataclass
class PartCollection:
    groups: OrderedDict
    file_count: int = 0
    skipped_files: list[str] = field(default_factory=list)

    @property
    def part_count(self):
        master = OrderedDict()
        for parts in self.groups.values():
            for part in parts:
                key = _part_key(part)
                if key and key not in master:
                    master[key] = part
        return len(master)


class BomReader:
    def __init__(self):
        self._excel = None
        self._pythoncom = None
        self._temp_dir = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        self.close()

    def read_part_names(self, file_path, start_row=2, filter_na_column=False, exclude_board_parts=False):
        path = Path(file_path)
        if path.suffix.lower() in OPENPYXL_EXTENSIONS:
            try:
                return _read_part_names_openpyxl(path, start_row, filter_na_column, exclude_board_parts)
            except ServiceError:
                raise
            except Exception:
                pass

        return self._read_part_names_com(path, start_row, filter_na_column, exclude_board_parts)

    def close(self):
        if self._excel is not None:
            _restore_and_quit_excel(self._excel)
            self._excel = None
        if self._pythoncom is not None:
            self._pythoncom.CoUninitialize()
            self._pythoncom = None
        if self._temp_dir is not None:
            self._temp_dir.cleanup()
            self._temp_dir = None

    def _ensure_excel(self):
        if self._excel is not None:
            return self._excel

        try:
            import pythoncom
            import win32timezone
            import win32com.client as win32
        except ImportError as exc:
            raise ServiceError(
                "File Excel lama (.xls/.xlsb) membutuhkan pywin32 dan Microsoft Excel terinstall.",
                title="Excel tidak tersedia",
            ) from exc

        pythoncom.CoInitialize()
        self._pythoncom = pythoncom
        try:
            self._excel = _excel_call(lambda: win32.DispatchEx("Excel.Application"))
        except Exception as exc:
            pythoncom.CoUninitialize()
            self._pythoncom = None
            raise ServiceError(
                "Microsoft Excel tidak bisa dibuka. Pastikan Excel terinstall dan tidak sedang menampilkan dialog.",
                title="Excel tidak tersedia",
            ) from exc

        _configure_excel_for_background(self._excel)
        self._temp_dir = tempfile.TemporaryDirectory(prefix="smt_used_parts_")
        return self._excel

    def _read_part_names_com(self, path, start_row, filter_na_column, exclude_board_parts):
        excel = self._ensure_excel()
        workbook = None
        temp_path = None

        try:
            workbook, temp_path = self._open_workbook(excel, path)
            try:
                worksheet = _excel_call(lambda: workbook.Worksheets("BOM"))
            except Exception as exc:
                raise ServiceError('Sheet "BOM" tidak ditemukan.', title="Format Excel tidak valid") from exc

            last_row = int(_excel_call(lambda: worksheet.Cells(worksheet.Rows.Count, "C").End(XL_UP).Row))
            if last_row < start_row:
                return []

            if filter_na_column:
                data = _excel_call(lambda: worksheet.Range(f"C{start_row}:F{last_row}").Value)
                rows = _ensure_2d(data)
                values = []
                for row in rows:
                    part_value = row[0] if row else None
                    marker_value = row[3] if len(row) > 3 else None
                    if _is_excel_error(part_value) or _is_excel_error(marker_value):
                        continue
                    if _is_na_marker(marker_value):
                        continue
                    values.append(part_value)
            else:
                data = _excel_call(lambda: worksheet.Range(f"C{start_row}:C{last_row}").Value)
                rows = _ensure_2d(data)
                values = [row[0] if row else None for row in rows]

            return _clean_part_values(values, exclude_board_parts)

        finally:
            if workbook is not None:
                _close_workbook(workbook)
            if temp_path and os.path.exists(temp_path):
                os.remove(temp_path)

    def _open_workbook(self, excel, path):
        source = str(path)
        try:
            if len(source) <= 200:
                return _open_workbook(excel, source), None
        except Exception:
            pass

        suffix = path.suffix or ".xlsx"
        temp_path = os.path.join(
            self._temp_dir.name,
            f"TEMP_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}{suffix}",
        )
        shutil.copy2(source, temp_path)
        return _open_workbook(excel, temp_path), temp_path


def _collect_by_program_folder(config, reader):
    folder = Path(config.source_folder)
    files = _excel_files_in_folder(folder)
    if not files:
        raise ServiceError("Tidak ada file Excel program di folder yang dipilih.", title="Data kosong")

    groups = OrderedDict()
    skipped_files = []
    file_count = 0

    for file_path in files:
        group_name = _unique_name(_extract_program_name(file_path.name), groups)
        try:
            parts = reader.read_part_names(file_path, start_row=2, filter_na_column=True, exclude_board_parts=False)
        except Exception as exc:
            skipped_files.append(f"{file_path.name}: {_error_message(exc)}")
            continue

        if not parts:
            skipped_files.append(f"{file_path.name}: tidak ada part name di sheet BOM kolom C")
            continue

        groups[group_name] = _unique_parts(parts)
        file_count += 1

    if not groups:
        raise ServiceError("Tidak ada data part yang berhasil dibaca dari folder ini.", title="Data kosong")

    return PartCollection(groups=groups, file_count=file_count, skipped_files=skipped_files)


def _collect_by_pcb_list(config, reader):
    folder = Path(config.source_folder)
    pcb_numbers = _parse_pcb_part_numbers(config.pcb_part_numbers)
    folder_map = _map_pcb_folders(folder, pcb_numbers)
    if not folder_map:
        raise ServiceError("Tidak ada folder PCB yang cocok dengan list part number.", title="Folder tidak ditemukan")

    groups = OrderedDict()
    skipped_files = []
    file_count = 0

    for pcb_number in pcb_numbers:
        matched_folders = folder_map.get(pcb_number, [])
        if not matched_folders:
            skipped_files.append(f"{pcb_number}: folder tidak ditemukan")
            continue

        merged_parts = []
        for pcb_folder in matched_folders:
            files = _excel_files_in_folder(pcb_folder)
            if not files:
                skipped_files.append(f"{pcb_number}: tidak ada file Excel di {pcb_folder.name}")
                continue

            for file_path in files:
                try:
                    parts = reader.read_part_names(file_path, start_row=1, filter_na_column=True, exclude_board_parts=True)
                except Exception as exc:
                    skipped_files.append(f"{pcb_number} / {file_path.name}: {_error_message(exc)}")
                    continue

                if parts:
                    merged_parts.extend(parts)
                    file_count += 1
                else:
                    skipped_files.append(f"{pcb_number} / {file_path.name}: tidak ada part valid")

        unique = _unique_parts(merged_parts)
        if unique:
            groups[pcb_number] = unique

    if not groups:
        raise ServiceError("Tidak ada data part valid dari PCB part number yang dipilih.", title="Data kosong")

    return PartCollection(groups=groups, file_count=file_count, skipped_files=skipped_files)


def _write_mode_1_workbook(collection, output_path):
    _write_matrix_workbook(collection, output_path, "MASTER")


def _write_mode_2_workbook(collection, output_path):
    _write_matrix_workbook(collection, output_path, "P/N COMPONENT")


def _write_matrix_workbook(collection, output_path, master_header):
    workbook = Workbook()
    master = workbook.active
    master.title = "MASTER"
    master_parts = _write_master_matrix(master, collection, master_header)

    _style_simple_sheet(master, len(master_parts) + 1, len(collection.groups) + 1, auto_filter=True)
    _add_detail_sheets(workbook, collection.groups)
    workbook.save(output_path)


def _write_master_matrix(worksheet, collection, master_header):
    master_parts = _sorted_master_parts(collection.groups)
    group_lookup = {
        group_name: {_part_key(part) for part in parts}
        for group_name, parts in collection.groups.items()
    }

    worksheet.cell(1, 1).value = master_header
    for col_index, group_name in enumerate(collection.groups.keys(), start=2):
        worksheet.cell(1, col_index).value = group_name

    for row_index, part in enumerate(master_parts, start=2):
        worksheet.cell(row_index, 1).value = part
        key = _part_key(part)
        for col_index, group_name in enumerate(collection.groups.keys(), start=2):
            if key in group_lookup[group_name]:
                worksheet.cell(row_index, col_index).value = part

    return master_parts


def _add_detail_sheets(workbook, groups):
    used_names = {worksheet.title.upper() for worksheet in workbook.worksheets}
    for group_name, parts in groups.items():
        sheet_name = _safe_sheet_name(group_name, used_names)
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.cell(1, 1).value = "Part Name"
        for row_index, part in enumerate(parts, start=2):
            worksheet.cell(row_index, 1).value = part
        _style_simple_sheet(worksheet, len(parts) + 1, 1, auto_filter=True)


def _style_simple_sheet(worksheet, row_count, column_count, auto_filter):
    header_font = Font(name="Calibri", size=11, bold=True)
    body_font = Font(name="Calibri", size=11)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row in worksheet.iter_rows(min_row=2, max_row=max(row_count, 2), max_col=max(column_count, 1)):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for col_index in range(1, column_count + 1):
        letter = get_column_letter(col_index)
        max_length = 10
        for row_index in range(1, min(row_count, worksheet.max_row) + 1):
            value = worksheet.cell(row_index, col_index).value
            if value not in (None, ""):
                max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[letter].width = min(max_length + 2, 32)

    if auto_filter and row_count >= 1 and column_count >= 1:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(column_count)}{row_count}"


def _read_part_names_openpyxl(path, start_row, filter_na_column, exclude_board_parts):
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        worksheet = _find_sheet_case_insensitive(workbook, "BOM")
        if worksheet is None:
            raise ServiceError('Sheet "BOM" tidak ditemukan.', title="Format Excel tidak valid")

        values = []
        for row in worksheet.iter_rows(min_row=start_row, min_col=3, max_col=6 if filter_na_column else 3, values_only=True):
            part_value = row[0] if row else None
            if filter_na_column:
                marker_value = row[3] if len(row) > 3 else None
                if _is_na_marker(marker_value):
                    continue
            values.append(part_value)
        return _clean_part_values(values, exclude_board_parts)
    finally:
        workbook.close()


def _clean_part_values(values, exclude_board_parts):
    parts = []
    for value in values:
        part = _part_text(value)
        if not part:
            continue
        if _is_header_text(part):
            continue
        if exclude_board_parts and _is_excluded_part(part):
            continue
        parts.append(part)
    return parts


def _unique_parts(parts):
    unique = OrderedDict()
    for part in parts:
        key = _part_key(part)
        if key and key not in unique:
            unique[key] = part
    return list(unique.values())


def _sorted_master_parts(groups):
    master = OrderedDict()
    for parts in groups.values():
        for part in parts:
            key = _part_key(part)
            if key and key not in master:
                master[key] = part
    return sorted(master.values(), key=lambda value: value.upper())


def _excel_files_in_folder(folder):
    files = []
    try:
        entries = sorted(os.scandir(folder), key=lambda e: e.name.upper())
    except OSError:
        return files
    for entry in entries:
        if not entry.is_file():
            continue
        if entry.name.startswith("~$"):
            continue
        ext = os.path.splitext(entry.name)[1].lower()
        if ext in EXCEL_EXTENSIONS:
            files.append(Path(entry.path))
    return files


def _parse_pcb_part_numbers(text):
    values = []
    for token in re.split(r"[\s,;]+", str(text or "")):
        token = token.strip()
        if not token:
            continue
        match = re.search(r"EAX[A-Za-z0-9]{8}", token, re.IGNORECASE)
        value = (match.group(0) if match else token).upper()
        if value not in values:
            values.append(value)
    return values


def _map_pcb_folders(main_folder, pcb_numbers):
    targets = {pcb.upper(): pcb for pcb in pcb_numbers}
    folder_map = {pcb: [] for pcb in pcb_numbers}
    try:
        entries = sorted(os.scandir(main_folder), key=lambda e: e.name.upper())
    except OSError:
        return {}
    for entry in entries:
        if not entry.is_dir():
            continue

        child = Path(entry.path)
        name_upper = entry.name.upper()
        matched = None
        match = re.search(r"EAX[A-Za-z0-9]{8}", entry.name, re.IGNORECASE)
        if match:
            matched = targets.get(match.group(0).upper())
        if matched is None:
            matched = next((original for key, original in targets.items() if key in name_upper), None)

        if matched:
            folder_map[matched].append(child)

    return {pcb: folders for pcb, folders in folder_map.items() if folders}


def _extract_program_name(filename):
    match = re.search(r"\(([A-Za-z][\w-]*\d[\w-]*)", filename, re.IGNORECASE)
    if match:
        return match.group(1)
    return Path(filename).stem


def _unique_name(base_name, existing):
    clean_name = _clean_group_name(base_name) or "PROGRAM"
    if clean_name not in existing:
        return clean_name

    index = 2
    while f"{clean_name}_{index}" in existing:
        index += 1
    return f"{clean_name}_{index}"


def _clean_group_name(value):
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    return text.strip(" .")


def _safe_sheet_name(value, used_names):
    text = _clean_group_name(value) or "Sheet"
    text = re.sub(r"[\\/*?:\[\]]+", "", text).strip()
    text = text[:31] or "Sheet"
    candidate = text
    index = 2
    while candidate.upper() in used_names:
        suffix = f"_{index}"
        candidate = f"{text[:31 - len(suffix)]}{suffix}"
        index += 1
    used_names.add(candidate.upper())
    return candidate


def _find_sheet_case_insensitive(workbook, sheet_name):
    target = sheet_name.lower()
    for name in workbook.sheetnames:
        if name.lower() == target:
            return workbook[name]
    return None


def _part_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    text = text.replace("\r", " ").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def _part_key(value):
    return _part_text(value).upper()


def _is_header_text(value):
    return _part_key(value) in {
        "PART",
        "PART NAME",
        "PART NO",
        "PART NO.",
        "PART NUMBER",
        "PARTNUMBER",
        "PART_NAME",
        "P/N",
        "P/N COMPONENT",
    }


def _is_excluded_part(value):
    text = _part_key(value)
    if text.startswith("(") and text.endswith(")") and len(text) > 2:
        text = text[1:-1].strip()
    if len(text) >= 4 and text[:4] in {"EBU ", "EBT ", "EBR "}:
        return True
    return len(text) >= 3 and text[:3] in {"EAX", "EBU", "EBT", "EBR"}


def _is_na_marker(value):
    return _part_key(value) in {"#N/A", "#NA", "N/A"}


def _is_excel_error(value):
    return isinstance(value, int) and value < -1000000


def _ensure_2d(data):
    if data is None:
        return []
    if not isinstance(data, tuple):
        return ((data,),)
    if not data:
        return []
    if not isinstance(data[0], tuple):
        return (data,)
    return data


def _validate_config(config):
    if config.mode not in {MODE_PCB_LIST, MODE_PROGRAM_FOLDER}:
        raise ServiceError("Mode collect tidak valid.", title="Input tidak valid")
    if not config.source_folder:
        raise ServiceError("Folder source belum dipilih.", title="Input belum lengkap")
    if not Path(config.source_folder).is_dir():
        raise ServiceError(f"Folder source tidak ditemukan:\n{config.source_folder}", title="Folder tidak ditemukan")
    if config.mode == MODE_PCB_LIST and not _parse_pcb_part_numbers(config.pcb_part_numbers):
        raise ServiceError("List PCB Part Number belum diisi.", title="Input belum lengkap")
    if not config.output_path:
        raise ServiceError("Lokasi output belum dipilih.", title="Input belum lengkap")


def _normalize_output_path(path):
    output_path = Path(path)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")
    return output_path


def _open_workbook(excel, path, **kwargs):
    options = {
        "UpdateLinks": 0,
        "ReadOnly": True,
        "AddToMru": False,
        "IgnoreReadOnlyRecommended": True,
        "Notify": False,
    }
    options.update(kwargs)
    return _excel_call(lambda: excel.Workbooks.Open(path, **options))


def _close_workbook(workbook):
    try:
        _excel_call(lambda: workbook.Close(SaveChanges=False))
    except Exception:
        pass


def _configure_excel_for_background(excel):
    for attr, value in (
        ("Visible", False),
        ("DisplayAlerts", False),
        ("ScreenUpdating", False),
        ("EnableEvents", False),
        ("Calculation", XL_CALCULATION_MANUAL),
    ):
        _set_excel_option(excel, attr, value)


def _restore_and_quit_excel(excel):
    for attr, value in (
        ("Calculation", XL_CALCULATION_AUTOMATIC),
        ("ScreenUpdating", True),
        ("EnableEvents", True),
        ("DisplayAlerts", True),
    ):
        _set_excel_option(excel, attr, value)

    try:
        _excel_call(lambda: excel.Quit())
    except Exception:
        pass


def _set_excel_option(excel, attr, value):
    try:
        setattr(excel, attr, value)
    except Exception:
        pass


def _excel_call(fn, attempts=80, delay=0.25):
    for attempt in range(attempts):
        try:
            return fn()
        except Exception as exc:
            if not _is_excel_busy_error(exc) or attempt == attempts - 1:
                raise
            time.sleep(delay)
    return None


def _is_excel_busy_error(exc):
    hresult = getattr(exc, "hresult", None)
    if hresult is None and getattr(exc, "args", None):
        hresult = exc.args[0]
    return hresult in COM_BUSY_HRESULTS


def _error_message(exc):
    return getattr(exc, "message", str(exc))
