from __future__ import annotations

import math
import re
from collections import Counter, OrderedDict, defaultdict
from dataclasses import dataclass, field, replace
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services.errors import ServiceError
from utils.sort import natural_sort_key


MACHINE_GENERIC = "Generic"
MACHINE_CM602 = "CM602"
MACHINE_NPM_CUSTOM = "NPM Custom"
MACHINE_MODES = [MACHINE_GENERIC, MACHINE_CM602, MACHINE_NPM_CUSTOM]

REQUIRED_FIELDS = ["slot", "part_number", "component_insert"]
OPTIONAL_FIELDS = [
    "used",
    "feeder_width",
    "package",
    "nozzle",
    "head_type",
    "machine",
    "side_bank",
    "reserved_fixed",
]

FIELD_LABELS = OrderedDict(
    [
        ("slot", "SLOT"),
        ("part_number", "PART NUMBER"),
        ("component_insert", "COMPONENT INSERT"),
        ("used", "USED"),
        ("feeder_width", "FEEDER WIDTH"),
        ("package", "PACKAGE"),
        ("nozzle", "NOZZLE"),
        ("head_type", "HEAD TYPE"),
        ("machine", "MACHINE"),
        ("side_bank", "SIDE/BANK"),
        ("reserved_fixed", "RESERVED/FIXED"),
    ]
)

BALANCED_COLUMNS = [
    ("slot", "SLOT"),
    ("zone", "ZONE/TABLE/MODULE"),
    ("part_number", "PART NUMBER"),
    ("component_insert", "COMPONENT INSERT"),
    ("used", "USED"),
    ("effective_insert", "EFFECTIVE INSERT"),
    ("copy_no", "COPY NO"),
    ("assigned_zones", "ASSIGNED ZONES"),
    ("source_slot", "SOURCE SLOT"),
    ("note", "NOTE"),
]

ZONE_SUMMARY_COLUMNS = [
    ("zone", "ZONE/TABLE/MODULE"),
    ("slot_count", "SLOT COUNT"),
    ("usable_slot_count", "USABLE SLOTS"),
    ("assigned_feeder_count", "ASSIGNED FEEDERS"),
    ("effective_insert_total", "EFFECTIVE INSERT TOTAL"),
    ("target_effective_insert", "TARGET EFFECTIVE INSERT"),
    ("deviation", "DEVIATION"),
    ("raw_insert_visible", "RAW INSERT VISIBLE"),
]

DUPLICATE_PLAN_COLUMNS = [
    ("part_number", "PART NUMBER"),
    ("component_insert", "COMPONENT INSERT"),
    ("feeder_copies", "FEEDER COPIES"),
    ("assigned_zones", "ASSIGNED ZONES"),
    ("effective_insert_per_feeder", "EFFECTIVE INSERT PER FEEDER"),
    ("source_slot", "SOURCE SLOT"),
    ("reason", "REASON"),
]

WARNING_COLUMNS = [("severity", "SEVERITY"), ("message", "MESSAGE")]


@dataclass
class FeederBalancerConfig:
    source_path: str
    machine_mode: str = MACHINE_GENERIC
    column_mapping: dict = field(default_factory=dict)
    profile_text: str = ""


@dataclass
class FeederBalancerPreview:
    source_path: str
    source_file: str
    sheet_name: str
    columns: list[str]
    suggested_mapping: dict
    preview_rows: list[dict]
    row_count: int


@dataclass
class DetectedZoneResult:
    zone_records: list[dict]
    warnings: list[str]
    row_count: int
    slot_count: int
    part_count: int
    detected_zone_count: int


@dataclass
class FeederBalanceResult:
    source_path: str
    source_file: str
    machine_mode: str
    balanced_rows: list[dict]
    zone_summary_rows: list[dict]
    duplicate_plan_rows: list[dict]
    original_rows: list[dict]
    metrics: OrderedDict
    warnings: list[str]
    optimization_status: str
    output_path: str = ""


@dataclass
class _SourceTable:
    path: Path
    sheet_name: str
    headers: list[str]
    rows: list[dict]


@dataclass
class _ParsedSlot:
    original: str
    zone: str = ""
    slot_number: int | None = None
    side: str = ""
    parsed_zone: bool = False
    parsed_slot_number: bool = False
    parsed_side: bool = False
    parse_status: str = "UNPARSED"


@dataclass
class _SlotInfo:
    original_slot: str
    slot_key: str
    original_order: int
    parsed: _ParsedSlot
    source_rows: list[int] = field(default_factory=list)
    reserved_empty: bool = False

    @property
    def zone(self):
        return self.parsed.zone

    @zone.setter
    def zone(self, value):
        self.parsed.zone = str(value or "").strip()


@dataclass
class _PartDemand:
    key: str
    part_number: str
    insert: float
    insert_values: list[float] = field(default_factory=list)
    source_slots: list[str] = field(default_factory=list)
    source_slot_keys: list[str] = field(default_factory=list)
    fixed_slot_keys: list[str] = field(default_factory=list)
    optionals: dict = field(default_factory=dict)


@dataclass(frozen=True)
class _Assignment:
    assignment_id: int
    part_key: str
    zone: str
    fixed_slot_key: str = ""
    fixed: bool = False


@dataclass
class _BalanceInput:
    table: _SourceTable
    mapping: dict
    slots: OrderedDict
    parts: OrderedDict
    zone_order: list[str]
    zone_slots: OrderedDict
    warnings: list[str]
    profile: dict


def suggest_export_name(source_path=None):
    stem = _clean_filename_part(Path(source_path or "Feeder_Balancer").stem) or "Feeder_Balancer"
    return f"Feeder_Balancer_{stem}_{datetime.now().strftime('%y%m%d')}.xlsx"


def load_feeder_balancer_preview(source_path):
    table = _load_source_table(source_path)
    suggested_mapping = _suggest_column_mapping(table.headers)
    preview_rows = []
    for row in table.rows[:20]:
        preview_rows.append({header: row.get(header, "") for header in table.headers[:12]})

    return FeederBalancerPreview(
        source_path=str(table.path),
        source_file=table.path.name,
        sheet_name=table.sheet_name,
        columns=table.headers,
        suggested_mapping=suggested_mapping,
        preview_rows=preview_rows,
        row_count=len(table.rows),
    )


def detect_feeder_balancer_zones(config: FeederBalancerConfig):
    balance_input = _prepare_balance_input(config, require_parts=False)
    zone_records = _build_zone_records(balance_input.zone_slots, balance_input.slots)
    return DetectedZoneResult(
        zone_records=zone_records,
        warnings=balance_input.warnings,
        row_count=len(balance_input.table.rows),
        slot_count=len(balance_input.slots),
        part_count=len(balance_input.parts),
        detected_zone_count=len(balance_input.zone_order),
    )


def analyze_feeder_balance(config: FeederBalancerConfig, progress_callback=None):
    _emit_progress(progress_callback, 0, "Loading feeder data...")
    balance_input = _prepare_balance_input(config, require_parts=True)
    if not balance_input.parts:
        raise ServiceError("Tidak ada PART NUMBER valid yang bisa dibalance.", title="Data kosong")

    total_slot_capacity = len(balance_input.slots)
    unique_part_count = len(balance_input.parts)
    available_capacity = _available_capacity(balance_input.zone_slots)
    if unique_part_count > available_capacity:
        raise ServiceError(
            (
                "Kapasitas feeder tidak cukup.\n"
                f"Unique part: {unique_part_count}\n"
                f"Usable slot capacity: {available_capacity}\n"
                "Kurangi part atau tambahkan SLOT di input."
            ),
            title="Kapasitas tidak cukup",
        )

    _emit_progress(progress_callback, 35, "Balancing feeder load per zone...")
    assignments, planning_warnings = _build_assignments(balance_input)
    balance_input.warnings.extend(planning_warnings)

    _emit_progress(progress_callback, 65, "Assigning balanced plan to actual slots...")
    slot_assignments, copy_numbers = _assign_slots(balance_input, assignments)

    _emit_progress(progress_callback, 85, "Building summary tables...")
    balanced_rows = _build_balanced_rows(balance_input, assignments, slot_assignments, copy_numbers)
    zone_summary_rows, metrics = _build_summary(balance_input, assignments, total_slot_capacity, unique_part_count)
    duplicate_plan_rows = _build_duplicate_plan(balance_input, assignments)
    original_rows = _build_original_rows(balance_input.table)
    optimization_status = "DETERMINISTIC_HEURISTIC"

    _emit_progress(progress_callback, 100, "Feeder balance complete")
    return FeederBalanceResult(
        source_path=str(balance_input.table.path),
        source_file=balance_input.table.path.name,
        machine_mode=config.machine_mode or MACHINE_GENERIC,
        balanced_rows=balanced_rows,
        zone_summary_rows=zone_summary_rows,
        duplicate_plan_rows=duplicate_plan_rows,
        original_rows=original_rows,
        metrics=metrics,
        warnings=balance_input.warnings,
        optimization_status=optimization_status,
    )


def export_feeder_balance_result(result: FeederBalanceResult, output_path):
    if result is None:
        raise ServiceError("Belum ada hasil Feeder Balancer untuk diexport.", title="Data kosong")

    output = _normalize_output_path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    balanced_sheet = workbook.active
    balanced_sheet.title = "BALANCED_FEEDER"
    _write_records_sheet(
        balanced_sheet,
        result.balanced_rows,
        BALANCED_COLUMNS,
        duplicate_key="note",
        number_format_keys={"effective_insert": "0.000", "component_insert": "0.###"},
        text_keys={"slot", "source_slot"},
    )

    summary_sheet = workbook.create_sheet("SUMMARY")
    _write_summary_sheet(summary_sheet, result)

    duplicate_sheet = workbook.create_sheet("DUPLICATE_PLAN")
    _write_records_sheet(
        duplicate_sheet,
        result.duplicate_plan_rows,
        DUPLICATE_PLAN_COLUMNS,
        duplicate_key="feeder_copies",
        number_format_keys={"effective_insert_per_feeder": "0.000", "component_insert": "0.###"},
    )

    original_sheet = workbook.create_sheet("ORIGINAL_DATA")
    original_columns = [(key, key) for key in result.original_rows[0].keys()] if result.original_rows else []
    _write_records_sheet(original_sheet, result.original_rows, original_columns, text_keys={"SLOT"})

    if result.warnings:
        warning_sheet = workbook.create_sheet("WARNINGS")
        warning_rows = [{"severity": "WARNING", "message": warning} for warning in result.warnings]
        _write_records_sheet(warning_sheet, warning_rows, WARNING_COLUMNS, warning_rows=True)

    workbook.save(output)
    result.output_path = str(output)
    return str(output)


def _prepare_balance_input(config: FeederBalancerConfig, require_parts=True):
    table = _load_source_table(config.source_path)
    mapping = _resolve_mapping(table.headers, config.column_mapping)
    profile = _parse_profile_text(config.profile_text)
    warnings = []
    slots = OrderedDict()
    parts = OrderedDict()
    duplicate_slots = {}

    for source_index, row in enumerate(table.rows, start=1):
        source_row_number = int(row.get("_source_row_number", source_index))
        slot_text = _row_value(row, mapping.get("slot"))
        part_number = _row_value(row, mapping.get("part_number"))
        insert_text = _row_value(row, mapping.get("component_insert"))
        side_bank = _row_value(row, mapping.get("side_bank"))
        reserved_text = _row_value(row, mapping.get("reserved_fixed"))

        if not slot_text:
            warnings.append(f"Row {source_row_number}: SLOT kosong, row dilewati dari kapasitas feeder.")
            continue

        slot_key = _slot_key(slot_text)
        is_profile_reserved = slot_key in profile.get("reserved_slot_keys", set())
        is_reserved = _truthy_reserved(reserved_text) or is_profile_reserved

        parsed_slot = _parse_slot(slot_text, config.machine_mode, fallback_zone=side_bank)
        if slot_key not in slots:
            slots[slot_key] = _SlotInfo(
                original_slot=slot_text,
                slot_key=slot_key,
                original_order=len(slots),
                parsed=parsed_slot,
                source_rows=[source_row_number],
            )
        else:
            slots[slot_key].source_rows.append(source_row_number)
            duplicate_slots.setdefault(slot_key, slots[slot_key].source_rows[0])

        insert_value = _parse_number(insert_text)
        part_key = _part_key(part_number)

        if not part_key:
            if is_reserved:
                slots[slot_key].reserved_empty = True
            warnings.append(f"Row {source_row_number}: PART NUMBER kosong untuk SLOT {slot_text}. Slot tetap dihitung sebagai kapasitas.")
            continue

        if insert_value is None:
            warnings.append(f"Row {source_row_number}: COMPONENT INSERT tidak valid untuk {part_number}; part dilewati.")
            if is_reserved:
                slots[slot_key].reserved_empty = True
            continue

        part = parts.get(part_key)
        if part is None:
            part = _PartDemand(
                key=part_key,
                part_number=part_number,
                insert=insert_value,
                insert_values=[insert_value],
                source_slots=[],
                source_slot_keys=[],
                fixed_slot_keys=[],
                optionals={},
            )
            parts[part_key] = part
        else:
            part.insert_values.append(insert_value)
            if not _same_number(insert_value, part.insert):
                part.insert = max(part.insert, insert_value)

        if slot_text not in part.source_slots:
            part.source_slots.append(slot_text)
        if slot_key not in part.source_slot_keys:
            part.source_slot_keys.append(slot_key)
        if is_reserved and slot_key not in part.fixed_slot_keys:
            part.fixed_slot_keys.append(slot_key)

        for field_name in OPTIONAL_FIELDS:
            source_column = mapping.get(field_name)
            if not source_column:
                continue
            value = _row_value(row, source_column)
            if value and field_name not in part.optionals:
                part.optionals[field_name] = value

    for slot_key, first_row in duplicate_slots.items():
        rows = ", ".join(str(row_number) for row_number in slots[slot_key].source_rows)
        warnings.append(
            f"Duplicate SLOT {slots[slot_key].original_slot} ditemukan di rows {rows}. Kapasitas dihitung satu kali dari row {first_row}."
        )

    if not slots:
        raise ServiceError("Tidak ada SLOT valid di file input.", title="Data kosong")

    _finalize_zones(slots, warnings)
    zone_order, zone_slots = _build_zone_slots(slots)
    _validate_fixed_slots(parts, slots, warnings)
    _warn_insert_conflicts(parts, warnings)

    if require_parts and not parts:
        raise ServiceError("Tidak ada PART NUMBER valid setelah validasi input.", title="Data kosong")

    return _BalanceInput(
        table=table,
        mapping=mapping,
        slots=slots,
        parts=parts,
        zone_order=zone_order,
        zone_slots=zone_slots,
        warnings=warnings,
        profile=profile,
    )


def _build_assignments(balance_input: _BalanceInput):
    parts = balance_input.parts
    zone_order = balance_input.zone_order
    zone_caps = {zone: _usable_slot_count(slots) for zone, slots in balance_input.zone_slots.items()}
    warnings = []
    assignments = []
    next_id = 1

    for part in parts.values():
        fixed_slots = _valid_fixed_slot_keys(part, balance_input.slots)
        for slot_key in fixed_slots:
            zone = balance_input.slots[slot_key].zone
            assignments.append(_Assignment(next_id, part.key, zone, fixed_slot_key=slot_key, fixed=True))
            next_id += 1

    _validate_zone_capacity(assignments, zone_caps)

    for part in _sorted_parts(parts):
        if any(assignment.part_key == part.key for assignment in assignments):
            continue
        zone = _best_zone_for_part(part.key, assignments, parts, zone_order, zone_caps)
        if not zone:
            raise ServiceError(
                f"Tidak ada zone dengan kapasitas kosong untuk PART NUMBER {part.part_number}.",
                title="Arrangement tidak valid",
            )
        assignments.append(_Assignment(next_id, part.key, zone))
        next_id += 1

    _validate_zone_capacity(assignments, zone_caps)

    available_capacity = sum(zone_caps.values())
    duplicate_slots_to_use = max(0, available_capacity - len(assignments))
    max_copy = int(balance_input.profile.get("max_copy_per_part") or len(zone_order) or 1)
    max_copy = max(1, min(max_copy, len(zone_order) or 1))
    target = _target_insert(parts, zone_order)

    while duplicate_slots_to_use > 0:
        added = False
        for part in _sorted_parts(parts):
            copy_count = _copy_count(assignments, part.key)
            if copy_count >= max_copy:
                continue
            candidate_zones = _candidate_duplicate_zones(part.key, assignments, zone_order, zone_caps)
            if not candidate_zones:
                continue
            zone = _best_duplicate_zone(part.key, candidate_zones, assignments, parts, zone_order, target)
            assignments.append(_Assignment(next_id, part.key, zone))
            next_id += 1
            duplicate_slots_to_use -= 1
            added = True
            if duplicate_slots_to_use <= 0:
                break
        if not added:
            warnings.append(
                f"{duplicate_slots_to_use} spare slot tidak bisa dipakai untuk duplicate karena constraint zone/copy."
            )
            break

    improved = _local_improve(assignments, parts, zone_order, zone_caps, target)
    if improved != assignments:
        assignments = improved

    _validate_no_duplicate_part_zone(assignments, parts)
    _validate_zone_capacity(assignments, zone_caps)
    return assignments, warnings


def _assign_slots(balance_input: _BalanceInput, assignments: list[_Assignment]):
    slot_assignments = {}
    assignment_to_slot = {}

    fixed_slot_keys = set()
    for assignment in assignments:
        if not assignment.fixed_slot_key:
            continue
        if assignment.fixed_slot_key in slot_assignments:
            slot = balance_input.slots[assignment.fixed_slot_key].original_slot
            raise ServiceError(f"SLOT fixed {slot} dipakai lebih dari satu feeder.", title="Arrangement tidak valid")
        slot_assignments[assignment.fixed_slot_key] = assignment.assignment_id
        assignment_to_slot[assignment.assignment_id] = assignment.fixed_slot_key
        fixed_slot_keys.add(assignment.fixed_slot_key)

    assignments_by_zone = defaultdict(list)
    for assignment in assignments:
        if assignment.fixed:
            continue
        assignments_by_zone[assignment.zone].append(assignment)

    copy_counts = _copy_counts(assignments)
    for zone in balance_input.zone_order:
        free_slots = [
            slot
            for slot in _center_out_slots(balance_input.zone_slots[zone])
            if slot.slot_key not in fixed_slot_keys and not slot.reserved_empty
        ]
        feeder_assignments = sorted(
            assignments_by_zone.get(zone, []),
            key=lambda assignment: (
                -_effective_insert(balance_input.parts[assignment.part_key], copy_counts[assignment.part_key]),
                natural_sort_key(balance_input.parts[assignment.part_key].part_number),
                assignment.assignment_id,
            ),
        )
        if len(feeder_assignments) > len(free_slots):
            raise ServiceError(f"Zone {zone} tidak punya slot kosong cukup untuk hasil balancing.", title="Arrangement tidak valid")
        for slot, assignment in zip(free_slots, feeder_assignments):
            slot_assignments[slot.slot_key] = assignment.assignment_id
            assignment_to_slot[assignment.assignment_id] = slot.slot_key

    display_order = _slot_display_order_map(balance_input)
    assignments_by_part = defaultdict(list)
    for assignment in assignments:
        assignments_by_part[assignment.part_key].append(assignment)

    copy_numbers = {}
    for part_key, part_assignments in assignments_by_part.items():
        part_assignments.sort(
            key=lambda assignment: (
                _zone_index(balance_input.zone_order, assignment.zone),
                display_order.get(assignment_to_slot.get(assignment.assignment_id, ""), 999999),
                assignment.assignment_id,
            )
        )
        for copy_no, assignment in enumerate(part_assignments, start=1):
            copy_numbers[assignment.assignment_id] = copy_no

    return slot_assignments, copy_numbers


def _build_balanced_rows(balance_input, assignments, slot_assignments, copy_numbers):
    assignments_by_id = {assignment.assignment_id: assignment for assignment in assignments}
    copy_counts = _copy_counts(assignments)
    assigned_zones = _assigned_zones_by_part(assignments, balance_input.zone_order)
    rows = []

    for zone in balance_input.zone_order:
        for slot in _display_slots(balance_input.zone_slots[zone]):
            assignment_id = slot_assignments.get(slot.slot_key)
            if assignment_id is None:
                rows.append(
                    {
                        "slot": slot.original_slot,
                        "zone": zone,
                        "part_number": "",
                        "component_insert": "",
                        "used": "",
                        "effective_insert": "",
                        "copy_no": "",
                        "assigned_zones": "",
                        "source_slot": "",
                        "note": "RESERVED/FIXED EMPTY" if slot.reserved_empty else "SPARE",
                    }
                )
                continue

            assignment = assignments_by_id[assignment_id]
            part = balance_input.parts[assignment.part_key]
            used = copy_counts[part.key]
            note = "DUPLICATE/SPLIT" if used > 1 else "SINGLE"
            if assignment.fixed:
                note = f"{note}; FIXED"
            rows.append(
                {
                    "slot": slot.original_slot,
                    "zone": zone,
                    "part_number": part.part_number,
                    "component_insert": _round_number(part.insert),
                    "used": used,
                    "effective_insert": _round_number(_effective_insert(part, used), 3),
                    "copy_no": copy_numbers.get(assignment.assignment_id, ""),
                    "assigned_zones": ", ".join(assigned_zones.get(part.key, [])),
                    "source_slot": _join_limited(part.source_slots, 6),
                    "note": note,
                }
            )
    return rows


def _build_summary(balance_input, assignments, total_slot_capacity, unique_part_count):
    parts = balance_input.parts
    copy_counts = _copy_counts(assignments)
    loads = _zone_loads(assignments, parts, balance_input.zone_order)
    raw_loads = _zone_raw_loads(assignments, parts, balance_input.zone_order)
    target = _target_insert(parts, balance_input.zone_order)
    duplicate_feeder_count = max(0, len(assignments) - len(parts))
    spare_slot_count = max(0, total_slot_capacity - unique_part_count)
    max_deviation = max((abs(loads[zone] - target) for zone in balance_input.zone_order), default=0.0)

    zone_summary_rows = []
    for zone in balance_input.zone_order:
        slot_count = len(balance_input.zone_slots[zone])
        usable_count = _usable_slot_count(balance_input.zone_slots[zone])
        assigned_count = sum(1 for assignment in assignments if assignment.zone == zone)
        zone_summary_rows.append(
            {
                "zone": zone,
                "slot_count": slot_count,
                "usable_slot_count": usable_count,
                "assigned_feeder_count": assigned_count,
                "effective_insert_total": _round_number(loads[zone], 3),
                "target_effective_insert": _round_number(target, 3),
                "deviation": _round_number(loads[zone] - target, 3),
                "raw_insert_visible": _round_number(raw_loads[zone], 3),
            }
        )

    capacity_per_zone = OrderedDict((zone, len(balance_input.zone_slots[zone])) for zone in balance_input.zone_order)
    metrics = OrderedDict(
        [
            ("detected_zones", ", ".join(balance_input.zone_order)),
            ("capacity_per_zone", "; ".join(f"{zone}: {count}" for zone, count in capacity_per_zone.items())),
            ("total_feeder_capacity", total_slot_capacity),
            ("unique_part_count", unique_part_count),
            ("spare_slot_count", spare_slot_count),
            ("duplicate_feeder_count", duplicate_feeder_count),
            ("total_unique_component_insert", _round_number(sum(part.insert for part in parts.values()), 3)),
            ("target_effective_insert_per_zone", _round_number(target, 3)),
            ("max_deviation", _round_number(max_deviation, 3)),
            ("optimization_status", "DETERMINISTIC_HEURISTIC"),
            ("parsing_warnings", len(balance_input.warnings)),
        ]
    )

    unused_spares = sum(1 for row in zone_summary_rows if row["assigned_feeder_count"] < row["usable_slot_count"])
    if unused_spares:
        balance_input.warnings.append("Ada zone dengan spare slot kosong setelah balancing; lihat row NOTE=SPARE.")

    return zone_summary_rows, metrics


def _build_duplicate_plan(balance_input, assignments):
    copy_counts = _copy_counts(assignments)
    assigned_zones = _assigned_zones_by_part(assignments, balance_input.zone_order)
    rows = []
    for part in _sorted_parts(balance_input.parts):
        copies = copy_counts.get(part.key, 0)
        if copies <= 1:
            continue
        fixed_count = len(_valid_fixed_slot_keys(part, balance_input.slots))
        reason = "Fixed/reserved slot constraint" if fixed_count > 1 else "High COMPONENT INSERT prioritized for spare slot"
        rows.append(
            {
                "part_number": part.part_number,
                "component_insert": _round_number(part.insert),
                "feeder_copies": copies,
                "assigned_zones": ", ".join(assigned_zones.get(part.key, [])),
                "effective_insert_per_feeder": _round_number(_effective_insert(part, copies), 3),
                "source_slot": _join_limited(part.source_slots, 6),
                "reason": reason,
            }
        )
    return rows


def _build_zone_records(zone_slots, all_slots):
    rows = []
    for zone, slots in zone_slots.items():
        parsed_count = sum(1 for slot in slots if slot.parsed.parsed_zone)
        slot_number_count = sum(1 for slot in slots if slot.parsed.parsed_slot_number)
        side_count = sum(1 for slot in slots if slot.parsed.parsed_side)
        sample_slots = ", ".join(slot.original_slot for slot in slots[:6])
        if len(slots) > 6:
            sample_slots += f", +{len(slots) - 6} more"
        rows.append(
            {
                "zone": zone,
                "slot_count": len(slots),
                "usable_slot_count": _usable_slot_count(slots),
                "parsed_zone_count": parsed_count,
                "parsed_slot_number_count": slot_number_count,
                "parsed_side_count": side_count,
                "parse_status": "OK" if parsed_count == len(slots) else "PARTIAL",
                "sample_slots": sample_slots,
            }
        )
    return rows


def _build_original_rows(table):
    rows = []
    for row in table.rows:
        output = OrderedDict()
        output["SOURCE ROW"] = row.get("_source_row_number", "")
        for header in table.headers:
            output[header] = row.get(header, "")
        rows.append(output)
    return rows


def _load_source_table(source_path):
    path = Path(str(source_path or "").strip().replace('"', "").replace("'", ""))
    if not path.is_file():
        raise ServiceError(f"File input tidak ditemukan:\n{path}", title="File tidak ditemukan")

    suffix = path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xls", ".csv"}:
        raise ServiceError("Input Feeder Balancer harus Excel .xlsx/.xls/.xlsm atau CSV.", title="Format tidak valid")

    matrix, sheet_name = _read_source_matrix(path)
    if not matrix:
        raise ServiceError("File input kosong.", title="Data kosong")

    header_index = _detect_header_index(matrix)
    raw_headers = [_cell_text(value) for value in matrix[header_index]]
    headers = _unique_headers(raw_headers)
    rows = []
    for offset, raw_row in enumerate(matrix[header_index + 1 :], start=header_index + 2):
        values = [_cell_text(value) for value in raw_row]
        if not any(values):
            continue
        record = {"_source_row_number": offset}
        for index, header in enumerate(headers):
            record[header] = values[index] if index < len(values) else ""
        rows.append(record)

    if not rows:
        raise ServiceError("Tidak ada data setelah header input.", title="Data kosong")
    return _SourceTable(path=path, sheet_name=sheet_name, headers=headers, rows=rows)


def _read_source_matrix(path):
    try:
        import pandas as pd
    except ImportError as exc:
        raise ServiceError("Feeder Balancer membutuhkan pandas untuk membaca Excel/CSV.", title="Dependency belum lengkap") from exc

    suffix = path.suffix.lower()
    if suffix == ".csv":
        last_error = None
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
            try:
                dataframe = pd.read_csv(
                    path,
                    header=None,
                    dtype=object,
                    na_filter=False,
                    sep=None,
                    engine="python",
                    encoding=encoding,
                )
                return dataframe.values.tolist(), "CSV"
            except Exception as exc:
                last_error = exc
        raise ServiceError("CSV tidak bisa dibaca. Cek delimiter atau encoding file.", title="CSV tidak valid") from last_error

    try:
        excel = pd.ExcelFile(path)
        sheet_name = excel.sheet_names[0]
        dataframe = pd.read_excel(excel, sheet_name=sheet_name, header=None, dtype=object, na_filter=False)
        excel.close()
        return dataframe.values.tolist(), sheet_name
    except Exception as exc:
        raise ServiceError("File Excel tidak bisa dibaca.", title="Excel tidak valid") from exc


def _detect_header_index(matrix):
    best_index = 0
    best_score = -1
    for index, row in enumerate(matrix[:30]):
        values = [_cell_text(value) for value in row]
        normalized = {_normalize_header(value) for value in values if _normalize_header(value)}
        score = 0
        for field_name in REQUIRED_FIELDS:
            if normalized & _alias_set(field_name):
                score += 3
        for field_name in OPTIONAL_FIELDS:
            if normalized & _alias_set(field_name):
                score += 1
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _unique_headers(headers):
    output = []
    seen = Counter()
    for index, header in enumerate(headers, start=1):
        name = header or f"Column {index}"
        seen[name] += 1
        if seen[name] > 1:
            name = f"{name} ({seen[name]})"
        output.append(name)
    return output


def _suggest_column_mapping(headers):
    mapping = {}
    normalized_headers = [(_normalize_header(header), header) for header in headers]
    for field_name in list(FIELD_LABELS.keys()):
        aliases = _alias_set(field_name)
        for normalized, header in normalized_headers:
            if normalized in aliases:
                mapping[field_name] = header
                break
        if field_name not in mapping:
            mapping[field_name] = ""
    return mapping


def _resolve_mapping(headers, user_mapping):
    user_mapping = dict(user_mapping or {})
    suggested = _suggest_column_mapping(headers)
    header_lookup = {header: header for header in headers}
    normalized_lookup = {_normalize_header(header): header for header in headers}
    mapping = {}

    for field_name in list(FIELD_LABELS.keys()):
        selected = str(user_mapping.get(field_name) or "").strip()
        if selected in {"", "-", "Not used"}:
            selected = suggested.get(field_name, "")
        if selected and selected not in header_lookup:
            selected = normalized_lookup.get(_normalize_header(selected), "")
        mapping[field_name] = selected

    missing = [FIELD_LABELS[field_name] for field_name in REQUIRED_FIELDS if not mapping.get(field_name)]
    if missing:
        raise ServiceError(
            "Kolom wajib belum lengkap: " + ", ".join(missing) + ". Gunakan Column Mapping untuk memilih kolom yang benar.",
            title="Missing required column",
        )
    return mapping


def _parse_slot(slot_text, machine_mode, fallback_zone=""):
    text = _cell_text(slot_text)
    parsed = _ParsedSlot(original=text)
    if not text:
        return parsed

    bracket = re.search(r"\[([^\]]+)\]\s*(\d+)?\s*([A-Za-z]+)?", text)
    if bracket:
        parsed.zone = bracket.group(1).strip()
        parsed.parsed_zone = bool(parsed.zone)
        if bracket.group(2):
            parsed.slot_number = int(bracket.group(2))
            parsed.parsed_slot_number = True
        if bracket.group(3):
            parsed.side = _normalize_side(bracket.group(3))
            parsed.parsed_side = bool(parsed.side)
        parsed.parse_status = "CM602" if machine_mode == MACHINE_CM602 else "BRACKET"

    if not parsed.zone:
        label = re.search(r"\b(?:zone|table|module|mod|head|bank)\s*[:#=\-]?\s*([A-Za-z0-9]+)\b", text, flags=re.IGNORECASE)
        if label:
            parsed.zone = label.group(1).strip()
            parsed.parsed_zone = True
            parsed.parse_status = "LABELED"

    if parsed.slot_number is None:
        slot_label = re.search(r"\b(?:slot|pu|pos|position)\s*[:#=\-]?\s*(\d+)\b", text, flags=re.IGNORECASE)
        if slot_label:
            parsed.slot_number = int(slot_label.group(1))
            parsed.parsed_slot_number = True

    if not parsed.zone or parsed.slot_number is None:
        compact = re.match(
            r"^\s*([A-Za-z]*\d+[A-Za-z]?|\d+)\s*[\-_:/ ]+\s*(?:S(?:LOT)?\s*)?(\d+)\s*([A-Za-z]+)?\s*$",
            text,
            flags=re.IGNORECASE,
        )
        if compact:
            if not parsed.zone:
                parsed.zone = compact.group(1).strip()
                parsed.parsed_zone = True
                parsed.parse_status = "PREFIX"
            if parsed.slot_number is None:
                parsed.slot_number = int(compact.group(2))
                parsed.parsed_slot_number = True
            if not parsed.side and compact.group(3):
                parsed.side = _normalize_side(compact.group(3))
                parsed.parsed_side = bool(parsed.side)

    if parsed.slot_number is None:
        trailing_slot = re.search(r"(\d+)\s*([A-Za-z]+)?\s*$", text)
        if trailing_slot:
            parsed.slot_number = int(trailing_slot.group(1))
            parsed.parsed_slot_number = True
            if not parsed.side and trailing_slot.group(2):
                parsed.side = _normalize_side(trailing_slot.group(2))
                parsed.parsed_side = bool(parsed.side)

    if not parsed.side:
        side_label = re.search(r"\b(?:side|position)\s*[:#=\-]?\s*([A-Za-z]+)\b", text, flags=re.IGNORECASE)
        if side_label:
            parsed.side = _normalize_side(side_label.group(1))
            parsed.parsed_side = bool(parsed.side)

    if not parsed.zone and fallback_zone:
        parsed.zone = str(fallback_zone).strip()
        parsed.parse_status = "ZONE_FROM_SIDE_BANK"

    if parsed.zone and parsed.parse_status == "UNPARSED":
        parsed.parse_status = "PARTIAL"
    return parsed


def _finalize_zones(slots, warnings):
    parsed_or_fallback = [slot for slot in slots.values() if slot.zone]
    if not parsed_or_fallback:
        raise ServiceError(
            (
                "Zone parsing gagal total. Mapping kolom SIDE/BANK jika tersedia, atau tambahkan prefix zone/table/module "
                "di kolom SLOT sebelum Generate Balance."
            ),
            title="Zone parsing gagal",
        )

    missing = [slot for slot in slots.values() if not slot.zone]
    if missing:
        for slot in missing:
            slot.zone = "UNPARSED"
        sample = ", ".join(slot.original_slot for slot in missing[:8])
        if len(missing) > 8:
            sample += f", +{len(missing) - 8} more"
        warnings.append(f"Slot parsing gagal sebagian untuk zone: {sample}. Slot tersebut masuk zone UNPARSED.")

    slot_number_missing = [slot for slot in slots.values() if not slot.parsed.parsed_slot_number]
    if slot_number_missing:
        sample = ", ".join(slot.original_slot for slot in slot_number_missing[:8])
        if len(slot_number_missing) > 8:
            sample += f", +{len(slot_number_missing) - 8} more"
        warnings.append(f"Slot number gagal diparse sebagian: {sample}. Urutan original dipakai untuk slot tersebut.")


def _build_zone_slots(slots):
    zone_slots = OrderedDict()
    for slot in slots.values():
        zone_slots.setdefault(slot.zone, []).append(slot)
    return list(zone_slots.keys()), zone_slots


def _validate_fixed_slots(parts, slots, warnings):
    for part in parts.values():
        fixed_zones = []
        valid_fixed_slots = []
        for slot_key in part.fixed_slot_keys:
            slot = slots.get(slot_key)
            if slot is None:
                continue
            valid_fixed_slots.append(slot_key)
            fixed_zones.append(slot.zone)
        duplicate_zones = [zone for zone, count in Counter(fixed_zones).items() if count > 1]
        if duplicate_zones:
            raise ServiceError(
                (
                    f"PART NUMBER {part.part_number} punya lebih dari satu RESERVED/FIXED slot di zone yang sama: "
                    + ", ".join(duplicate_zones)
                ),
                title="Duplicate part di zone sama",
            )
        if len(valid_fixed_slots) > 1:
            warnings.append(f"{part.part_number}: RESERVED/FIXED membuat minimal {len(valid_fixed_slots)} feeder copy.")


def _warn_insert_conflicts(parts, warnings):
    for part in parts.values():
        rounded_values = {_round_number(value, 6) for value in part.insert_values}
        if len(rounded_values) > 1:
            values = ", ".join(str(value) for value in sorted(rounded_values))
            warnings.append(
                f"{part.part_number}: COMPONENT INSERT berbeda antar row ({values}). Feeder Balancer memakai nilai terbesar: {_round_number(part.insert, 6)}."
            )


def _best_zone_for_part(part_key, assignments, parts, zone_order, zone_caps):
    used_zones = _zones_for_part(assignments, part_key)
    loads = _zone_loads(assignments, parts, zone_order)
    counts = _zone_counts(assignments, zone_order)
    candidates = [
        zone
        for zone in zone_order
        if zone not in used_zones and counts[zone] < zone_caps.get(zone, 0)
    ]
    if not candidates:
        return ""
    return min(
        candidates,
        key=lambda zone: (
            loads[zone],
            counts[zone],
            natural_sort_key(zone),
        ),
    )


def _candidate_duplicate_zones(part_key, assignments, zone_order, zone_caps):
    used_zones = _zones_for_part(assignments, part_key)
    counts = _zone_counts(assignments, zone_order)
    return [
        zone
        for zone in zone_order
        if zone not in used_zones and counts[zone] < zone_caps.get(zone, 0)
    ]


def _best_duplicate_zone(part_key, candidate_zones, assignments, parts, zone_order, target):
    best_zone = None
    best_score = None
    next_id = max((assignment.assignment_id for assignment in assignments), default=0) + 1
    for zone in candidate_zones:
        candidate = list(assignments) + [_Assignment(next_id, part_key, zone)]
        score = _objective(candidate, parts, zone_order, target)
        tie = (score, natural_sort_key(zone))
        if best_score is None or tie < best_score:
            best_score = tie
            best_zone = zone
    return best_zone or candidate_zones[0]


def _local_improve(assignments, parts, zone_order, zone_caps, target):
    current = list(assignments)
    best_score = _objective(current, parts, zone_order, target)

    for _ in range(60):
        improved = False
        for index, assignment in enumerate(list(current)):
            if assignment.fixed:
                continue
            for zone in zone_order:
                if zone == assignment.zone:
                    continue
                if not _can_move_to_zone(current, index, zone, zone_order, zone_caps):
                    continue
                candidate = list(current)
                candidate[index] = replace(assignment, zone=zone)
                score = _objective(candidate, parts, zone_order, target)
                if score < best_score:
                    current = candidate
                    best_score = score
                    improved = True
                    break
            if improved:
                break
        if improved:
            continue

        for left in range(len(current)):
            if current[left].fixed:
                continue
            for right in range(left + 1, len(current)):
                if current[right].fixed or current[left].zone == current[right].zone:
                    continue
                if not _can_swap(current, left, right):
                    continue
                candidate = list(current)
                candidate[left] = replace(current[left], zone=current[right].zone)
                candidate[right] = replace(current[right], zone=current[left].zone)
                score = _objective(candidate, parts, zone_order, target)
                if score < best_score:
                    current = candidate
                    best_score = score
                    improved = True
                    break
            if improved:
                break
        if not improved:
            break
    return current


def _can_move_to_zone(assignments, index, target_zone, zone_order, zone_caps):
    assignment = assignments[index]
    counts = _zone_counts(assignments, zone_order)
    if counts[target_zone] >= zone_caps.get(target_zone, 0):
        return False
    for other_index, other in enumerate(assignments):
        if other_index == index:
            continue
        if other.part_key == assignment.part_key and other.zone == target_zone:
            return False
    return True


def _can_swap(assignments, left, right):
    left_assignment = assignments[left]
    right_assignment = assignments[right]
    for index, other in enumerate(assignments):
        if index in {left, right}:
            continue
        if other.part_key == left_assignment.part_key and other.zone == right_assignment.zone:
            return False
        if other.part_key == right_assignment.part_key and other.zone == left_assignment.zone:
            return False
    return True


def _objective(assignments, parts, zone_order, target):
    loads = _zone_loads(assignments, parts, zone_order)
    deviations = [abs(loads[zone] - target) for zone in zone_order]
    max_deviation = max(deviations, default=0.0)
    total_deviation = sum(deviations)
    spread = max(loads.values(), default=0.0) - min(loads.values(), default=0.0)
    return (
        round(max_deviation, 9),
        round(total_deviation, 9),
        round(spread, 9),
        tuple(round(loads[zone], 9) for zone in zone_order),
    )


def _zone_loads(assignments, parts, zone_order):
    loads = {zone: 0.0 for zone in zone_order}
    copy_counts = _copy_counts(assignments)
    for assignment in assignments:
        part = parts[assignment.part_key]
        loads[assignment.zone] += _effective_insert(part, copy_counts[assignment.part_key])
    return loads


def _zone_raw_loads(assignments, parts, zone_order):
    loads = {zone: 0.0 for zone in zone_order}
    for assignment in assignments:
        loads[assignment.zone] += parts[assignment.part_key].insert
    return loads


def _zone_counts(assignments, zone_order):
    counts = {zone: 0 for zone in zone_order}
    for assignment in assignments:
        counts[assignment.zone] = counts.get(assignment.zone, 0) + 1
    return counts


def _copy_counts(assignments):
    counts = Counter()
    for assignment in assignments:
        counts[assignment.part_key] += 1
    return counts


def _copy_count(assignments, part_key):
    return sum(1 for assignment in assignments if assignment.part_key == part_key)


def _zones_for_part(assignments, part_key):
    return {assignment.zone for assignment in assignments if assignment.part_key == part_key}


def _assigned_zones_by_part(assignments, zone_order):
    zones = defaultdict(list)
    for assignment in assignments:
        if assignment.zone not in zones[assignment.part_key]:
            zones[assignment.part_key].append(assignment.zone)
    for part_key in zones:
        zones[part_key].sort(key=lambda zone: _zone_index(zone_order, zone))
    return zones


def _validate_zone_capacity(assignments, zone_caps):
    counts = Counter(assignment.zone for assignment in assignments)
    for zone, count in counts.items():
        if count > zone_caps.get(zone, 0):
            raise ServiceError(f"Zone {zone} melebihi kapasitas slot ({count}/{zone_caps.get(zone, 0)}).", title="Arrangement tidak valid")


def _validate_no_duplicate_part_zone(assignments, parts):
    seen = set()
    for assignment in assignments:
        key = (assignment.part_key, assignment.zone)
        if key in seen:
            part_number = parts[assignment.part_key].part_number
            raise ServiceError(
                f"PART NUMBER {part_number} terpasang lebih dari 1 kali di zone {assignment.zone}.",
                title="Duplicate part di zone sama",
            )
        seen.add(key)


def _center_out_slots(slots):
    if not slots:
        return []
    if not all(slot.parsed.parsed_slot_number for slot in slots):
        return list(sorted(slots, key=lambda slot: slot.original_order))

    grouped = defaultdict(list)
    for slot in slots:
        grouped[slot.parsed.slot_number].append(slot)
    numbers = sorted(grouped)
    center_index = (len(numbers) - 1) // 2
    ordered_numbers = []
    for offset in range(len(numbers)):
        if offset == 0:
            ordered_numbers.append(numbers[center_index])
            continue
        left_index = center_index - offset
        right_index = center_index + offset
        if left_index >= 0:
            ordered_numbers.append(numbers[left_index])
        if right_index < len(numbers):
            ordered_numbers.append(numbers[right_index])

    ordered_slots = []
    for number in ordered_numbers:
        ordered_slots.extend(sorted(grouped[number], key=_slot_side_order))
    return ordered_slots


def _display_slots(slots):
    if not slots:
        return []
    if all(slot.parsed.parsed_slot_number for slot in slots):
        return sorted(slots, key=lambda slot: (slot.parsed.slot_number, _slot_side_order(slot), natural_sort_key(slot.original_slot)))
    return sorted(slots, key=lambda slot: slot.original_order)


def _slot_display_order_map(balance_input):
    order = {}
    index = 0
    for zone in balance_input.zone_order:
        for slot in _display_slots(balance_input.zone_slots[zone]):
            order[slot.slot_key] = index
            index += 1
    return order


def _slot_side_order(slot):
    side = str(slot.parsed.side or "").upper()
    side_rank = {
        "L": 0,
        "A": 0,
        "F": 0,
        "FRONT": 0,
        "R": 1,
        "B": 1,
        "REAR": 1,
    }.get(side, 5)
    return (side_rank, natural_sort_key(side), slot.original_order)


def _valid_fixed_slot_keys(part, slots):
    return [slot_key for slot_key in part.fixed_slot_keys if slot_key in slots and not slots[slot_key].reserved_empty]


def _available_capacity(zone_slots):
    return sum(_usable_slot_count(slots) for slots in zone_slots.values())


def _usable_slot_count(slots):
    return sum(1 for slot in slots if not slot.reserved_empty)


def _target_insert(parts, zone_order):
    if not zone_order:
        return 0.0
    return sum(part.insert for part in parts.values()) / len(zone_order)


def _effective_insert(part, used):
    return part.insert / max(1, used)


def _sorted_parts(parts):
    return sorted(parts.values(), key=lambda part: (-part.insert, natural_sort_key(part.part_number)))


def _zone_index(zone_order, zone):
    try:
        return zone_order.index(zone)
    except ValueError:
        return len(zone_order) + 999


def _write_summary_sheet(worksheet, result):
    worksheet.append(["ITEM", "VALUE"])
    for key, value in result.metrics.items():
        worksheet.append([key.replace("_", " ").upper(), value])
    worksheet.append([])
    worksheet.append(["ZONE SUMMARY", ""])
    header_row = worksheet.max_row + 1
    worksheet.append([header for _, header in ZONE_SUMMARY_COLUMNS])
    for row in result.zone_summary_rows:
        worksheet.append([row.get(key, "") for key, _ in ZONE_SUMMARY_COLUMNS])
    _style_sheet(worksheet, header_rows={1, header_row}, number_format_headers={"EFFECTIVE INSERT TOTAL", "TARGET EFFECTIVE INSERT", "DEVIATION", "RAW INSERT VISIBLE"})


def _write_records_sheet(
    worksheet,
    rows,
    columns,
    duplicate_key=None,
    warning_rows=False,
    number_format_keys=None,
    text_keys=None,
):
    number_format_keys = number_format_keys or {}
    text_keys = text_keys or set()
    if not columns:
        worksheet.append(["NO DATA"])
        _style_sheet(worksheet)
        return
    worksheet.append([header for _, header in columns])
    for row in rows:
        worksheet.append([row.get(key, "") for key, _ in columns])

    duplicate_columns = set()
    if duplicate_key:
        for row_number, row in enumerate(rows, start=2):
            value = str(row.get(duplicate_key, "")).upper()
            if "DUPLICATE" in value or "SPLIT" in value:
                duplicate_columns.add(row_number)
            elif duplicate_key == "feeder_copies" and _safe_int(value) > 1:
                duplicate_columns.add(row_number)

    key_to_column = {key: index for index, (key, _) in enumerate(columns, start=1)}
    for key, number_format in number_format_keys.items():
        column_index = key_to_column.get(key)
        if not column_index:
            continue
        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(row_index, column_index).number_format = number_format

    for key in text_keys:
        column_index = key_to_column.get(key)
        if not column_index:
            continue
        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(row_index, column_index).number_format = "@"

    _style_sheet(worksheet, duplicate_rows=duplicate_columns, warning_rows=warning_rows)


def _style_sheet(worksheet, header_rows=None, duplicate_rows=None, warning_rows=False, number_format_headers=None):
    header_rows = header_rows or {1}
    duplicate_rows = duplicate_rows or set()
    number_format_headers = number_format_headers or set()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color="1F2937")
    duplicate_fill = PatternFill("solid", fgColor="FFF2CC")
    warning_fill = PatternFill("solid", fgColor="FCE4D6")

    for row_index in range(1, worksheet.max_row + 1):
        for cell in worksheet[row_index]:
            if row_index in header_rows:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = body_font
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                if row_index in duplicate_rows:
                    cell.fill = duplicate_fill
                if warning_rows and row_index >= 2:
                    cell.fill = warning_fill

    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"

    header_lookup = {}
    for column_index in range(1, worksheet.max_column + 1):
        header_lookup[column_index] = str(worksheet.cell(1, column_index).value or "")

    for column_index in range(1, worksheet.max_column + 1):
        letter = get_column_letter(column_index)
        max_length = 10
        for row_index in range(1, min(worksheet.max_row, 250) + 1):
            value = worksheet.cell(row_index, column_index).value
            if value not in (None, ""):
                max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[letter].width = min(max_length + 2, 48)
        if header_lookup.get(column_index) in number_format_headers:
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row_index, column_index).number_format = "0.000"


def _parse_profile_text(profile_text):
    profile = {"reserved_slot_keys": set()}
    for raw_line in str(profile_text or "").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = [part.strip() for part in line.split("=", 1)]
        normalized_key = _normalize_header(key)
        if normalized_key == "maxcopyperpart":
            number = _safe_int(value)
            if number > 0:
                profile["max_copy_per_part"] = number
        elif normalized_key == "reservedslots":
            profile["reserved_slot_keys"].update(_slot_key(item) for item in re.split(r"[,;|]+", value) if _slot_key(item))
    return profile


def _normalize_output_path(path):
    output_path = Path(path)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")
    return output_path


def _alias_set(field_name):
    aliases = {
        "slot": {
            "slot",
            "location",
            "locationcode",
            "feederlocation",
            "feederslot",
            "fixedslot",
            "pu",
            "pickupunit",
        },
        "part_number": {
            "partnumber",
            "partno",
            "part",
            "pn",
            "pncomponent",
            "componentpn",
            "componentpartnumber",
            "partname",
        },
        "component_insert": {
            "componentinsert",
            "insert",
            "insertpoint",
            "insertcount",
            "insertqty",
            "mountcount",
            "usage",
            "qty",
            "quantity",
            "totalinsert",
        },
        "used": {"used", "use", "feedercopies", "copies"},
        "feeder_width": {"feederwidth", "width", "feedersize"},
        "package": {"package", "pkg"},
        "nozzle": {"nozzle"},
        "head_type": {"headtype", "head"},
        "machine": {"machine", "machinemodel"},
        "side_bank": {"sidebank", "side", "bank", "zone", "table", "module", "feederbank"},
        "reserved_fixed": {"reservedfixed", "reserved", "fixed", "fixedslot", "lock", "locked", "pin", "pinned"},
    }
    return aliases.get(field_name, set())


def _normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "", _cell_text(value).lower())


def _cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
    text = str(value)
    text = text.replace("\r", " ").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def _row_value(row, column_name):
    if not column_name:
        return ""
    return _cell_text(row.get(column_name, ""))


def _part_key(value):
    return _cell_text(value).upper()


def _slot_key(value):
    return _cell_text(value).upper()


def _parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)

    text = _cell_text(value)
    if not text:
        return None
    text = re.sub(r"[^0-9,.\-]+", "", text)
    if not text or text in {"-", ".", ","}:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.split(",")
        if len(parts[-1]) == 3 and len(parts) > 1:
            text = "".join(parts)
        else:
            text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _same_number(left, right):
    return abs(float(left) - float(right)) < 0.000001


def _truthy_reserved(value):
    text = _cell_text(value).upper()
    return text in {
        "1",
        "Y",
        "YES",
        "TRUE",
        "T",
        "FIXED",
        "RESERVED",
        "RESERVE",
        "LOCK",
        "LOCKED",
        "PIN",
        "PINNED",
        "R",
        "F",
    }


def _normalize_side(value):
    text = _cell_text(value).upper()
    if text in {"LEFT", "L"}:
        return "L"
    if text in {"RIGHT", "R"}:
        return "R"
    if text in {"FRONT", "F"}:
        return "F"
    if text in {"REAR"}:
        return "REAR"
    if text:
        return text[:8]
    return ""


def _round_number(value, digits=3):
    if value == "":
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    if abs(number - round(number)) < 0.0000001:
        return int(round(number))
    return round(number, digits)


def _join_limited(values, limit):
    values = [str(value) for value in values if str(value)]
    if not values:
        return ""
    if len(values) <= limit:
        return ", ".join(values)
    return ", ".join(values[:limit]) + f", +{len(values) - limit} more"


def _safe_int(value):
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return 0


def _clean_filename_part(value):
    text = str(value or "").strip()
    text = re.sub(r'[<>:"/\\|?*]+', "_", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" .")


def _emit_progress(progress_callback, percent, message):
    if progress_callback:
        progress_callback(percent, message)
