import os
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime
from itertools import combinations
from pathlib import Path

from openpyxl import Workbook

from services.common_feeder_reuse_service import (
    EXCEL_EXTENSIONS,
    _error_message,
    _normalize_output_path,
    _part_key,
    _read_bom_parts,
    _style_sheet,
)
from services.component_usage_finder_service import format_pcb_part_number, parse_pcb_part_number
from services.errors import ServiceError


STATUS_GROUPED = "GROUPED"
STATUS_SINGLE = "SINGLE"
STATUS_OPTIONS = ["SHOW ALL", STATUS_GROUPED, STATUS_SINGLE]

GROUP_COLUMNS = [
    ("group_id", "Group"),
    ("status", "Status"),
    ("member_count", "PCB Count"),
    ("avg_similarity_percent", "Avg Similarity %"),
    ("min_similarity_percent", "Min Similarity %"),
    ("members", "PCB Members"),
]

PAIR_COLUMNS = [
    ("status", "Status"),
    ("model_a", "PCB A"),
    ("model_b", "PCB B"),
    ("similarity_percent", "Similarity %"),
    ("jaccard_percent", "Jaccard %"),
    ("shared_component_count", "Shared Parts"),
    ("model_a_component_count", "A Parts"),
    ("model_b_component_count", "B Parts"),
    ("shared_components", "Shared Component P/N"),
]

GROUP_COMPONENT_COLUMNS = [
    ("group_id", "Group"),
    ("component_part_number", "Component P/N"),
    ("status", "Status"),
    ("used_in_count", "Used In PCB"),
    ("member_count", "PCB Count"),
    ("coverage_percent", "Coverage %"),
    ("members", "PCB Members"),
]

MODEL_COLUMNS = [
    ("pcb_part_number", "PCB Part Number"),
    ("component_count", "Component Count"),
    ("excel_file_count", "Excel Files"),
    ("source_folder", "Source Folder"),
    ("source_files", "Source Files"),
    ("components", "Component P/N"),
]


@dataclass
class ModelFeederGroupConfig:
    source_folder: str
    min_similarity_percent: int = 70
    min_shared_components: int = 20
    target_pcb_list: list = None


@dataclass
class ModelUsage:
    key: str
    display_name: str
    pcb_part_number: str
    source_folder: str
    source_files: list[str]
    components: OrderedDict
    component_frequencies: dict = None
    insert_averages: dict = None
    variant_components: dict = None

    @property
    def component_count(self):
        return len(self.components)


@dataclass
class ModelFeederGroupResult:
    group_rows: list[dict]
    pair_rows: list[dict]
    model_rows: list[dict]
    total_files: int
    read_files: int
    skipped_files: list[str]
    model_count: int
    group_count: int
    single_count: int
    min_similarity_percent: int
    min_shared_components: int


def analyze_model_feeder_groups(config: ModelFeederGroupConfig, progress_callback=None):
    _validate_config(config)
    _emit_progress(progress_callback, 0, "Scanning PCB folders...")

    models, total_files, read_files, skipped_files = _scan_models(config.source_folder, config.target_pcb_list, progress_callback)
    if not models:
        return ModelFeederGroupResult(
            group_rows=[],
            pair_rows=[],
            model_rows=[],
            total_files=total_files,
            read_files=read_files,
            skipped_files=skipped_files,
            model_count=0,
            group_count=0,
            single_count=0,
            min_similarity_percent=config.min_similarity_percent,
            min_shared_components=config.min_shared_components,
        )

    _emit_progress(progress_callback, 96, "Calculating PCB similarity...")
    pair_rows, pair_lookup = _build_pair_rows(models, config)

    _emit_progress(progress_callback, 98, "Building recommended fixed-feeder groups...")
    groups = _build_groups(models, pair_rows, pair_lookup)
    group_rows = _build_group_outputs(groups, models, pair_lookup)
    model_rows = _build_model_rows(models)
    group_count = sum(1 for row in group_rows if row["status"] == STATUS_GROUPED)
    single_count = sum(1 for row in group_rows if row["status"] == STATUS_SINGLE)

    _emit_progress(progress_callback, 100, "Analysis complete")
    return ModelFeederGroupResult(
        group_rows=group_rows,
        pair_rows=pair_rows,
        model_rows=model_rows,
        total_files=total_files,
        read_files=read_files,
        skipped_files=skipped_files,
        model_count=len(models),
        group_count=group_count,
        single_count=single_count,
        min_similarity_percent=config.min_similarity_percent,
        min_shared_components=config.min_shared_components,
    )


def suggest_export_name():
    return f"PCB_Fix_Feeder_Groups_{datetime.now().strftime('%y%m%d')}.xlsx"


def export_model_feeder_group_result(result, output_path):
    if result is None:
        raise ServiceError("Belum ada hasil analisa untuk diexport.", title="Data kosong")

    output = _normalize_output_path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    group_sheet = workbook.active
    group_sheet.title = "Recommended Groups"
    _write_records_sheet(group_sheet, result.group_rows, GROUP_COLUMNS)

    pair_sheet = workbook.create_sheet("Pair Similarity")
    _write_records_sheet(pair_sheet, result.pair_rows, PAIR_COLUMNS)

    model_sheet = workbook.create_sheet("PCB Components")
    _write_records_sheet(model_sheet, result.model_rows, MODEL_COLUMNS)

    log_sheet = workbook.create_sheet("Scan Log")
    _write_scan_log(log_sheet, result)

    workbook.save(output)
    return str(output)


def _scan_models(source_folder, target_pcb_list, progress_callback=None):
    folder = Path(source_folder)
    models = OrderedDict()
    skipped_files = []
    read_files = 0
    pcb_folders = _pcb_folders(folder)
    
    if target_pcb_list:
        def _folder_matches_target(folder_name, targets):
            name_upper = folder_name.upper()
            for target in targets:
                # Match folders that start with INI_<target> (the primary PCB folder)
                if name_upper.startswith(f"INI_{target}"):
                    return True
            return False
        pcb_folders = [f for f in pcb_folders if _folder_matches_target(f.name, target_pcb_list)]

    folder_files_map = {pcb_folder: _excel_files_in_folder(pcb_folder) for pcb_folder in pcb_folders}
    total_files = sum(len(files) for files in folder_files_map.values())

    file_index = 0
    for pcb_folder in pcb_folders:
        excel_files = folder_files_map[pcb_folder]
        if not excel_files:
            skipped_files.append(f"{pcb_folder.name}: tidak ada file Excel program")
            continue

        merged_components = OrderedDict()
        component_frequencies = {}
        source_files = []
        total_inserts = {}
        variant_components = {}
        for file_path in excel_files:
            file_index += 1
            percent = max(1, min(95, int((file_index - 1) / max(1, total_files) * 95)))
            _emit_progress(
                progress_callback,
                percent,
                f"Reading {pcb_folder.name} ({file_index}/{total_files}): {file_path.name}",
            )

            try:
                part_values = _read_bom_parts(file_path)
            except Exception as exc:
                skipped_files.append(f"{pcb_folder.name} / {file_path.name}: {_error_message(exc)}")
                continue

            read_files += 1
            components = _unique_components(part_values)
            if not components:
                skipped_files.append(f"{pcb_folder.name} / {file_path.name}: Sheet BOM tidak punya component P/N yang valid.")
                continue

            source_files.append(file_path.name)

            import re
            m_prog = re.search(r'(EB[TU]\d+)', file_path.name)
            vname = m_prog.group(1) if m_prog else file_path.stem
            variant_components[vname] = set(components.keys())

            file_inserts = {}
            for val in part_values:
                k = _part_key(val)
                if k:
                    file_inserts[k] = file_inserts.get(k, 0) + 1
            for k, count in file_inserts.items():
                total_inserts[k] = total_inserts.get(k, 0) + count

            for key, part in components.items():
                if key not in merged_components:
                    merged_components[key] = part
                component_frequencies[key] = component_frequencies.get(key, 0) + 1

        if not merged_components:
            skipped_files.append(f"{pcb_folder.name}: tidak ada component P/N valid dari semua Excel program")
            continue

        insert_averages = {}
        num_files = len(source_files)
        if num_files > 0:
            for k, count in total_inserts.items():
                insert_averages[k] = count / num_files

        pcb_part_number = _pcb_part_number_for_folder(pcb_folder, source_files)
        folder_key = str(pcb_folder.resolve()).upper()
        models[folder_key] = ModelUsage(
            key=folder_key,
            display_name=pcb_part_number,
            pcb_part_number=pcb_part_number,
            source_folder=pcb_folder.name,
            source_files=source_files,
            components=merged_components,
            component_frequencies=component_frequencies,
            insert_averages=insert_averages,
            variant_components=variant_components,
        )

    models = OrderedDict(sorted(models.items(), key=lambda item: item[1].display_name.upper()))
    return models, total_files, read_files, skipped_files


def _build_pair_rows(models, config):
    pair_rows = []
    pair_lookup = {}

    for first_key, second_key in combinations(models.keys(), 2):
        first = models[first_key]
        second = models[second_key]
        first_components = set(first.components.keys())
        second_components = set(second.components.keys())
        shared_keys = first_components & second_components
        union_keys = first_components | second_components
        shared_count = len(shared_keys)
        min_count = min(len(first_components), len(second_components)) or 1
        union_count = len(union_keys) or 1
        similarity_percent = round(shared_count / min_count * 100, 1)
        jaccard_percent = round(shared_count / union_count * 100, 1)
        is_match = (
            similarity_percent >= config.min_similarity_percent
            and shared_count >= config.min_shared_components
        )
        row = {
            "status": STATUS_GROUPED if is_match else STATUS_SINGLE,
            "model_a_key": first_key,
            "model_b_key": second_key,
            "model_a": first.display_name,
            "model_b": second.display_name,
            "similarity_percent": similarity_percent,
            "jaccard_percent": jaccard_percent,
            "shared_component_count": shared_count,
            "model_a_component_count": len(first_components),
            "model_b_component_count": len(second_components),
            "shared_components": _format_components(shared_keys, first.components, limit=60),
            "_is_match": is_match,
        }
        pair_rows.append(row)
        pair_lookup[frozenset((first_key, second_key))] = row

    pair_rows.sort(key=_pair_sort_key)
    return pair_rows, pair_lookup


def _build_groups(models, pair_rows, pair_lookup):
    unassigned = set(models.keys())
    groups = []
    eligible_pairs = [row for row in pair_rows if row.get("_is_match")]
    eligible_pairs.sort(key=lambda row: (-row["similarity_percent"], -row["shared_component_count"], row["model_a"], row["model_b"]))

    while True:
        seed = None
        for row in eligible_pairs:
            if row["model_a_key"] in unassigned and row["model_b_key"] in unassigned:
                seed = row
                break
        if seed is None:
            break

        group = [seed["model_a_key"], seed["model_b_key"]]
        while True:
            candidates = []
            for candidate_key in sorted(unassigned - set(group), key=lambda key: models[key].display_name.upper()):
                pair_values = [_pair_for(pair_lookup, candidate_key, member_key) for member_key in group]
                if not all(pair and pair.get("_is_match") for pair in pair_values):
                    continue
                avg_similarity = sum(pair["similarity_percent"] for pair in pair_values) / len(pair_values)
                min_shared = min(pair["shared_component_count"] for pair in pair_values)
                candidates.append((avg_similarity, min_shared, models[candidate_key].display_name.upper(), candidate_key))

            if not candidates:
                break
            candidates.sort(key=lambda item: (-item[0], -item[1], item[2]))
            group.append(candidates[0][3])

        for key in group:
            unassigned.remove(key)
        groups.append(group)

    for key in sorted(unassigned, key=lambda item: models[item].display_name.upper()):
        groups.append([key])

    groups.sort(key=lambda group: (-len(group), models[group[0]].display_name.upper()))
    return groups


def _build_group_outputs(groups, models, pair_lookup):
    group_rows = []

    for index, group_keys in enumerate(groups, start=1):
        group_id = f"FFG{index:02d}"
        group_models = [models[key] for key in group_keys]
        pair_values = [
            _pair_for(pair_lookup, first_key, second_key)
            for first_key, second_key in combinations(group_keys, 2)
        ]
        pair_values = [pair for pair in pair_values if pair]
        avg_similarity = round(sum(pair["similarity_percent"] for pair in pair_values) / len(pair_values), 1) if pair_values else 100.0
        min_similarity = min((pair["similarity_percent"] for pair in pair_values), default=100.0)
        member_names = "; ".join(model.display_name for model in group_models)
        group_status = STATUS_GROUPED if len(group_keys) > 1 else STATUS_SINGLE

        group_rows.append(
            {
                "group_id": group_id,
                "status": group_status,
                "member_count": len(group_keys),
                "avg_similarity_percent": avg_similarity,
                "min_similarity_percent": min_similarity,
                "members": member_names,
            }
        )

    return group_rows


def _build_model_rows(models):
    rows = []
    for model in models.values():
        rows.append(
            {
                "pcb_part_number": model.pcb_part_number,
                "component_count": model.component_count,
                "excel_file_count": len(model.source_files),
                "source_folder": model.source_folder,
                "source_files": "; ".join(model.source_files),
                "components": _format_components(model.components.keys(), model.components, limit=200),
            }
        )
    return rows


def _write_records_sheet(worksheet, rows, columns):
    worksheet.append([header for _, header in columns])
    for row in rows:
        worksheet.append([row.get(key, "") for key, _ in columns])
    _style_sheet(worksheet)


def _write_scan_log(worksheet, result):
    rows = [
        ("Excel files found", result.total_files),
        ("Files read", result.read_files),
        ("PCB folders analyzed", result.model_count),
        ("Recommended groups", result.group_count),
        ("Single PCB", result.single_count),
        ("Minimum similarity %", result.min_similarity_percent),
        ("Minimum shared components", result.min_shared_components),
        ("Skipped/error files", len(result.skipped_files)),
    ]

    worksheet.append(["Item", "Value"])
    for item, value in rows:
        worksheet.append([item, value])

    if result.skipped_files:
        worksheet.append([])
        worksheet.append(["Skipped/error detail", ""])
        for skipped in result.skipped_files:
            worksheet.append([skipped, ""])

    _style_sheet(worksheet)


def _pair_for(pair_lookup, first_key, second_key):
    return pair_lookup.get(frozenset((first_key, second_key)))


def _unique_components(parts):
    unique = OrderedDict()
    for part in parts:
        key = _part_key(part)
        if key and key not in unique:
            unique[key] = str(part).strip()
    return unique


def _pcb_folders(source_folder):
    try:
        entries = sorted(os.scandir(source_folder), key=lambda e: e.name.upper())
    except OSError:
        return [source_folder]
    child_folders = [Path(e.path) for e in entries if e.is_dir()]
    return child_folders or [source_folder]


def _excel_files_in_folder(folder):
    files = []
    try:
        entries = sorted(os.scandir(folder), key=lambda e: e.name.upper())
    except OSError:
        return files
    for entry in entries:
        if not entry.is_file():
            continue
        if entry.name.startswith("~$"):
            continue
        ext = os.path.splitext(entry.name)[1].lower()
        if ext in EXCEL_EXTENSIONS:
            files.append(Path(entry.path))
    return files


def _pcb_part_number_for_folder(pcb_folder, source_files):
    pcb_part_number, revision = parse_pcb_part_number(pcb_folder)
    display = format_pcb_part_number(pcb_part_number, revision)
    if display != "-":
        return display

    for file_name in source_files:
        pcb_part_number, revision = parse_pcb_part_number(pcb_folder / file_name)
        display = format_pcb_part_number(pcb_part_number, revision)
        if display != "-":
            return display

    return pcb_folder.name


def _format_components(component_keys, component_catalog, limit=30):
    keys = sorted(component_keys, key=lambda key: component_catalog.get(key, key).upper())
    if not keys:
        return "-"
    values = [component_catalog.get(key, key) for key in keys[:limit]]
    if len(keys) > limit:
        values.append(f"... +{len(keys) - limit} more")
    return "; ".join(values)


def _pair_sort_key(row):
    return (
        0 if row.get("_is_match") else 1,
        -row["similarity_percent"],
        -row["shared_component_count"],
        row["model_a"].upper(),
        row["model_b"].upper(),
    )


def _validate_config(config):
    if not config.source_folder:
        raise ServiceError("Folder Induk PCB belum dipilih.", title="Input belum lengkap")
    if not Path(config.source_folder).is_dir():
        raise ServiceError(f"Folder Induk PCB tidak ditemukan:\n{config.source_folder}", title="Folder tidak ditemukan")
    if config.min_similarity_percent < 1 or config.min_similarity_percent > 100:
        raise ServiceError("Minimum similarity harus di antara 1 sampai 100.", title="Input tidak valid")
    if config.min_shared_components < 1:
        raise ServiceError("Minimum shared components minimal 1.", title="Input tidak valid")


def _emit_progress(progress_callback, percent, message):
    if progress_callback:
        progress_callback(percent, message)
