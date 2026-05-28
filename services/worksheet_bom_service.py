from dataclasses import dataclass
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook

from services import bom_service
from services.errors import ServiceError
from utils.sort import natural_sort_key


WORKSHEET_EXCLUDED_FEED_IDS = {"1401", "1402"}
BOM_WORKSHEET_SCOPE_SIDE = "TOP"


@dataclass
class WorksheetBomSource:
    dataframe: pd.DataFrame
    filename: str
    row_count: int
    total_qty: int
    skipped_rows: int = 0
    sheet_name: str = ""
    chassis_pn: str = ""
    pcb_pn: str = ""


@dataclass
class WorksheetBomCompareResult:
    worksheet: WorksheetBomSource
    bom: WorksheetBomSource
    diffs: list


def load_worksheet_summary(file_path):
    path = _clean_path(file_path)
    if not Path(path).is_file():
        raise ServiceError(f"Worksheet tidak ditemukan:\n{path}", title="File tidak ditemukan")

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ServiceError(f"Worksheet Excel tidak bisa dibaca:\n{exc}", title="Excel error") from exc

    try:
        worksheet, header_row, columns = _find_worksheet_table(workbook)
        records, skipped_rows = _read_worksheet_records(worksheet, header_row, columns)
    finally:
        workbook.close()

    if not records:
        raise ServiceError("Tidak ada data feeder valid di Worksheet.", title="Format Worksheet tidak valid")

    summary_records = _summarize_worksheet_records(records)
    dataframe = pd.DataFrame(summary_records)
    return WorksheetBomSource(
        dataframe=dataframe,
        filename=Path(path).name,
        row_count=len(dataframe),
        total_qty=int(dataframe["WorksheetQty"].sum()) if not dataframe.empty else 0,
        skipped_rows=skipped_rows,
        sheet_name=worksheet.title,
    )


def load_bom_top_summary(file_path):
    path = _clean_path(file_path)
    if not Path(path).is_file():
        raise ServiceError(f"BOM File tidak ditemukan:\n{path}", title="File tidak ditemukan")

    result = bom_service.load_raw_bom(path, check_duplicate_circuits=False)
    dataframe = result.dataframe.copy()
    dataframe = dataframe[dataframe["Side"].astype(str).str.upper() == BOM_WORKSHEET_SCOPE_SIDE].copy()

    if dataframe.empty:
        raise ServiceError(
            "Tidak ada data TOP SMT pada BOM.\n"
            "Worksheet feeder dibandingkan dengan item BOM dari Upper SMT/S/W Package.",
            title="BOM kosong untuk Worksheet",
        )

    summary_records = _summarize_bom_records(dataframe)
    summary_df = pd.DataFrame(summary_records)
    return WorksheetBomSource(
        dataframe=summary_df,
        filename=Path(path).name,
        row_count=len(summary_df),
        total_qty=int(summary_df["BomQty"].sum()) if not summary_df.empty else 0,
        chassis_pn=result.chassis_pn,
        pcb_pn=result.pcb_pn,
    )


def run_worksheet_bom_compare(worksheet_path, bom_path):
    worksheet_summary = load_worksheet_summary(worksheet_path)
    bom_summary = load_bom_top_summary(bom_path)
    diffs = compare_worksheet_bom(worksheet_summary.dataframe, bom_summary.dataframe)
    return WorksheetBomCompareResult(worksheet_summary, bom_summary, diffs)


def compare_worksheet_bom(worksheet_df, bom_df):
    worksheet_lookup = {
        str(row["PartNo"]).strip(): row
        for _, row in worksheet_df.iterrows()
        if str(row.get("PartNo", "")).strip()
    }
    bom_lookup = {
        str(row["PartNo"]).strip(): row
        for _, row in bom_df.iterrows()
        if str(row.get("PartNo", "")).strip()
    }

    diffs = []
    all_parts = sorted(set(worksheet_lookup) | set(bom_lookup), key=natural_sort_key)
    for part_no in all_parts:
        worksheet_row = worksheet_lookup.get(part_no)
        bom_row = bom_lookup.get(part_no)
        worksheet_qty = int(worksheet_row["WorksheetQty"]) if worksheet_row is not None else 0
        bom_qty = int(bom_row["BomQty"]) if bom_row is not None else 0

        if worksheet_qty == bom_qty:
            continue

        if worksheet_row is None:
            diff_type = "ADD"
            description = "Part ada di BOM TOP, tapi tidak ada di Worksheet."
        elif bom_row is None:
            diff_type = "DEL"
            description = "Part ada di Worksheet, tapi tidak ada di BOM TOP."
        else:
            diff_type = "CNG"
            description = "Total CNT Worksheet berbeda dengan jumlah designator BOM TOP."

        diffs.append(
            {
                "PartNo": part_no,
                "WorksheetQty": worksheet_qty,
                "BomQty": bom_qty,
                "Delta": worksheet_qty - bom_qty,
                "WorksheetRows": _row_value(worksheet_row, "WorksheetRows"),
                "FeedIds": _row_value(worksheet_row, "FeedIds"),
                "RefDes": _row_value(bom_row, "RefDes"),
                "Spec": _row_value(worksheet_row, "Spec") or _row_value(bom_row, "Spec"),
                "Type": diff_type,
                "Description": description,
            }
        )

    type_order = {"ADD": 1, "CNG": 2, "DEL": 3}
    diffs.sort(key=lambda item: (type_order.get(item["Type"], 99), natural_sort_key(item["PartNo"])))
    return diffs


def export_worksheet_bom_results(diff_results, file_path):
    output_path = Path(file_path)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")
    pd.DataFrame(
        diff_results,
        columns=[
            "PartNo",
            "WorksheetQty",
            "BomQty",
            "Delta",
            "WorksheetRows",
            "FeedIds",
            "RefDes",
            "Spec",
            "Type",
            "Description",
        ],
    ).to_excel(output_path, index=False)
    return str(output_path)


def _find_worksheet_table(workbook):
    for worksheet in workbook.worksheets:
        max_row = min(30, worksheet.max_row or 0)
        for row_index in range(1, max_row + 1):
            row_values = [worksheet.cell(row_index, column).value for column in range(1, worksheet.max_column + 1)]
            normalized = [_normalize_header(value) for value in row_values]
            if "PARTNO" not in normalized or "CNT" not in normalized:
                continue

            columns = {
                "part": normalized.index("PARTNO") + 1,
                "qty": normalized.index("CNT") + 1,
                "table": _find_header_column(normalized, "TBL"),
                "feed_no": _find_header_column(normalized, "FEED"),
                "pos": _find_header_column(normalized, "POS"),
                "side": _find_header_column(normalized, "SIDE"),
                "spec": _find_header_column(normalized, "SPEC"),
                "nz": _find_header_column(normalized, "NZ"),
                "fd": _find_header_column(normalized, "FD"),
            }

            subheader = [
                _normalize_header(worksheet.cell(row_index + 1, column).value)
                for column in range(1, worksheet.max_column + 1)
            ]
            columns["feed_ar"] = _find_header_column(subheader, "AR")
            columns["feed_id"] = _find_header_column(subheader, "ID") or _find_header_column(normalized, "ID")
            return worksheet, row_index, columns

    raise ServiceError(
        "Header Worksheet tidak ditemukan.\n"
        "File harus punya kolom PART NO dan CNT.",
        title="Format Worksheet tidak valid",
    )


def _read_worksheet_records(worksheet, header_row, columns):
    records = []
    skipped_rows = 0
    start_row = header_row + 1
    if _looks_like_subheader_row(worksheet, start_row, columns):
        start_row += 1

    for row_index in range(start_row, worksheet.max_row + 1):
        part_no = _cell_text(worksheet.cell(row_index, columns["part"]).value)
        if not part_no or part_no.upper() in {"PART NO", "NAN", "NONE"}:
            continue

        feed_id = _cell_text(_cell_value(worksheet, row_index, columns.get("feed_id")))
        if feed_id in WORKSHEET_EXCLUDED_FEED_IDS:
            skipped_rows += 1
            continue

        qty = _parse_int(worksheet.cell(row_index, columns["qty"]).value)
        records.append(
            {
                "PartNo": part_no,
                "Qty": qty,
                "WorksheetRow": row_index,
                "Table": _cell_text(_cell_value(worksheet, row_index, columns.get("table"))),
                "FeedNo": _cell_text(_cell_value(worksheet, row_index, columns.get("feed_no"))),
                "FeedAr": _cell_text(_cell_value(worksheet, row_index, columns.get("feed_ar"))),
                "FeedId": feed_id,
                "Pos": _cell_text(_cell_value(worksheet, row_index, columns.get("pos"))),
                "Side": _cell_text(_cell_value(worksheet, row_index, columns.get("side"))),
                "Spec": _cell_text(_cell_value(worksheet, row_index, columns.get("spec"))),
            }
        )

    return records, skipped_rows


def _summarize_worksheet_records(records):
    grouped = {}
    for record in records:
        part_no = record["PartNo"]
        group = grouped.setdefault(
            part_no,
            {
                "PartNo": part_no,
                "WorksheetQty": 0,
                "WorksheetRows": [],
                "FeedIds": [],
                "FeedSlots": [],
                "Side": "",
                "Spec": "",
            },
        )
        group["WorksheetQty"] += record["Qty"]
        group["WorksheetRows"].append(str(record["WorksheetRow"]))
        if record["FeedId"]:
            group["FeedIds"].append(record["FeedId"])
        slot = _format_feed_slot(record)
        if slot:
            group["FeedSlots"].append(slot)
        if not group["Side"] and record["Side"]:
            group["Side"] = record["Side"]
        if not group["Spec"] and record["Spec"]:
            group["Spec"] = record["Spec"]

    summary = []
    for group in grouped.values():
        summary.append(
            {
                "PartNo": group["PartNo"],
                "WorksheetQty": group["WorksheetQty"],
                "WorksheetRows": _compact_list(group["WorksheetRows"], limit=18),
                "FeedIds": _compact_list(group["FeedIds"], limit=18),
                "FeedSlots": _compact_list(group["FeedSlots"], limit=18),
                "Side": group["Side"],
                "Spec": group["Spec"],
            }
        )

    summary.sort(key=lambda item: natural_sort_key(item["PartNo"]))
    return summary


def _summarize_bom_records(dataframe):
    grouped = {}
    for _, row in dataframe.iterrows():
        part_no = str(row.get("PartNo", "") or "").strip()
        if not part_no:
            continue
        group = grouped.setdefault(
            part_no,
            {
                "PartNo": part_no,
                "BomQty": 0,
                "RefDes": [],
                "Side": [],
                "Spec": "",
            },
        )
        group["BomQty"] += 1
        circuit = str(row.get("Circuit", "") or "").strip()
        if circuit:
            group["RefDes"].append(circuit)
        side = str(row.get("Side", "") or "").strip()
        if side:
            group["Side"].append(side)
        spec = str(row.get("Spec", "") or "").strip()
        if spec and not group["Spec"]:
            group["Spec"] = spec

    summary = []
    for group in grouped.values():
        summary.append(
            {
                "PartNo": group["PartNo"],
                "BomQty": group["BomQty"],
                "RefDes": _compact_list(sorted(group["RefDes"], key=natural_sort_key), limit=24),
                "Side": _compact_list(sorted(set(group["Side"]), key=natural_sort_key), limit=8),
                "Spec": group["Spec"],
            }
        )

    summary.sort(key=lambda item: natural_sort_key(item["PartNo"]))
    return summary


def _format_feed_slot(record):
    values = [record.get("Table"), record.get("FeedNo"), record.get("FeedAr")]
    if not any(values):
        return ""
    return "/".join(value for value in values if value)


def _looks_like_subheader_row(worksheet, row_index, columns):
    part = _cell_text(worksheet.cell(row_index, columns["part"]).value)
    qty = _cell_text(worksheet.cell(row_index, columns["qty"]).value)
    feed_id = _cell_text(_cell_value(worksheet, row_index, columns.get("feed_id")))
    return not part and not qty and feed_id.upper() == "ID"


def _find_header_column(normalized_row, target):
    try:
        return normalized_row.index(target) + 1
    except ValueError:
        return None


def _cell_value(worksheet, row_index, column_index):
    if not column_index:
        return None
    return worksheet.cell(row_index, column_index).value


def _cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_int(value):
    if value is None or str(value).strip() == "":
        return 0
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return 0


def _normalize_header(value):
    return re.sub(r"[^A-Z0-9]+", "", _cell_text(value).upper())


def _compact_list(values, limit=20):
    clean_values = [str(value).strip() for value in values if str(value).strip()]
    if not clean_values:
        return ""
    shown = clean_values[:limit]
    suffix = f" ... (+{len(clean_values) - limit})" if len(clean_values) > limit else ""
    return ", ".join(shown) + suffix


def _row_value(row, key):
    if row is None:
        return ""
    value = row.get(key, "")
    return "" if pd.isna(value) else value


def _clean_path(file_path):
    return str(file_path or "").strip().replace('"', "").replace("'", "")
