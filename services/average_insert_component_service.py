from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from services.used_part_component_service import (
    MODE_PCB_LIST,
    MODE_PROGRAM_FOLDER,
    BomReader,
    _error_message,
    _excel_files_in_folder,
    _extract_program_name,
    _map_pcb_folders,
    _normalize_output_path,
    _parse_pcb_part_numbers,
    _part_key,
    _safe_sheet_name,
    _unique_name,
    _validate_config,
)
from services.errors import ServiceError


@dataclass
class AverageInsertComponentConfig:
    mode: str
    source_folder: str
    output_path: str
    pcb_part_numbers: str = ""


@dataclass
class AverageInsertComponentResult:
    output_path: str
    mode: str
    group_count: int
    file_count: int
    part_count: int
    skipped_files: list[str] = field(default_factory=list)


@dataclass
class PartInsertCount:
    part_name: str
    insert_count: int


@dataclass
class PartInsertAverage:
    part_name: str
    avg_insert: int


@dataclass
class AverageInsertCollection:
    groups: OrderedDict
    file_count: int = 0
    skipped_files: list[str] = field(default_factory=list)

    @property
    def part_count(self):
        master = OrderedDict()
        for parts in self.groups.values():
            for key, row in parts.items():
                if key and key not in master:
                    master[key] = row.part_name
        return len(master)


def suggest_output_name(mode=MODE_PROGRAM_FOLDER):
    mode_label = "MODE_1" if mode == MODE_PCB_LIST else "MODE_2"
    return f"Used_Part_Component_AVG_Insert_{mode_label}_{datetime.now().strftime('%y%m%d')}.xlsx"


def generate_average_insert_component(config: AverageInsertComponentConfig):
    _validate_config(config)
    output_path = _normalize_output_path(config.output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with BomReader() as reader:
        if config.mode == MODE_PCB_LIST:
            collection = _collect_by_pcb_list(config, reader)
            _write_mode_1_workbook(collection, output_path)
        else:
            collection = _collect_by_program_folder(config, reader)
            _write_mode_2_workbook(collection, output_path)

    return AverageInsertComponentResult(
        output_path=str(output_path),
        mode=config.mode,
        group_count=len(collection.groups),
        file_count=collection.file_count,
        part_count=collection.part_count,
        skipped_files=collection.skipped_files,
    )


def _collect_by_program_folder(config, reader):
    folder = Path(config.source_folder)
    files = _excel_files_in_folder(folder)
    if not files:
        raise ServiceError("Tidak ada file Excel program di folder yang dipilih.", title="Data kosong")

    groups = OrderedDict()
    skipped_files = []
    file_count = 0

    for file_path in files:
        group_name = _unique_name(_extract_program_name(file_path.name), groups)
        try:
            counts = _read_part_counts(
                reader,
                file_path,
                start_row=2,
                filter_na_column=True,
                exclude_board_parts=False,
            )
        except Exception as exc:
            skipped_files.append(f"{file_path.name}: {_error_message(exc)}")
            continue

        if not counts:
            skipped_files.append(f"{file_path.name}: tidak ada part name di sheet BOM kolom C")
            continue

        groups[group_name] = _average_count_sets([counts])
        file_count += 1

    if not groups:
        raise ServiceError("Tidak ada data part yang berhasil dibaca dari folder ini.", title="Data kosong")

    return AverageInsertCollection(groups=groups, file_count=file_count, skipped_files=skipped_files)


def _collect_by_pcb_list(config, reader):
    folder = Path(config.source_folder)
    pcb_numbers = _parse_pcb_part_numbers(config.pcb_part_numbers)
    folder_map = _map_pcb_folders(folder, pcb_numbers)
    if not folder_map:
        raise ServiceError("Tidak ada folder PCB yang cocok dengan list part number.", title="Folder tidak ditemukan")

    groups = OrderedDict()
    skipped_files = []
    file_count = 0

    for pcb_number in pcb_numbers:
        matched_folders = folder_map.get(pcb_number, [])
        if not matched_folders:
            skipped_files.append(f"{pcb_number}: folder tidak ditemukan")
            continue

        count_sets = []
        for pcb_folder in matched_folders:
            files = _excel_files_in_folder(pcb_folder)
            if not files:
                skipped_files.append(f"{pcb_number}: tidak ada file Excel di {pcb_folder.name}")
                continue

            for file_path in files:
                try:
                    counts = _read_part_counts(
                        reader,
                        file_path,
                        start_row=1,
                        filter_na_column=True,
                        exclude_board_parts=True,
                    )
                except Exception as exc:
                    skipped_files.append(f"{pcb_number} / {file_path.name}: {_error_message(exc)}")
                    continue

                if counts:
                    count_sets.append(counts)
                    file_count += 1
                else:
                    skipped_files.append(f"{pcb_number} / {file_path.name}: tidak ada part valid")

        averages = _average_count_sets(count_sets)
        if averages:
            groups[pcb_number] = averages

    if not groups:
        raise ServiceError("Tidak ada data part valid dari PCB part number yang dipilih.", title="Data kosong")

    return AverageInsertCollection(groups=groups, file_count=file_count, skipped_files=skipped_files)


def _read_part_counts(reader, file_path, start_row, filter_na_column, exclude_board_parts):
    parts = reader.read_part_names(file_path, start_row, filter_na_column, exclude_board_parts)
    counts = OrderedDict()
    for part in parts:
        key = _part_key(part)
        if not key:
            continue
        if key not in counts:
            counts[key] = PartInsertCount(part_name=part, insert_count=0)
        counts[key].insert_count += 1
    return counts


def _average_count_sets(count_sets):
    merged = OrderedDict()
    for counts in count_sets:
        for key, row in counts.items():
            if key not in merged:
                merged[key] = {"part_name": row.part_name, "insert_counts": []}
            merged[key]["insert_counts"].append(row.insert_count)

    averages = OrderedDict()
    for key, row in merged.items():
        averages[key] = PartInsertAverage(
            part_name=row["part_name"],
            avg_insert=_rounded_average(row["insert_counts"]),
        )

    return _sort_average_rows(averages)


def _write_mode_1_workbook(collection, output_path):
    _write_matrix_workbook(collection, output_path, "MASTER")


def _write_mode_2_workbook(collection, output_path):
    _write_matrix_workbook(collection, output_path, "P/N COMPONENT")


def _write_matrix_workbook(collection, output_path, master_header):
    workbook = Workbook()
    master = workbook.active
    master.title = "MASTER"
    master_rows = _write_master_matrix(master, collection, master_header)

    _style_sheet(master, len(master_rows) + 1, len(collection.groups) + 2, auto_filter=True)
    _add_detail_sheets(workbook, collection.groups)
    workbook.save(output_path)


def _write_master_matrix(worksheet, collection, master_header):
    master_rows = _master_average_rows(collection.groups)
    group_names = list(collection.groups.keys())
    avg_column = len(group_names) + 2

    worksheet.cell(1, 1).value = master_header
    for col_index, group_name in enumerate(group_names, start=2):
        worksheet.cell(1, col_index).value = group_name
    worksheet.cell(1, avg_column).value = "AVG INSERT"

    for row_index, (part_key, row) in enumerate(master_rows.items(), start=2):
        worksheet.cell(row_index, 1).value = row.part_name
        for col_index, group_name in enumerate(group_names, start=2):
            group_row = collection.groups[group_name].get(part_key)
            if group_row is not None:
                worksheet.cell(row_index, col_index).value = group_row.part_name
        worksheet.cell(row_index, avg_column).value = row.avg_insert

    return master_rows


def _add_detail_sheets(workbook, groups):
    used_names = {worksheet.title.upper() for worksheet in workbook.worksheets}
    for group_name, parts in groups.items():
        sheet_name = _safe_sheet_name(group_name, used_names)
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.cell(1, 1).value = group_name
        worksheet.cell(1, 2).value = "AVG INSERT"
        for row_index, row in enumerate(parts.values(), start=2):
            worksheet.cell(row_index, 1).value = row.part_name
            worksheet.cell(row_index, 2).value = row.avg_insert
        _style_sheet(worksheet, len(parts) + 1, 2, auto_filter=True)


def _master_average_rows(groups):
    merged = OrderedDict()
    for parts in groups.values():
        for key, row in parts.items():
            if key not in merged:
                merged[key] = {"part_name": row.part_name, "avg_inserts": []}
            merged[key]["avg_inserts"].append(row.avg_insert)

    rows = OrderedDict()
    for key, row in merged.items():
        rows[key] = PartInsertAverage(
            part_name=row["part_name"],
            avg_insert=_rounded_average(row["avg_inserts"]),
        )
    return _sort_average_rows(rows)


def _sort_average_rows(rows):
    return OrderedDict(
        sorted(
            rows.items(),
            key=lambda item: (-item[1].avg_insert, item[1].part_name.upper()),
        )
    )


def _rounded_average(values):
    if not values:
        return 0
    average = Decimal(sum(values)) / Decimal(len(values))
    return int(average.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _style_sheet(worksheet, row_count, column_count, auto_filter):
    header_font = Font(name="Calibri", size=11, bold=True)
    body_font = Font(name="Calibri", size=11)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row in worksheet.iter_rows(min_row=2, max_row=max(row_count, 2), max_col=max(column_count, 1)):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if isinstance(cell.value, int):
                cell.number_format = "0"

    for col_index in range(1, column_count + 1):
        letter = get_column_letter(col_index)
        max_length = 10
        for row_index in range(1, min(row_count, worksheet.max_row) + 1):
            value = worksheet.cell(row_index, col_index).value
            if value not in (None, ""):
                max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[letter].width = min(max_length + 2, 32)

    if auto_filter and row_count >= 1 and column_count >= 1:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(column_count)}{row_count}"
