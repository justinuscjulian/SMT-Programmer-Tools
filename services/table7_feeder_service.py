import pandas as pd
from dataclasses import dataclass
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

from services.errors import ServiceError
from services.model_feeder_group_service import _scan_models, _emit_progress
from services.common_feeder_reuse_service import _normalize_output_path, _style_sheet


@dataclass
class Table7FeederConfig:
    source_folder: str
    target_pcb_list: list = None
    table7_ref_file: str = ""
    master_mapping_file: str = ""


@dataclass
class Table7FeederResult:
    pcb_rows: list[dict]
    total_files: int
    read_files: int
    skipped_files: list[str]
    model_count: int


def get_table7_components(ref_file: str) -> dict:
    try:
        df = pd.read_excel(ref_file)
        if df.empty:
            raise ServiceError("File komponen Table 7 kosong.", title="Data Kosong")
            
        components = {}
        for _, row in df.iterrows():
            part = str(row.iloc[0]).strip().upper()
            if part and part != "NAN":
                size = 1
                if len(row) > 1:
                    size_val = str(row.iloc[1]).strip().upper()
                    if "2" in size_val:
                        size = 2
                components[part] = size
                
        if not components:
            raise ServiceError("Tidak ada komponen valid di file.", title="Data Kosong")
        return components
    except Exception as exc:
        raise ServiceError(f"Gagal membaca file komponen Table 7: {exc}", title="Error Baca File")


def get_master_mapping(ref_file: str) -> dict:
    master_slots = {}
    if not ref_file or not Path(ref_file).is_file():
        return master_slots
    
    try:
        df = pd.read_excel(ref_file, sheet_name="Table 7 Slots Detail")
        slot_cols = [col for col in df.columns if str(col).startswith("[7]")]
        for _, row in df.iterrows():
            for i, col in enumerate(slot_cols):
                comp = str(row[col]).strip().upper()
                if comp and comp != "NAN" and comp != "BLOCKED":
                    if comp not in master_slots:
                        master_slots[comp] = i
    except Exception as exc:
        raise ServiceError(f"Gagal membaca file Master Mapping: {exc}", title="Error Baca File")
    return master_slots


def analyze_table7_feeders(config: Table7FeederConfig, progress_callback=None) -> Table7FeederResult:
    if not config.source_folder or not Path(config.source_folder).is_dir():
        raise ServiceError("Folder Induk PCB tidak valid.", title="Input Error")
    if not config.table7_ref_file or not Path(config.table7_ref_file).is_file():
        raise ServiceError("File referensi Table 7 tidak valid.", title="Input Error")

    _emit_progress(progress_callback, 0, "Membaca referensi & Master Mapping...")
    table7_components = get_table7_components(config.table7_ref_file)
    master_slots = get_master_mapping(config.master_mapping_file)

    _emit_progress(progress_callback, 5, "Scanning PCB folders...")
    models, total_files, read_files, skipped_files = _scan_models(
        config.source_folder, config.target_pcb_list, progress_callback
    )

    if not models:
        return Table7FeederResult([], total_files, read_files, skipped_files, 0)

    _emit_progress(progress_callback, 96, "Menganalisa komponen Table 7 per PCB...")
    pcb_rows = []
    
    for key, model in models.items():
        used_table7_parts = []
        for comp_key, comp_val in model.components.items():
            if comp_key.upper() in table7_components or comp_val.upper() in table7_components:
                # Always use the value (part number) to check against dict
                used_table7_parts.append(comp_val.upper())
                
        used_table7_parts = list(dict.fromkeys(used_table7_parts))
        used_table7_parts.sort()
        
        slots = [None] * 30
        unassigned_parts = []
        
        # 1. Place Master Mapping components
        for part in used_table7_parts:
            if part in master_slots:
                start_idx = master_slots[part]
                size = table7_components.get(part, 1)
                if start_idx < 30:
                    slots[start_idx] = part
                    if size == 2 and start_idx + 1 < 30:
                        slots[start_idx + 1] = "BLOCKED"
                else:
                    unassigned_parts.append(part)
            else:
                unassigned_parts.append(part)
                
        # 2. Place remaining components
        overload = False
        for part in unassigned_parts:
            size = table7_components.get(part, 1)
            placed = False
            for i in range(30):
                if size == 1:
                    if slots[i] is None:
                        slots[i] = part
                        placed = True
                        break
                elif size == 2:
                    if i + 1 < 30 and slots[i] is None and slots[i+1] is None:
                        slots[i] = part
                        slots[i+1] = "BLOCKED"
                        placed = True
                        break
            if not placed:
                overload = True
        
        comp_count = sum(1 for s in slots if s and s != "BLOCKED")
        
        if comp_count == 0:
            status = "NO TABLE 7 PARTS"
        elif overload:
            status = "OVERLOAD (> 30)"
        else:
            status = "OK"

        slot_assignments = []
        for idx, s in enumerate(slots, start=1):
            if s and s != "BLOCKED":
                size = table7_components.get(s, 1)
                if size == 2:
                    slot_assignments.append(f"[7]{idx}-{idx+1}: {s}")
                else:
                    slot_assignments.append(f"[7]{idx}: {s}")
        
        pcb_rows.append({
            "pcb_part_number": model.pcb_part_number,
            "status": status,
            "table7_part_count": comp_count,
            "excel_file_count": len(model.source_files),
            "members": "; ".join(model.source_files),
            "slot_assignments": "\n".join(slot_assignments) if slot_assignments else "-",
            "_slots_array": slots,
        })
        
    _emit_progress(progress_callback, 100, "Selesai")
    
    return Table7FeederResult(
        pcb_rows=pcb_rows,
        total_files=total_files,
        read_files=read_files,
        skipped_files=skipped_files,
        model_count=len(models),
    )


def suggest_table7_export_name():
    return f"Table7_Fix_Feeder_{datetime.now().strftime('%y%m%d')}.xlsx"


def export_table7_result(result: Table7FeederResult, output_path: str):
    if result is None or not result.pcb_rows:
        raise ServiceError("Belum ada hasil analisa untuk diexport.", title="Data kosong")

    output = _normalize_output_path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    
    summary_sheet = workbook.active
    summary_sheet.title = "Table 7 Fix Feeder Summary"
    columns = [
        ("pcb_part_number", "PCB Part Number"),
        ("status", "Status"),
        ("table7_part_count", "Total Parts T7"),
        ("excel_file_count", "Total Program Excel"),
        ("members", "Program Excel"),
    ]
    summary_sheet.append([header for _, header in columns])
    for row in result.pcb_rows:
        summary_sheet.append([row.get(key, "") for key, _ in columns])
    _style_sheet(summary_sheet)
    
    slot_sheet = workbook.create_sheet("Table 7 Slots Detail")
    slot_sheet.append(["PCB Part Number", "Status", "Total Parts T7"] + [f"[7]{i}" for i in range(1, 31)])
    for row in result.pcb_rows:
        out_row = [row["pcb_part_number"], row["status"], row["table7_part_count"]]
        slots = row["_slots_array"]
        for i in range(30):
            val = slots[i] if i < len(slots) and slots[i] else ""
            out_row.append(val)
        slot_sheet.append(out_row)
    _style_sheet(slot_sheet)
    
    log_sheet = workbook.create_sheet("Scan Log")
    log_rows = [
        ("Excel files found", result.total_files),
        ("Files read", result.read_files),
        ("PCB folders analyzed", result.model_count),
        ("Skipped/error files", len(result.skipped_files)),
    ]
    log_sheet.append(["Item", "Value"])
    for item, value in log_rows:
        log_sheet.append([item, value])
        
    if result.skipped_files:
        log_sheet.append([])
        log_sheet.append(["Skipped/error detail", ""])
        for skipped in result.skipped_files:
            log_sheet.append([skipped, ""])
    _style_sheet(log_sheet)

    workbook.save(output)
    return str(output)
