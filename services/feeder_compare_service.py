import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from services.feeder_mapping_service import load_feeder_mapping, load_cm602_feeder_mapping
from utils.sort import natural_sort_key


STATUS_ADD = "ADD"
STATUS_MOVE = "MOVE"
STATUS_CNG = "CNG"
STATUS_DEL = "DEL"
STATUS_FILTER_ALL = "SHOW ALL"
STATUS_OPTIONS = [STATUS_FILTER_ALL, STATUS_ADD, STATUS_MOVE, STATUS_CNG, STATUS_DEL]

RESULT_COLUMNS = [
    ("status", "Status"),
    ("old_location", "Program A Location"),
    ("old_part_number", "Program A Part Number"),
    ("old_table", "A Table"),
    ("old_slot", "A Slot"),
    ("old_position", "A Position"),
    ("new_location", "Program B Location"),
    ("new_part_number", "Program B Part Number"),
    ("new_table", "B Table"),
    ("new_slot", "B Slot"),
    ("new_position", "B Position"),
    ("description", "Description"),
]


@dataclass
class FeederCompareResult:
    old_file: str
    new_file: str
    old_count: int
    new_count: int
    rows: list
    add_count: int
    move_count: int
    cng_count: int
    del_count: int


def compare_feeder_files(old_path, new_path, old_parser="NPM", new_parser="NPM"):
    def load_mapping(path, parser_type):
        if parser_type == "CM602":
            return load_cm602_feeder_mapping(path)
        return load_feeder_mapping(path)
        
    old_mapping = load_mapping(old_path, old_parser)
    new_mapping = load_mapping(new_path, new_parser)
    rows = compare_feeder_records(old_mapping.records, new_mapping.records)
    return FeederCompareResult(
        old_file=old_mapping.source_file,
        new_file=new_mapping.source_file,
        old_count=old_mapping.row_count,
        new_count=new_mapping.row_count,
        rows=rows,
        add_count=sum(1 for row in rows if row["status"] == STATUS_ADD),
        move_count=sum(1 for row in rows if row["status"] == STATUS_MOVE),
        cng_count=sum(1 for row in rows if row["status"] == STATUS_CNG),
        del_count=sum(1 for row in rows if row["status"] == STATUS_DEL),
    )


def compare_feeder_records(old_records, new_records):
    old_entries = [_entry(record, index) for index, record in enumerate(old_records or [])]
    new_entries = [_entry(record, index) for index, record in enumerate(new_records or [])]

    old_used = set()
    new_used = set()
    new_exact_lookup = _group_entries(new_entries, _exact_key)
    for old_entry in old_entries:
        candidates = new_exact_lookup.get(_exact_key(old_entry), [])
        match = _take_first_unused(candidates, new_used)
        if match is None:
            continue
        old_used.add(old_entry["_index"])
        new_used.add(match["_index"])

    rows = []
    new_part_lookup = _group_entries([entry for entry in new_entries if entry["_index"] not in new_used], _part_key)
    for old_entry in _sorted_entries(entry for entry in old_entries if entry["_index"] not in old_used):
        candidates = [
            entry
            for entry in new_part_lookup.get(_part_key(old_entry), [])
            if entry["_location_key"] != old_entry["_location_key"]
        ]
        match = _take_first_unused(candidates, new_used)
        if match is None:
            continue
        old_used.add(old_entry["_index"])
        new_used.add(match["_index"])
        rows.append(_diff_row(STATUS_MOVE, old_entry, match))

    new_location_lookup = _group_entries([entry for entry in new_entries if entry["_index"] not in new_used], _location_key)
    for old_entry in _sorted_entries(entry for entry in old_entries if entry["_index"] not in old_used):
        match = _take_first_unused(new_location_lookup.get(_location_key(old_entry), []), new_used)
        if match is None:
            continue
        old_used.add(old_entry["_index"])
        new_used.add(match["_index"])
        rows.append(_diff_row(STATUS_CNG, old_entry, match))

    for new_entry in _sorted_entries(entry for entry in new_entries if entry["_index"] not in new_used):
        rows.append(_diff_row(STATUS_ADD, None, new_entry))
    for old_entry in _sorted_entries(entry for entry in old_entries if entry["_index"] not in old_used):
        rows.append(_diff_row(STATUS_DEL, old_entry, None))

    type_order = {STATUS_ADD: 1, STATUS_MOVE: 2, STATUS_CNG: 3, STATUS_DEL: 4}
    rows.sort(
        key=lambda row: (
            type_order.get(row["status"], 99),
            natural_sort_key(row.get("new_location") or row.get("old_location")),
            natural_sort_key(row.get("new_part_number") or row.get("old_part_number")),
        )
    )
    return rows


def suggest_export_name(old_path="", new_path=""):
    old_name = _clean_filename_part(Path(old_path or "Program_A").stem) or "Program_A"
    new_name = _clean_filename_part(Path(new_path or "Program_B").stem) or "Program_B"
    return f"Feeder_Compare_{old_name}_vs_{new_name}.xlsx"


def export_feeder_compare_result(result, output_path):
    output = Path(output_path)
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Program A", result.old_file])
    summary_sheet.append(["Program B", result.new_file])
    summary_sheet.append(["Program A Feeders", result.old_count])
    summary_sheet.append(["Program B Feeders", result.new_count])
    summary_sheet.append(["ADD", result.add_count])
    summary_sheet.append(["MOVE", result.move_count])
    summary_sheet.append(["CNG", result.cng_count])
    summary_sheet.append(["DEL", result.del_count])

    result_sheet = workbook.create_sheet("Comparison Results")
    result_sheet.append(["No"] + [header for _, header in RESULT_COLUMNS])
    for index, row in enumerate(result.rows, start=1):
        result_sheet.append([index] + [row.get(key, "") for key, _ in RESULT_COLUMNS])

    _style_summary_sheet(summary_sheet)
    _style_result_sheet(result_sheet)
    workbook.save(output)
    return str(output)


def _entry(record, index):
    entry = dict(record)
    entry["_index"] = index
    entry["_location_key"] = _normalize_key(entry.get("location_code"))
    entry["_part_key"] = _normalize_key(entry.get("part_number"))
    return entry


def _exact_key(entry):
    return (
        _location_key(entry),
        _part_key(entry),
        _normalize_key(entry.get("table")),
        _normalize_key(entry.get("slot")),
        _normalize_key(entry.get("position")),
    )


def _location_key(entry):
    return entry["_location_key"]


def _part_key(entry):
    return entry["_part_key"]


def _group_entries(entries, key_fn):
    grouped = {}
    for entry in entries:
        key = key_fn(entry)
        if not key or key == "":
            continue
        grouped.setdefault(key, []).append(entry)
    for key in grouped:
        grouped[key] = _sorted_entries(grouped[key])
    return grouped


def _take_first_unused(entries, used_indexes):
    for entry in entries or []:
        if entry["_index"] not in used_indexes:
            return entry
    return None


def _sorted_entries(entries):
    return sorted(
        list(entries),
        key=lambda entry: (
            natural_sort_key(entry.get("location_code", "")),
            natural_sort_key(entry.get("part_number", "")),
            entry.get("position", ""),
            entry["_index"],
        ),
    )


def _diff_row(status, old_entry, new_entry):
    row = {
        "status": status,
        "old_table": _value(old_entry, "table"),
        "old_slot": _value(old_entry, "slot"),
        "old_position": _value(old_entry, "position"),
        "old_location": _value(old_entry, "location_code"),
        "old_part_number": _value(old_entry, "part_number"),
        "new_table": _value(new_entry, "table"),
        "new_slot": _value(new_entry, "slot"),
        "new_position": _value(new_entry, "position"),
        "new_location": _value(new_entry, "location_code"),
        "new_part_number": _value(new_entry, "part_number"),
        "_diff_keys": _diff_keys(status),
    }
    row["description"] = _description(row)
    return row


def _description(row):
    status = row["status"]
    if status == STATUS_MOVE:
        return f"Part feeder pindah dari {row['old_location']} ke {row['new_location']}."
    if status == STATUS_CNG:
        if row["old_part_number"] == row["new_part_number"]:
            return (
                f"Location {row['old_location']} detail feeder berubah dari "
                f"{row['old_position']} ke {row['new_position']}."
            )
        return (
            f"Location {row['old_location']} berubah part dari "
            f"{row['old_part_number']} ke {row['new_part_number']}."
        )
    if status == STATUS_ADD:
        return f"Feeder baru ada di Program B pada {row['new_location']}."
    if status == STATUS_DEL:
        return f"Feeder dari Program A pada {row['old_location']} tidak ada di Program B."
    return ""


def _diff_keys(status):
    if status == STATUS_MOVE:
        return ["old_location", "new_location"]
    if status == STATUS_CNG:
        return ["old_part_number", "new_part_number"]
    if status == STATUS_ADD:
        return ["new_location", "new_part_number"]
    if status == STATUS_DEL:
        return ["old_location", "old_part_number"]
    return []


def _value(entry, key):
    if not entry:
        return ""
    return entry.get(key, "")


def _normalize_key(value):
    return str(value or "").strip().upper()


def _style_summary_sheet(worksheet):
    header_font = Font(name="Calibri", size=11, bold=True)
    for row in worksheet.iter_rows():
        row[0].font = header_font
    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 38


def _style_result_sheet(worksheet):
    header_fill = PatternFill("solid", fgColor="FF2B3A4C")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    default_font = Font(name="Calibri", size=11)
    border_side = Side(style="thin", color="FFD3D3D3")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    widths = {
        "A": 6,
        "B": 10,
        "C": 18,
        "D": 22,
        "E": 9,
        "F": 8,
        "G": 18,
        "H": 18,
        "I": 22,
        "J": 9,
        "K": 8,
        "L": 18,
        "M": 48,
    }
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = default_font
            cell.border = border
            horizontal = "left" if cell.column in (4, 9, 13) else "center"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def _clean_filename_part(value):
    text = str(value or "").strip()
    text = re.sub(r'[<>:"/\\|?*]+', "_", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" .")
