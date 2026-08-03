import re
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from services.errors import ServiceError
from services.model_feeder_group_service import _scan_models, _emit_progress, ModelFeederGroupConfig, _build_pair_rows, _build_groups

def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', str(s))]

class VirtualMachine:
    def __init__(self, line_type="Line 1-5"):
        self.line_type = line_type
        self.tables = {}
        if line_type == "Line 1-5":
            for t in range(1, 5):
                self.tables[t] = [{'L': False, 'R': False} for _ in range(17)]
            for t in range(5, 9):
                self.tables[t] = [{'L': False, 'R': False} for _ in range(30)]
            self.tables[9] = [{'L': False, 'R': False} for _ in range(30)]
        elif line_type == "Line 8":
            for t in range(1, 7):
                self.tables[t] = [{'L': False, 'R': False} for _ in range(30)]
            self.tables[7] = [{'L': False, 'R': False} for _ in range(30)]
        else: # Line 6-7
            for t in range(1, 10):
                self.tables[t] = [{'L': False, 'R': False} for _ in range(30)]

    def _parse_loc(self, loc_str):
        match = re.match(r'^\[(\d+)\](\d+)(?:-(\d+))?([LRlr])?$', loc_str)
        if not match:
            return None
        t = int(match.group(1))
        s1 = int(match.group(2))
        s2 = int(match.group(3)) if match.group(3) else s1
        pos = match.group(4).upper() if match.group(4) else None
        return t, s1, s2, pos

    def can_add(self, loc_str):
        parsed = self._parse_loc(loc_str)
        if not parsed:
            return False
        t, s1, s2, pos = parsed
        
        if t not in self.tables:
            return False
            
        if s1 < 1 or s2 > len(self.tables[t]):
            return False
            
        for s in range(s1, s2 + 1):
            idx = s - 1
            if pos == 'L':
                if self.tables[t][idx]['L']: return False
            elif pos == 'R':
                if self.tables[t][idx]['R']: return False
            else:
                if self.tables[t][idx]['L'] or self.tables[t][idx]['R']: return False
                
        return True

    def add(self, loc_str):
        if not self.can_add(loc_str):
            raise ValueError(f'Cannot add {loc_str}')
            
        t, s1, s2, pos = self._parse_loc(loc_str)
        for s in range(s1, s2 + 1):
            idx = s - 1
            if pos == 'L':
                self.tables[t][idx]['L'] = True
            elif pos == 'R':
                self.tables[t][idx]['R'] = True
            else:
                self.tables[t][idx]['L'] = True
                self.tables[t][idx]['R'] = True

    def copy(self):
        import copy
        new_vm = VirtualMachine(self.line_type)
        new_vm.tables = copy.deepcopy(self.tables)
        return new_vm

    def remove(self, loc_str):
        parsed = self._parse_loc(loc_str)
        if not parsed:
            return
        t, s1, s2, pos = parsed
        if t not in self.tables:
            return
        for s in range(s1, s2 + 1):
            idx = s - 1
            if idx >= len(self.tables[t]):
                continue
            if pos == 'L':
                self.tables[t][idx]['L'] = False
            elif pos == 'R':
                self.tables[t][idx]['R'] = False
            else:
                self.tables[t][idx]['L'] = False
                self.tables[t][idx]['R'] = False

    def find_fallback(self, loc_str):
        parsed = self._parse_loc(loc_str)
        if not parsed:
            return None
        pref_t, s1, s2, pos = parsed
        needs_slots = s2 - s1 + 1
        
        if pref_t not in self.tables:
            return None
            
        allowed_tables = [pref_t]
            
        for t in allowed_tables:
            if t not in self.tables:
                continue
                
            num_slots = self.tables[t]
            for start_s in range(1, len(num_slots) - needs_slots + 2):
                end_s = start_s + needs_slots - 1
                idx = start_s - 1
                
                if pos in ['L', 'R'] and needs_slots == 1:
                    # Can fallback to either L or R on this slot
                    if not num_slots[idx]['L']:
                        return f"[{t}]{start_s}L"
                    if not num_slots[idx]['R']:
                        return f"[{t}]{start_s}R"
                else:
                    # Needs full slot(s)
                    can_fit = True
                    for s in range(start_s, end_s + 1):
                        s_idx = s - 1
                        if num_slots[s_idx]['L'] or num_slots[s_idx]['R']:
                            can_fit = False
                            break
                    if can_fit:
                        if needs_slots > 1:
                            return f"[{t}]{start_s}-{end_s}"
                        else:
                            return f"[{t}]{start_s}"
        return None


class PcbInfo:
    def __init__(self, part_number, file_path, components, insert_averages=None, variant_components=None):
        self.part_number = part_number
        self.file_path = file_path
        self.components = components
        self.insert_averages = insert_averages or {}
        self.variant_components = variant_components or {}


class GroupResult:
    def __init__(self, group_name, pcbs, slot_mapping, part_mapping, unassigned_parts=None, substitute_mapping=None, master_excel_path=None):
        self.group_name = group_name
        self.pcbs = pcbs
        self.slot_mapping = slot_mapping
        self.part_mapping = part_mapping
        self.unassigned_parts = unassigned_parts or []
        self.substitute_mapping = substitute_mapping or {}
        self.master_excel_path = master_excel_path

def get_master_mapping(excel_path, line_type=None):
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    try:
        idx_part = headers.index("Part Number")
        idx_loc = headers.index("Feeder Paling Sering")
        idx_freq = headers.index("Total Muncul")
        idx_other = headers.index("Feeder Lain") if "Feeder Lain" in headers else -1
    except ValueError as e:
        raise ServiceError(f"Format Master Mapping salah: {e}")
        
    master = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx_part]:
            continue
        part = str(row[idx_part]).strip().upper()
        loc = str(row[idx_loc]).strip() if row[idx_loc] else ""
        freq = int(row[idx_freq]) if row[idx_freq] else 0
        
        # Ignore Tables
        if "[10]" in loc:
            continue
        if line_type == "Line 8" and ("[8]" in loc or "[9]" in loc):
            continue
            
        other_locs = []
        if idx_other != -1 and row[idx_other]:
            matches = re.findall(r'(\[\d+\]\d+(?:-\d+)?(?:[LRlr])?)', str(row[idx_other]))
            for m in matches:
                if "[10]" in m:
                    continue
                if line_type == "Line 8" and ("[8]" in m or "[9]" in m):
                    continue
                other_locs.append(m)

        if part not in master:
            master[part] = []
        master[part].append({
            "location": loc,
            "frequency": freq,
            "alternatives": other_locs
        })
    return master


def _bipartition_pcbs(raw_group, models):
    if len(raw_group) <= 1:
        return raw_group, []
    if len(raw_group) == 2:
        return [raw_group[0]], [raw_group[1]]

    comp_sets = {}
    for pcb_name in raw_group:
        m = models[pcb_name]
        comp_sets[pcb_name] = set(str(val).strip().upper() for val in m.components.values() if val)

    min_sim = 1.1
    best_pair = (raw_group[0], raw_group[1])
    for i in range(len(raw_group)):
        for j in range(i + 1, len(raw_group)):
            p1, p2 = raw_group[i], raw_group[j]
            s1, s2 = comp_sets[p1], comp_sets[p2]
            inter = len(s1 & s2)
            union = len(s1 | s2)
            sim = inter / max(1, union)
            if sim < min_sim:
                min_sim = sim
                best_pair = (p1, p2)

    seed1, seed2 = best_pair
    sub1 = [seed1]
    sub2 = [seed2]

    for pcb_name in raw_group:
        if pcb_name in (seed1, seed2):
            continue
        p_set = comp_sets[pcb_name]
        overlap1 = sum(len(p_set & comp_sets[s]) for s in sub1) / len(sub1)
        overlap2 = sum(len(p_set & comp_sets[s]) for s in sub2) / len(sub2)
        if overlap1 >= overlap2:
            sub1.append(pcb_name)
        else:
            sub2.append(pcb_name)

    return sub1, sub2


def _process_and_split_group(raw_group, group_label, models, master, global_vm, global_slot_mapping, global_part_mapping, global_unassigned, line_type, cross_group_count=None, master_excel_path=None):
    group_pcbs = []
    for pcb_name in raw_group:
        model = models[pcb_name]
        group_pcbs.append(PcbInfo(
            part_number=model.display_name,
            file_path="",
            components=set(str(val).strip().upper() for val in model.components.values() if val),
            insert_averages=model.insert_averages,
            variant_components=getattr(model, "variant_components", {})
        ))

    variant_usage = {}
    for pcb in group_pcbs:
        if pcb.variant_components:
            for vname, vparts in pcb.variant_components.items():
                vlabel = f"{pcb.part_number} ({vname})" if (vname != pcb.part_number and not vname.startswith(pcb.part_number)) else pcb.part_number
                if len(pcb.variant_components) > 1 and vname not in vlabel:
                    vlabel = f"{pcb.part_number} ({vname})"
                for comp in vparts:
                    comp_u = str(comp).strip().upper()
                    if comp_u not in variant_usage:
                        variant_usage[comp_u] = set()
                    variant_usage[comp_u].add(vlabel)
        else:
            for comp in pcb.components:
                comp_u = str(comp).strip().upper()
                if comp_u not in variant_usage:
                    variant_usage[comp_u] = set()
                variant_usage[comp_u].add(pcb.part_number)

    def part_sort_key(p):
        loc_list = master.get(p, [])
        num_options = 0
        for item in loc_list:
            if item["location"]:
                num_options += 1
            num_options += len(item.get("alternatives", []))
        if num_options == 0:
            num_options = 999
        # cross_group_count: how many top-level groups use this part (cross-group priority)
        cgc = (cross_group_count or {}).get(p, 0)
        v_count = len(variant_usage.get(p, set()))
        master_freq = max([item["frequency"] for item in loc_list], default=0)
        # CGC is the highest priority: parts shared across more groups claim canonical slots first
        return (-cgc, -v_count, num_options, -master_freq)

    group_all_parts = set(variant_usage.keys())
    sorted_parts = sorted(list(group_all_parts), key=part_sort_key)

    vm = global_vm.copy()
    slot_mapping = global_slot_mapping.copy()
    part_mapping = global_part_mapping.copy()
    unassigned_parts = []

    placed_global_parts = set(global_slot_mapping.values())
    parts_to_place = [p for p in sorted_parts if p not in placed_global_parts]

    for part in parts_to_place:
        loc_list = master.get(part, [])
        if not loc_list:
            unassigned_parts.append(part)
            continue
        loc_list = sorted(loc_list, key=lambda x: x["frequency"], reverse=True)
        inserts_in_group = [pcb.insert_averages.get(part, 0) for pcb in group_pcbs]
        avg_inserts = sum(inserts_in_group) / len(group_pcbs) if group_pcbs else 0
        is_balancing = (avg_inserts >= 20) and (len(loc_list) > 1)

        if is_balancing:
            placed_slots = []
            for loc_item in loc_list:
                loc = loc_item["location"]
                if not loc:
                    continue
                if vm.can_add(loc):
                    vm.add(loc)
                    slot_mapping[loc] = part
                    placed_slots.append(loc)
                else:
                    fallback = vm.find_fallback(loc)
                    if fallback:
                        vm.add(fallback)
                        slot_mapping[fallback] = part
                        placed_slots.append(fallback)
            if placed_slots:
                part_mapping[part] = placed_slots[0]
            else:
                unassigned_parts.append(part)
        else:
            primary_loc = loc_list[0]["location"]
            if not primary_loc:
                unassigned_parts.append(part)
                continue
            candidates = [primary_loc]
            for alt in loc_list[0].get("alternatives", []):
                if alt and alt not in candidates:
                    candidates.append(alt)
            placed = False
            for loc in candidates:
                if vm.can_add(loc):
                    vm.add(loc)
                    slot_mapping[loc] = part
                    part_mapping[part] = loc
                    placed = True
                    break
                else:
                    fallback = vm.find_fallback(loc)
                    if fallback:
                        vm.add(fallback)
                        slot_mapping[fallback] = part
                        part_mapping[fallback] = fallback
                        placed = True
                        break
            if not placed:
                unassigned_parts.append(part)

    # Strategy 2: Swap on Conflict
    def _get_valid_locs(p):
        locs = []
        for item in master.get(p, []):
            if item["location"]:
                locs.append(item["location"])
            for alt in item.get("alternatives", []):
                if alt:
                    locs.append(alt)
        seen = set()
        unique = []
        for loc in locs:
            if loc not in seen:
                seen.add(loc)
                unique.append(loc)
        return unique

    swap_resolved = []
    for part in unassigned_parts:
        valid_locs = _get_valid_locs(part)
        if not valid_locs:
            swap_resolved.append(part)
            continue
        swapped = False
        for loc in valid_locs:
            if loc not in slot_mapping and vm.can_add(loc):
                vm.add(loc)
                slot_mapping[loc] = part
                part_mapping[part] = loc
                swapped = True
                break
            if loc not in slot_mapping:
                continue
            occupant = slot_mapping[loc]
            if occupant in placed_global_parts:
                continue
            # Chain Swap Protection: don't displace a component with higher CGC
            cgc_map = cross_group_count or {}
            if cgc_map.get(occupant, 0) > cgc_map.get(part, 0):
                continue
            occupant_valid = _get_valid_locs(occupant)
            for occ_alt in occupant_valid:
                if occ_alt == loc or occ_alt in slot_mapping:
                    continue
                if vm.can_add(occ_alt):
                    vm.remove(loc)
                    vm.add(occ_alt)
                    vm.add(loc)
                    del slot_mapping[loc]
                    slot_mapping[occ_alt] = occupant
                    slot_mapping[loc] = part
                    part_mapping[occupant] = occ_alt
                    part_mapping[part] = loc
                    swapped = True
                    break
            if swapped:
                break
        if not swapped:
            swap_resolved.append(part)

    unassigned_parts = swap_resolved

    # Strategy 3: Substitute Component Slots (Per-Variant & Per-PCB aware)
    variant_usage = {}
    for pcb in group_pcbs:
        if pcb.variant_components:
            for vname, vparts in pcb.variant_components.items():
                vlabel = f"{pcb.part_number} ({vname})" if (vname != pcb.part_number and not vname.startswith(pcb.part_number)) else pcb.part_number
                if len(pcb.variant_components) > 1 and vname not in vlabel:
                    vlabel = f"{pcb.part_number} ({vname})"
                for comp in vparts:
                    comp_u = str(comp).strip().upper()
                    if comp_u not in variant_usage:
                        variant_usage[comp_u] = set()
                    variant_usage[comp_u].add(vlabel)
        else:
            for comp in pcb.components:
                comp_u = str(comp).strip().upper()
                if comp_u not in variant_usage:
                    variant_usage[comp_u] = set()
                variant_usage[comp_u].add(pcb.part_number)

    substitute_mapping = {}
    slot_sub_vars = {}
    final_unassigned = []

    for part in unassigned_parts:
        part_vars = variant_usage.get(part, set())
        if not part_vars:
            final_unassigned.append(part)
            continue
        valid_locs = _get_valid_locs(part)
        if not valid_locs:
            final_unassigned.append(part)
            continue
        valid_tables = set()
        part_spans = set()
        part_positions = set()
        for vloc in valid_locs:
            parsed_vloc = vm._parse_loc(vloc)
            if parsed_vloc:
                valid_tables.add(parsed_vloc[0])
                part_spans.add(parsed_vloc[2] - parsed_vloc[1] + 1)
                part_positions.add(parsed_vloc[3])  # 'L', 'R', or None
        if not valid_tables:
            final_unassigned.append(part)
            continue

        # Prefer candidate slots that are in valid_locs of this part first, then fall back to other compatible slots
        preferred_slots = [loc for loc in valid_locs if loc in slot_mapping]
        other_slots = [loc for loc in slot_mapping if loc not in preferred_slots]
        candidate_slots = preferred_slots + other_slots

        best_slot = None
        for loc in candidate_slots:
            occupant = slot_mapping[loc]
            parsed_loc = vm._parse_loc(loc)
            if not parsed_loc or parsed_loc[0] not in valid_tables:
                continue
            loc_span = parsed_loc[2] - parsed_loc[1] + 1
            if part_spans and loc_span not in part_spans:
                continue
            loc_pos = parsed_loc[3]
            if part_positions and loc_pos not in part_positions:
                continue
            occupant_vars = variant_usage.get(occupant, set())
            existing_sub_vars = slot_sub_vars.get(loc, set())
            blocked_vars = occupant_vars | existing_sub_vars
            if not (part_vars & blocked_vars):
                best_slot = loc
                break
        if best_slot:
            if best_slot not in substitute_mapping:
                substitute_mapping[best_slot] = []
            var_names = sorted(list(part_vars))
            substitute_mapping[best_slot].append((part, var_names))
            if best_slot not in slot_sub_vars:
                slot_sub_vars[best_slot] = set()
            slot_sub_vars[best_slot].update(part_vars)
        else:
            final_unassigned.append(part)

    unassigned_parts = final_unassigned

    # AUTO-SPLIT CHECK (splits multi-PCB groups ONLY if physical capacity is exceeded for master parts):
    overload_unassigned = [p for p in unassigned_parts if master.get(p)]

    if overload_unassigned and len(raw_group) > 1:
        sub1, sub2 = _bipartition_pcbs(raw_group, models)

        # Determine components shared between sub1 and sub2
        sub1_parts = set()
        for pcb_name in sub1:
            m = models[pcb_name]
            sub1_parts.update(set(str(val).strip().upper() for val in m.components.values() if val))

        sub2_parts = set()
        for pcb_name in sub2:
            m = models[pcb_name]
            sub2_parts.update(set(str(val).strip().upper() for val in m.components.values() if val))

        shared_parent_parts = (sub1_parts & sub2_parts) - set(global_slot_mapping.values())

        # Build parent base state by locking shared_parent_parts into parent_vm
        parent_vm = global_vm.copy()
        parent_slot_mapping = global_slot_mapping.copy()
        parent_part_mapping = global_part_mapping.copy()
        parent_unassigned = list(global_unassigned)

        def _parent_sort_key(p):
            loc_list = master.get(p, [])
            num_options = 0
            for item in loc_list:
                if item["location"]:
                    num_options += 1
                num_options += len(item.get("alternatives", []))
            if num_options == 0:
                num_options = 999
            # Use cross_group_count as highest priority — same as global sort logic
            cgc = (cross_group_count or {}).get(p, 0)
            master_freq = max([item["frequency"] for item in loc_list], default=0)
            return (-cgc, num_options, -master_freq)

        sorted_parent_shared = sorted(list(shared_parent_parts), key=_parent_sort_key)

        for part in sorted_parent_shared:
            if part in parent_slot_mapping.values():
                continue
            loc_list = master.get(part, [])
            if not loc_list:
                parent_unassigned.append(part)
                continue
            loc_list = sorted(loc_list, key=lambda x: x["frequency"], reverse=True)

            primary_loc = loc_list[0]["location"]
            if not primary_loc:
                parent_unassigned.append(part)
                continue
            candidates = [primary_loc]
            for alt in loc_list[0].get("alternatives", []):
                if alt and alt not in candidates:
                    candidates.append(alt)
            placed = False
            for loc in candidates:
                if parent_vm.can_add(loc):
                    parent_vm.add(loc)
                    parent_slot_mapping[loc] = part
                    parent_part_mapping[part] = loc
                    placed = True
                    break
                else:
                    fallback = parent_vm.find_fallback(loc)
                    if fallback:
                        parent_vm.add(fallback)
                        parent_slot_mapping[fallback] = part
                        parent_part_mapping[part] = fallback
                        placed = True
                        break
            if not placed:
                parent_unassigned.append(part)

        label1 = f"{group_label}A" if not group_label[-1].isalpha() else f"{group_label}-1"
        label2 = f"{group_label}B" if not group_label[-1].isalpha() else f"{group_label}-2"
        res1 = _process_and_split_group(sub1, label1, models, master, parent_vm, parent_slot_mapping, parent_part_mapping, parent_unassigned, line_type, cross_group_count, master_excel_path)
        res2 = _process_and_split_group(sub2, label2, models, master, parent_vm, parent_slot_mapping, parent_part_mapping, parent_unassigned, line_type, cross_group_count, master_excel_path)
        return res1 + res2

    return [GroupResult(
        group_name=group_label,
        pcbs=group_pcbs,
        slot_mapping=slot_mapping,
        part_mapping=part_mapping,
        unassigned_parts=unassigned_parts,
        substitute_mapping=substitute_mapping,
        master_excel_path=master_excel_path,
    )]


def generate_all_table_groups(crb_folder, master_excel_path, target_pcbs_text, line_type, min_sim, min_shared, progress_callback=None, base_npm_path=None):
    _emit_progress(progress_callback, 0, "Membaca referensi & Master Mapping...")
    master = get_master_mapping(master_excel_path, line_type)
    
    # Strip L/R suffixes for tables that are for large/full-slot components
    new_master = {}
    for part, loc_list in master.items():
        new_items = []
        for item in loc_list:
            loc = item["location"]
            if not loc:
                continue
            
            if line_type == "Line 8":
                # For Line 8, Table 7 is for large components (no L/R, 2-3 slots)
                if "[7]" in loc:
                    match = re.match(r'^(\[7\]\d+(?:-\d+)?)[LRlr]?$', loc)
                    if match:
                        loc = match.group(1)
            else:
                # For Line 1-5 and Line 6-7, Table 9 is for large components (no L/R)
                if "[9]" in loc:
                    match = re.match(r'^(\[9\]\d+(?:-\d+)?)[LRlr]?$', loc)
                    if match:
                        loc = match.group(1)
            
            # Clean alternatives similarly
            new_alts = []
            for alt in item.get("alternatives", []):
                if line_type == "Line 8":
                    if "[7]" in alt:
                        match = re.match(r'^(\[7\]\d+(?:-\d+)?)[LRlr]?$', alt)
                        if match:
                            alt = match.group(1)
                else:
                    if "[9]" in alt:
                        match = re.match(r'^(\[9\]\d+(?:-\d+)?)[LRlr]?$', alt)
                        if match:
                            alt = match.group(1)
                new_alts.append(alt)
            
            new_item = dict(item)
            new_item["location"] = loc
            new_item["alternatives"] = new_alts
            new_items.append(new_item)
        if new_items:
            new_master[part] = new_items
    master = new_master
    
    target_list = []
    if target_pcbs_text:
        for line in target_pcbs_text.splitlines():
            line = line.strip().upper()
            if line:
                target_list.append(line)

    _emit_progress(progress_callback, 10, "Scanning PCB folders...")
    models, _, _, _ = _scan_models(crb_folder, target_list, progress_callback)
    if not models:
        raise ServiceError("Tidak ada file Excel program yang valid ditemukan atau sesuai dengan target PCB.")

    # Step 1: Group PCBs based on similarity
    _emit_progress(progress_callback, 40, "Mengelompokkan PCB berdasarkan kemiripan...")
    config = ModelFeederGroupConfig(
        source_folder=crb_folder,
        min_similarity_percent=min_sim,
        min_shared_components=min_shared,
        target_pcb_list=target_list
    )
    pair_rows, pair_lookup = _build_pair_rows(models, config)
    raw_groups = _build_groups(models, pair_rows, pair_lookup)
    
    groups = []
    _emit_progress(progress_callback, 60, "Membangun setup feeder untuk masing-masing grup...")
    
    total_groups = len(raw_groups)

    # Phase 1: Determine Global Base Components (parts used across ALL groups)
    # Also compute cross_group_count: how many top-level groups use each part
    group_parts_sets = []
    cross_group_count = {}  # {part: number of top-level groups that use this part}
    for raw_group in raw_groups:
        parts_in_g = set()
        for pcb_name in raw_group:
            m = models[pcb_name]
            parts_in_g.update(set(str(val).strip().upper() for val in m.components.values() if val))
        group_parts_sets.append(parts_in_g)
        for part in parts_in_g:
            cross_group_count[part] = cross_group_count.get(part, 0) + 1
        
    # global_base_parts: all parts that appear in 2+ top-level groups → lock to canonical slot globally
    global_base_parts = set()
    if len(raw_groups) > 1:
        global_base_parts = {part for part, count in cross_group_count.items() if count >= 2}
    elif len(group_parts_sets) == 1:
        all_pcb_sets = []
        for pcb_name in raw_groups[0]:
            m = models[pcb_name]
            all_pcb_sets.append(set(str(val).strip().upper() for val in m.components.values() if val))
        if all_pcb_sets:
            global_base_parts = set.intersection(*all_pcb_sets)

    # Phase 2: Lock Global Base Components to identical slots across all groups
    global_vm = VirtualMachine(line_type=line_type)
    global_slot_mapping = {}
    global_part_mapping = {}
    global_unassigned = []

    if base_npm_path and Path(base_npm_path).is_file():
        import services.feeder_mapping_service as fms
        base_mapping_res = fms.load_feeder_mapping(base_npm_path)
        for r in base_mapping_res.records:
            loc = r['location_code']
            part = r['part_number']
            if global_vm.can_add(loc):
                global_vm.add(loc)
                global_slot_mapping[loc] = part
                global_part_mapping[part] = loc

    def _global_sort_key(p):
        loc_list = master.get(p, [])
        num_options = 0
        for item in loc_list:
            if item["location"]:
                num_options += 1
            num_options += len(item.get("alternatives", []))
        if num_options == 0:
            num_options = 999
        # CGC first: lock highest-cross-group components to canonical slots before others
        cgc = cross_group_count.get(p, 0)
        master_freq = max([item["frequency"] for item in loc_list], default=0)
        return (-cgc, num_options, -master_freq)

    sorted_global_base = sorted(list(global_base_parts), key=_global_sort_key)

    # Compute global average inserts across all models for balancing check
    all_model_count = max(1, len(models))

    for part in sorted_global_base:
        if part in global_part_mapping:
            continue
            
        loc_list = master.get(part, [])
        if not loc_list:
            global_unassigned.append(part)
            continue
        loc_list = sorted(loc_list, key=lambda x: x["frequency"], reverse=True)

        # Check global average inserts for balancing
        inserts_total = sum(model.insert_averages.get(part, 0) for model in models.values())
        avg_inserts = inserts_total / all_model_count
        is_balancing = (avg_inserts >= 20) and (len(loc_list) > 1)

        if is_balancing:
            placed_slots = []
            for loc_item in loc_list:
                loc = loc_item["location"]
                if not loc:
                    continue
                if global_vm.can_add(loc):
                    global_vm.add(loc)
                    global_slot_mapping[loc] = part
                    placed_slots.append(loc)
                # No fallback: only canonical/listed slots (no drifting to adjacent slots)
            if placed_slots:
                global_part_mapping[part] = placed_slots[0]
            else:
                global_unassigned.append(part)
        else:
            primary_loc = loc_list[0]["location"]
            if not primary_loc:
                global_unassigned.append(part)
                continue
            candidates = [primary_loc]
            for alt in loc_list[0].get("alternatives", []):
                if alt and alt not in candidates:
                    candidates.append(alt)
            placed = False
            for loc in candidates:
                if global_vm.can_add(loc):
                    global_vm.add(loc)
                    global_slot_mapping[loc] = part
                    global_part_mapping[part] = loc
                    placed = True
                    break
                # No fallback: if canonical and alternatives are taken, fall to local processing
            if not placed:
                global_unassigned.append(part)
    
    for i, raw_group in enumerate(raw_groups):
        label = f"Group {i + 1}"
        sub_results = _process_and_split_group(raw_group, label, models, master, global_vm, global_slot_mapping, global_part_mapping, global_unassigned, line_type, cross_group_count, master_excel_path)
        groups.extend(sub_results)
        
        percent = 60 + int((i + 1) / max(1, total_groups) * 35)
        _emit_progress(progress_callback, percent, f"Setup Group {i + 1} selesai...")
        
    _emit_progress(progress_callback, 100, "Selesai")
    return groups

def _load_raw_master_locs(excel_path):
    if not excel_path or not Path(excel_path).is_file():
        return {}
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        if "Part Number" not in headers or "Feeder Paling Sering" not in headers:
            wb.close()
            return {}
        idx_part = headers.index("Part Number")
        idx_loc = headers.index("Feeder Paling Sering")
        raw_locs = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(idx_part, idx_loc) or not row[idx_part]:
                continue
            part = str(row[idx_part]).strip().upper()
            loc = str(row[idx_loc]).strip() if row[idx_loc] else ""
            if part not in raw_locs:
                raw_locs[part] = loc
        wb.close()
        return raw_locs
    except Exception:
        return {}


def export_all_table_groups(groups, output_path, master_excel_path=None):
    wb = Workbook()
    
    if not master_excel_path and groups and getattr(groups[0], "master_excel_path", None):
        master_excel_path = groups[0].master_excel_path

    raw_master_locs = _load_raw_master_locs(master_excel_path)

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Group Name", "PCB Name", "Total PCBs in Group", "Total Fixed Slots", "Total Substitute Components", "Total Skipped Components"])
    
    link_font = Font(name="Calibri", size=11, color="0002D9", underline="single")

    for g in groups:
        # Filter unassigned parts to exclude Table [10] (TRAY) components
        display_unassigned = []
        for p in g.unassigned_parts:
            if p in raw_master_locs:
                loc = raw_master_locs[p]
                if "[10]" in loc or loc.startswith("[10]"):
                    continue  # Skip Table 10 (TRAY) components completely
                display_unassigned.append((p, loc if loc else "#N/A"))
            else:
                display_unassigned.append((p, "#N/A"))

        g._display_unassigned = display_unassigned

        for p in g.pcbs:
            total_subs = sum(len(subs) for subs in g.substitute_mapping.values())
            row = [g.group_name, p.part_number, len(g.pcbs), len(g.slot_mapping), total_subs, len(display_unassigned)]
            ws_summary.append(row)
            summary_row_idx = ws_summary.max_row
            
            # Make Group Name clickable hyperlink to its worksheet
            cell_group = ws_summary.cell(row=summary_row_idx, column=1)
            cell_group.hyperlink = f"#'{g.group_name}'!A1"
            cell_group.font = link_font
        
    # Calculate base components used in ALL groups
    group_parts = []
    for g in groups:
        parts_in_group = set(g.slot_mapping.values())
        group_parts.append(parts_in_group)
        
    common_parts = set()
    if group_parts:
        common_parts = group_parts[0]
        for s in group_parts[1:]:
            common_parts = common_parts & s
            
    common_parts_list = sorted(list(common_parts))
    
    # Style and write base components
    ws_summary.append([])
    ws_summary.append([])
    
    title_font = Font(name="Calibri", size=12, bold=True, color="1F4E78")
    ws_summary.append(["Base Components (Kerangka Dasar - Terpakai di Semua Group)"])
    title_row_idx = ws_summary.max_row
    ws_summary.cell(row=title_row_idx, column=1).font = title_font
    
    base_headers = ["Base Component P/N"] + [f"{g.group_name} Slot" for g in groups]
    ws_summary.append(base_headers)
    header_row_idx = ws_summary.max_row
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    for col_idx in range(1, len(base_headers) + 1):
        cell = ws_summary.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align if col_idx > 1 else left_align
        
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    body_font = Font(name="Calibri", size=11)
    
    if common_parts_list:
        for part in common_parts_list:
            row = [part]
            for g in groups:
                slots = [loc for loc, p in g.slot_mapping.items() if p == part]
                row.append(", ".join(slots))
            ws_summary.append(row)
            
            data_row_idx = ws_summary.max_row
            for col_idx in range(1, len(row) + 1):
                cell = ws_summary.cell(row=data_row_idx, column=col_idx)
                cell.font = body_font
                cell.alignment = center_align if col_idx > 1 else left_align
                cell.border = thin_border
    else:
        ws_summary.append(["Tidak ada komponen yang terpakai di semua group."])
        data_row_idx = ws_summary.max_row
        cell = ws_summary.cell(row=data_row_idx, column=1)
        cell.font = body_font
        cell.border = thin_border
        
    # Auto-adjust column widths on summary sheet
    for col in ws_summary.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or "")
            if cell.row == title_row_idx:
                continue
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    for g in groups:
        ws = wb.create_sheet(title=g.group_name)

        # Row 1: Return to Summary Hyperlink Button
        cell_back = ws.cell(row=1, column=1, value="[ 🔙 Kembali ke Summary ]")
        cell_back.hyperlink = "#'Summary'!A1"
        cell_back.font = Font(name="Calibri", size=11, bold=True, color="0002D9", underline="single")

        # Row 2: Headers
        ws.cell(row=2, column=1, value="Table")
        ws.cell(row=2, column=2, value="Slot")
        ws.cell(row=2, column=3, value="Position")
        ws.cell(row=2, column=4, value="Location Code")
        ws.cell(row=2, column=5, value="Part Number")
        ws.cell(row=2, column=6, value="Type")
        ws.cell(row=2, column=7, value="Active When")

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for col_idx in range(1, 8):
            c = ws.cell(row=2, column=col_idx)
            c.font = header_font
            c.fill = header_fill
        
        records = []
        for loc, part in g.slot_mapping.items():
            match = re.match(r"^\[(\d+)\](\d+)(?:-\d+)?([A-Za-z]+)?$", loc)
            if match:
                table = int(match.group(1))
                slot = int(match.group(2))
                pos = str(match.group(3) or "").upper()
            else:
                table = 99
                slot = 99
                pos = ""
            records.append({
                "table": table,
                "slot": slot,
                "pos": pos,
                "loc": loc,
                "part": part,
                "type": "FIXED",
                "active_when": ""
            })
            if loc in g.substitute_mapping:
                for sub_part, sub_pcbs in g.substitute_mapping[loc]:
                    records.append({
                        "table": table,
                        "slot": slot,
                        "pos": pos,
                        "loc": loc,
                        "part": sub_part,
                        "type": "SUBSTITUTE",
                        "active_when": ", ".join(sub_pcbs)
                    })
                
        records.sort(key=lambda x: (x["table"], x["slot"], x["pos"], natural_sort_key(x["loc"]), x["type"] != "FIXED"))
        
        for r in records:
            ws.append([r["table"], r["slot"], r["pos"], r["loc"], r["part"], r["type"], r["active_when"]])
            
        # Write truly unresolvable skipped parts at the bottom (excluding TRAY components)
        display_unassigned = getattr(g, "_display_unassigned", None)
        if display_unassigned is None:
            display_unassigned = []
            for p in g.unassigned_parts:
                if p in raw_master_locs:
                    loc = raw_master_locs[p]
                    if "[10]" in loc or loc.startswith("[10]"):
                        continue
                    display_unassigned.append((p, loc if loc else "#N/A"))
                else:
                    display_unassigned.append((p, "#N/A"))

        if display_unassigned:
            ws.append([])
            ws.append(["SKIPPED / TRULY UNRESOLVABLE COMPONENTS:"])
            ws.append(["Part Number", "Feeder Paling Sering"])
            for p, loc in display_unassigned:
                ws.append([p, loc])
            
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 35
        
    wb.save(output_path)
    return output_path
